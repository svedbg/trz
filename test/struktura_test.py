# -*- coding: utf-8 -*-
"""Проверки върху генерирана ведомост от втория комплект.

    python test/struktura_test.py --seed 7            # генерира и проверява
    python test/struktura_test.py --seed 7 --tiho     # само резултатът

Тестът е кръгов: `generate_struktura.py` строи коректна ведомост и я чупи по
известен списък, а този файл трябва да намери точно счупеното — нито по-малко
(пропусната находка), нито повече (фалшиво положително). Фалшивите положителни
се броят наравно с пропуските: скил, който вижда нарушения навсякъде, е толкова
безполезен, колкото и скил, който не вижда никакви.

Какво НЕ прави тестът: не решава спорните въпроси. Дали доходът в натура влиза в
осигурителния доход има повече от едно защитимо четене (proverki.md, F10).
Затова проверката е за **последователност** — политиката на файла се извежда от
самия файл и се търсят редовете, които се отклоняват от нея. Същото важи и за
превишението над необлагаемия праг.

Методът за F1/F6/F10 е „решаване на състава": търси се кое подмножество от
начисленията и придобивките обяснява обявения осигурителен доход и обявената
данъчна основа. Ако един и същ елемент е вътре в едната база и вън от другата,
поне едно от двете е грешно — независимо от тълкуването. Точно това е находката,
която не изисква произнасяне по спорния въпрос.
"""
import argparse
import json
import os
import sys
from collections import Counter

import openpyxl

import trz_model as M
from trz_model import r2

HERE = os.path.dirname(os.path.abspath(__file__))


def _podmnozhestva(elementi):
    """Всички подмножества на [(име, стойност)] като [(frozenset, сума)]."""
    out = [(frozenset(), 0.0)]
    for ime, v in elementi:
        if not v:
            continue
        out += [(frozenset(m | {ime}), r2(s + v)) for m, s in out]
    return out


class Nahodki:
    def __init__(self):
        self.spisak = []
        self._vidyani = set()

    def add(self, ident, kade, opisanie, nach=None, dalj=None):
        if (kade, ident) in self._vidyani:
            return                     # една находка на (ред, вид), не по няколко
        self._vidyani.add((kade, ident))
        self.spisak.append(dict(id=ident, kade=kade, opisanie=opisanie,
                                nachisleno=nach, dalzhimo=dalj))

    def kljuchove(self):
        return {(n["kade"], n["id"]) for n in self.spisak}


def proveri(xlsx, manifest, tiho=False):
    man = json.load(open(manifest, encoding="utf8"))
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[man["list"]]
    HDR, TOT = man["hdr"], man["red_obshto"]
    norma = man["norma_dni"]
    max_osig = man["max_osig_prilozhim"]
    mod_samoosig = man["mod_samoosig"]
    tzpb_dalzhim = man["tzpb_dalzhim"]
    drug_tavan = [v["max_osig"] for v in M.REZHIMI.values() if v["max_osig"] != max_osig][0]

    kol = {ws.cell(row=HDR, column=c).value: c for c in range(1, ws.max_column + 1)}

    def g(r, ime):
        v = ws.cell(row=r, column=kol[ime]).value
        return 0.0 if v in (None, "") else v

    redove = list(range(HDR + 1, TOT))
    dogovori = {l["red"]: l["vhod"] for l in man["lica"]}
    N = Nahodki()
    rd = []

    # =================================================== пас 1: ред по ред
    for r in redove:
        v = {k: g(r, k) for k in M.KOLONI}
        wd, pl, sd, md = (v[k] for k in M.KOLONI_DNI)
        osnovna = v["Основна за отработеното"]
        klas_pr, klas = v["Клас %"], v["Клас сума"]
        bonus, otpusk = v["Бонус"], v["Платен отпуск"]
        ob224, boln = v["Обезщетение чл. 224"], v["Болнични (работодател)"]
        bruto, lichni = v["БРУТО"], v["Лични вноски общо"]
        osig, dan_osnova, danak = v["Осигурителен доход"], v["Данъчна основа"], v["ДДФЛ"]
        udr_osig, udr_karta = v["Удръжка доброволно осиг. (лична)"], v["Удръжка карта (лична част)"]
        neto_predi, neto = v["НЕТО преди удръжки"], v["НЕТО за изплащане"]
        platено, razlika = v["Изплатено"], v["Разлика"]
        er_doo, er_zo_boln = v["Вноски работодател ДОО+ТЗПБ"], v["ЗО при болничен/майчинство"]
        er_obshto = v["Вноски работодател общо"]
        natura = v["Карта (за сметка на работодателя)"]
        premia = v["Доброволно здравно осигуряване (премия)"]
        razhod = v["Общ разход за труд"]
        prev = r2(max(0.0, premia - M.PRAG_SOTSIALNI_RAZHODI)) if premia else 0.0

        # --- K2: сума, въведена в колона за дни -------------------------
        for kd in M.KOLONI_DNI:
            x = v[kd]
            if x and (abs(x - round(x)) > 1e-9 or x > norma):
                N.add("K2_suma_v_kolona_dni", r,
                      f"колона за дни съдържа {x} при норма {norma} — това е сума, не дни",
                      x, None)
                break

        # --- I5: сборът на дните ---------------------------------------
        sbor_dni = wd + pl + sd + md
        dni_ok = abs(sbor_dni - norma) < 1e-9
        if not dni_ok:
            N.add("I5_dni_ne_se_vrazvat", r,
                  f"дни {wd}+{pl}+{sd}+{md} = {sbor_dni} при норма {norma} работни дни",
                  sbor_dni, norma)

        # --- K6: незакръглени начисления --------------------------------
        for kn in M.NACHISLENIYA + ["БРУТО"]:
            if v[kn] and abs(v[kn] - r2(v[kn])) > 1e-9:
                N.add("K6_nezakraglen", r,
                      f"начисление с повече от два знака: {v[kn]!r}", v[kn], r2(v[kn]))
                break

        # --- K1: брутото = сборът на начисленията -----------------------
        sbor_n = r2(sum(v[k] for k in M.NACHISLENIYA))
        if abs(bruto - sbor_n) > M.TOL:
            lipsva = [k for k in M.NACHISLENIYA
                      if v[k] and abs(r2(sbor_n - v[k]) - bruto) <= M.TOL]
            N.add("K1_sbor_izpuska_kolona", r,
                  "БРУТО не е сборът на начисленията"
                  + (f"; извън сбора остава колоната за {lipsva[0]}" if lipsva else ""),
                  bruto, sbor_n)

        # --- K3: вноските като процент от осигурителния доход -----------
        loshi = [(kv, v[kv], r2(osig * M.LICHNI[kl] / 100.0))
                 for kv, kl in (("ДОО пенсии", "pensii"), ("ДОО ОЗМ", "ozm"),
                                ("ДОО безработица", "bezrab"), ("ЗО лична", "zo"),
                                ("ДЗПО-УПФ лична", "upf"))
                 if abs(v[kv] - r2(osig * M.LICHNI[kl] / 100.0)) > M.TOL]
        if loshi:
            N.add("K3_tvardi_stoynosti", r,
                  f"вноските не са процент от осигурителния доход {osig:.2f} "
                  f"(първо разминаване: {loshi[0][0]})", loshi[0][1], loshi[0][2])
        else:
            # Сборът на петте отделно закръглени вноски може да се разминава с
            # 13.78% от осигурителния доход с до 0.03 — по 0.005 на компонент
            # плюс закръглянето на самата отправна стойност. Точната контрола е
            # сборът на компонентите; процентът е ориентировъчен.
            sbor_komp = r2(sum(v[k] for k in ("ДОО пенсии", "ДОО ОЗМ", "ДОО безработица",
                                              "ЗО лична", "ДЗПО-УПФ лична")))
            if abs(lichni - sbor_komp) > M.TOL_STROG:
                N.add("K3_tvardi_stoynosti", r,
                      "„Лични вноски общо“ ≠ сборът на петте вноски", lichni, sbor_komp)
            elif abs(lichni - r2(osig * M.LICHNI_OBSHTO / 100.0)) > 0.03:
                N.add("K3_tvardi_stoynosti", r,
                      f"личните вноски общо ≠ {M.LICHNI_OBSHTO}% от {osig:.2f}",
                      lichni, r2(osig * M.LICHNI_OBSHTO / 100.0))

        # --- I1 / F6: вертикална сверка ---------------------------------
        if abs(neto_predi - r2(bruto - lichni - danak)) > M.TOL:
            N.add("I1_vertikalna", r, "НЕТО преди удръжки ≠ бруто − вноски − данък",
                  neto_predi, r2(bruto - lichni - danak))
        if abs(neto - r2(neto_predi - udr_osig - udr_karta)) > M.TOL:
            N.add("I1_vertikalna", r, "НЕТО за изплащане ≠ нето преди удръжки − удръжки",
                  neto, r2(neto_predi - udr_osig - udr_karta))
        if abs(danak - r2(dan_osnova * M.DDFL)) > M.TOL:
            N.add("F6_danak", r, "данъкът ≠ 10% от данъчната основа",
                  danak, r2(dan_osnova * M.DDFL))

        # --- K4: контролната колона -------------------------------------
        if abs(razlika - r2(neto - platено)) > M.TOL_STROG:
            N.add("K4_kontrola_ne_hvashta", r,
                  f"колона „Разлика“ показва {razlika:.2f}, а нето − изплатено = "
                  f"{r2(neto - platено):.2f}; контролата не отчита разликата",
                  razlika, r2(neto - platено))

        # --- K7: общият разход ------------------------------------------
        och_razhod = r2(bruto + er_obshto + natura + premia)
        if abs(razhod - och_razhod) > M.TOL:
            udr = r2(udr_osig + udr_karta)
            zashto = " — занижен точно с удържаното от работника" \
                if udr and abs(r2(och_razhod - udr) - razhod) <= M.TOL else ""
            N.add("K7_razhod_ot_neto", r,
                  "общият разход ≠ бруто + вноски работодател + придобивки" + zashto,
                  razhod, och_razhod)

        # --- проверки, зависещи от дните и от договора -------------------
        if dni_ok and r in dogovori:
            d = dogovori[r]
            dneven = d["mesechna_zaplata"] / norma
            s_klas = 1 + klas_pr / 100.0
            if abs(osnovna - r2(dneven * wd)) > M.TOL:
                N.add("A6_osnovna_vs_dogovor", r,
                      "основната за отработеното не отговаря на договорената заплата",
                      osnovna, r2(dneven * wd))
            if pl:
                och = r2(dneven * s_klas * pl)
                if abs(otpusk - och) > M.TOL:
                    bez = r2(dneven * pl)
                    zashto = " — без включен клас" if abs(otpusk - bez) <= M.TOL else ""
                    N.add("E3_otpusk_bez_klas", r,
                          f"платеният отпуск за {int(pl)} дни{zashto}", otpusk, och)
            if sd:
                dni_er = min(sd, M.BOLN_DNI_RABOTODATEL)
                och = r2(dneven * s_klas * dni_er * M.BOLN_PROCENT)
                if abs(boln - och) > M.TOL:
                    N.add("F9_bolnichen_razmer", r,
                          f"обезщетението по чл. 40, ал. 5 КСО за {int(dni_er)} дни",
                          boln, och)
            if sd + md:
                och = r2(mod_samoosig * M.ZO_PRI_NERABOTOSPOSOBNOST / 100.0 * (sd + md) / norma)
                if abs(er_zo_boln - och) > M.TOL:
                    N.add("F9_bez_zo_bolnichen", r,
                          f"ЗО за {int(sd + md)} дни неработоспособност/майчинство "
                          f"(чл. 40, ал. 1, т. 5 ЗЗО)", er_zo_boln, och)

        # --- C2: база на класа ------------------------------------------
        if klas_pr:
            och = r2(osnovna * klas_pr / 100.0)
            if abs(klas - och) > M.TOL:
                shiroka = r2((osnovna + otpusk + bonus) * klas_pr / 100.0)
                zashto = " — начислен върху по-широка база от основната заплата" \
                    if abs(klas - shiroka) <= M.TOL else ""
                N.add("C2_klas_varhu_bruto", r,
                      f"класът {klas_pr}% не съответства на основната заплата{zashto}",
                      klas, och)

        # --- база за труд и „кой елемент е вътре" -------------------------
        # Съставът на осигурителния доход се решава на два прохода: първо се
        # изважда политиката на файла от редовете, при които базата не стига
        # тавана, а после всеки ред се мери спрямо нея. Само с изброяване на
        # подмножества капнатите редове са неразличими — там сумата е една и
        # съща за много различни комбинации.
        baza = r2(osnovna + klas + bonus + otpusk)
        elementi = dict(natura=natura, prev=prev, boln=boln, ob224=ob224)
        # Съставът е решим точно, когато редът НЕ стои на таван: тогава
        # осигурителният доход е самата сума и разликата спрямо базата за труд
        # сочи елемента. Стои ли на таван — сумата е една и съща за много
        # различни комбинации и проверката е длъжна да замълчи.
        na_tavan = any(abs(osig - c) <= M.TOL for c in (max_osig, drug_tavan))
        # Лице без начисления за труд през месеца (цял месец в майчинство или
        # неплатен отпуск): придобивките сами по себе си не създават осигурителен
        # доход, защото няма доход от трудова дейност, върху който да се начислят
        # (чл. 6, ал. 2 КСО). За такъв ред съставът не се анализира.
        bez_trud = baza <= 0
        vatre_ednoznachno = None
        if not na_tavan and not bez_trud:
            sred = [maska for maska, s in _podmnozhestva([("natura", natura), ("prev", prev)])
                    if abs(r2(baza + s) - osig) <= M.TOL]
            if len(sred) == 1:
                vatre_ednoznachno = sred[0]

        rd.append(dict(r=r, baza=baza, elementi=elementi, bruto=bruto, lichni=lichni,
                       vatre_ednoznachno=vatre_ednoznachno, na_tavan=na_tavan,
                       bez_trud=bez_trud,
                       osig=osig, boln=boln, natura=natura, prev=prev,
                       dan_osnova=dan_osnova, udr_osig=udr_osig, er_doo=er_doo))

    # ------------------------------- практиката на файла за двете придобивки
    # Дали доходът в натура и превишението над необлагаемия праг влизат в базите
    # е спорен въпрос (proverki.md, F10). Затова не се приема отговор по памет, а
    # се изважда от самия файл — но само ако извадката го позволява: поне три
    # използваеми реда и ясно мнозинство. Ред „използваем" значи ред, който не
    # стои на таван (там съставът е неразличим) и има начисления за труд.
    # Не се ли установи практика, за този елемент не се произнасяме — точно
    # както жив проверяващ би попитал, вместо да гадае.
    def praktika_za(el):
        izvadka = [el in x["vatre_ednoznachno"] for x in rd
                   if x["vatre_ednoznachno"] is not None and x["elementi"][el]]
        if len(izvadka) < 3:
            return None, len(izvadka)
        c = Counter(izvadka)
        stoynost, broy = c.most_common(1)[0]
        if broy / len(izvadka) < 2 / 3:
            return None, len(izvadka)
        return stoynost, len(izvadka)

    pol, pol_izvadka = {}, {}
    for el in ("natura", "prev"):
        pol[el], pol_izvadka[el] = praktika_za(el)
        if pol[el] is None and any(x["elementi"][el] for x in rd):
            N.add("F10_praktika_neustanovima", "fayl",
                  f"практиката на файла за елемента {el} не се установява от него самия "
                  f"({pol_izvadka[el]} използваеми реда) — съставът на базите за тези "
                  f"лица не подлежи на извод без вътрешните правила")

    # ============================== пас 2: състав, симетрия, файлови находки
    IMENA_EL = dict(natura="доходът в натура", prev="превишението над необлагаемия праг",
                    boln="обезщетението по чл. 40, ал. 5 КСО",
                    ob224="обезщетението по чл. 224 КТ")
    ID_DOBAVEN = dict(boln="F9_bolnichen_v_osig", ob224="F1_obezsht_v_osig",
                      natura="F10_natura_asimetria", prev="F10_previshenie_asim")

    for x in rd:
        r = x["r"]
        el = x["elementi"]
        if x["bez_trud"]:
            continue                   # няма доход от трудова дейност — няма състав
        # Неустановима практика спира само изводите за състава на осигурителния
        # доход. Проверките, които не зависят от нея — таванът, данъчната основа,
        # лимитът на облекчението — вървят нататък.
        praktika_yasna = not any(el[k] and pol[k] is None for k in ("natura", "prev"))
        dopustimi = {k for k in ("natura", "prev") if pol[k] and el[k]}
        s_dop = r2(sum(el[k] for k in dopustimi))
        ochakvan = r2(x["baza"] + s_dop)

        if x["na_tavan"]:
            # проверява се само дали таванът е правилният за периода
            if abs(x["osig"] - drug_tavan) <= M.TOL and drug_tavan < max_osig \
                    and ochakvan > drug_tavan + M.TOL:
                N.add("B4_taван_ot_drug_period", "fayl",
                      f"осигурителният доход е ограничен до {drug_tavan:.2f} — тавана от "
                      f"другия период на годината, а не до приложимия {max_osig:.2f}",
                      drug_tavan, max_osig)
        elif praktika_yasna and abs(x["osig"] - ochakvan) > M.TOL:
            dobaveni = [k for k, v in el.items() if v and k not in dopustimi
                        and abs(r2(ochakvan + v) - x["osig"]) <= M.TOL]
            izvadeni = [k for k in dopustimi
                        if abs(r2(ochakvan - el[k]) - x["osig"]) <= M.TOL]
            if len(dobaveni) == 1:
                k = dobaveni[0]
                N.add(ID_DOBAVEN[k], r,
                      f"{IMENA_EL[k]} ({el[k]:.2f}) е включен в осигурителния доход, "
                      f"а останалите редове не го включват", x["osig"], ochakvan)
            elif len(izvadeni) == 1:
                k = izvadeni[0]
                N.add(ID_DOBAVEN[k], r,
                      f"{IMENA_EL[k]} ({el[k]:.2f}) е изваден от осигурителния доход, "
                      f"а останалите редове го включват", x["osig"], ochakvan)
            else:
                kand = ", ".join(f"{IMENA_EL[k]} {el[k]:.2f}" for k in (dobaveni or izvadeni)) \
                    or "няма"
                N.add("F1_osig_neobyasnen", r,
                      f"осигурителният доход {x['osig']:.2f} не съвпада с базата за труд "
                      f"{x['baza']:.2f} плюс допустимото по практиката на файла "
                      f"({ochakvan:.2f}); кандидати за разликата: {kand}", x["osig"], ochakvan)

        if x["osig"] > max_osig + M.TOL:
            N.add("B4_taван_ot_drug_period", "fayl",
                  f"осигурителен доход {x['osig']:.2f} над максималния {max_osig:.2f}",
                  x["osig"], max_osig)

        # ---------------- състав на данъчната основа --------------------
        if not praktika_yasna:
            continue        # без установена практика няма и очаквана данъчна основа
        # Същата логика като при осигурителния доход: първо се проверява
        # хипотезата „както предвижда практиката на файла", после се търси едно
        # отклонение, което обяснява разликата. Изброяването на всички
        # подмножества дава по няколко решения и води до грешна локализация.
        def dan_za(delta, rezhim_oblek):
            predi = r2(x["bruto"] + s_dop + delta - x["lichni"])
            if rezhim_oblek == "limit":
                ob = r2(min(x["udr_osig"], r2(predi * M.OBLEKCHENIE_LIMIT))) \
                    if x["udr_osig"] else 0.0
            elif rezhim_oblek == "pulno":
                ob = x["udr_osig"]
            else:
                ob = 0.0
            return r2(predi - ob), ob, predi

        osnovna_hipoteza, oblek, dan_predi = dan_za(0.0, "limit")
        if abs(x["dan_osnova"] - osnovna_hipoteza) > M.TOL:
            kandidati = [(0.0, "pulno", "F7_oblekchenie_nad_limit",
                          "приспаднато е цялото удържано, без ограничението до 10%")]
            if x["boln"]:
                kandidati.append((-x["boln"], "limit", "F9_bolnichen_bez_danak",
                                  "обезщетението по чл. 40, ал. 5 КСО е извадено от "
                                  "данъчната основа; то е облагаем доход"))
            if el["ob224"]:
                kandidati.append((-el["ob224"], "limit", "F6_obezsht_bez_danak",
                                  "обезщетението по чл. 224 КТ е извадено от данъчната "
                                  "основа"))
            for k in ("natura", "prev"):
                if not el[k] or not praktika_yasna:
                    continue
                znak = -1.0 if k in dopustimi else 1.0
                kandidati.append((znak * el[k], "limit", ID_DOBAVEN[k],
                                  f"{IMENA_EL[k]} ({el[k]:.2f}) е "
                                  + ("изваден от" if znak < 0 else "добавен само в")
                                  + " данъчната основа, а осигурителният доход го "
                                  + ("включва" if znak < 0 else "не включва")))
            namereni = []
            for delta, rezh, ident, opis in kandidati:
                stoynost, ob, predi = dan_za(delta, rezh)
                if abs(x["dan_osnova"] - stoynost) <= M.TOL:
                    namereni.append((ident, opis, ob, predi))
            if len(namereni) == 1:
                ident, opis, ob, predi = namereni[0]
                if ident == "F7_oblekchenie_nad_limit":
                    limit = r2(predi * M.OBLEKCHENIE_LIMIT)
                    if ob > limit + M.TOL:
                        N.add(ident, r,
                              f"приспаднати са {ob:.2f} при лимит 10% = {limit:.2f} "
                              f"(чл. 42, ал. 3 във вр. с чл. 19 ЗДДФЛ)", ob, limit)
                else:
                    N.add(ident, r, opis, x["dan_osnova"], osnovna_hipoteza)
            else:
                N.add("F6_dan_neobyasnena", r,
                      f"данъчната основа {x['dan_osnova']:.2f} не се обяснява с бруто "
                      f"{x['bruto']:.2f} минус личните вноски {x['lichni']:.2f} "
                      f"(очаквано {osnovna_hipoteza:.2f}); "
                      + (f"{len(namereni)} възможни обяснения" if namereni
                         else "нито едно от известните отклонения не пасва"),
                      x["dan_osnova"], osnovna_hipoteza)

    # --- F5: производен процент ТЗПБ -------------------------------------
    proizvodni = [r2(x["er_doo"] / x["osig"] * 100.0 - M.RABOTODATEL_DOO)
                  for x in rd if x["osig"] > 0]
    if proizvodni:
        p = Counter(proizvodni).most_common(1)[0][0]
        if p < tzpb_dalzhim - 0.005:
            N.add("F5_tzpb_pod_dalzhimiya", "fayl",
                  f"вноските на работодателя дават ТЗПБ {p}% при дължим {tzpb_dalzhim}%",
                  p, tzpb_dalzhim)

    # --- K5: редът ОБЩО ---------------------------------------------------
    for k in M.KOLONI_SBOR:
        s = r2(sum(ws.cell(row=r, column=kol[k]).value or 0 for r in redove))
        val = ws.cell(row=TOT, column=kol[k]).value or 0
        # Над цент и половина. Един цент разлика идва от закръгляването на самия
        # сбор и от реда на събиране на числата с плаваща точка — това е шум, не
        # ръчно вписан сбор.
        if abs(val - s) > 0.015:
            N.add("K5_sbor_ne_e_sum", "fayl",
                  f"в реда ОБЩО „{k}“ е {val}, а сумата на клетките е {s}", val, s)

    # ------------------------------------------------------- сравнение
    ochakvani = {("fayl" if kade == "fayl" else HDR + 1 + idx, ident)
                 for kade, idx, ident in man["ochakvani"]}
    # Информативните находки не твърдят дефект, а описват докъде стигат данните.
    # Не влизат в сравнението, но се печатат.
    INFORMATIVNI = {"F10_praktika_neustanovima"}
    nameren = {k for k in N.kljuchove() if k[1] not in INFORMATIVNI}
    propusnati, izlishni = ochakvani - nameren, nameren - ochakvani

    if not tiho:
        print(f"=== {os.path.basename(xlsx)} · {man['mesec']:02d}.{man['godina']} · "
              f"режим {man['rezhim']} (таван {max_osig}) · {norma} дни · "
              f"{len(rd)} лица · ТЗПБ {tzpb_dalzhim}% ===")
        print(f"политика на файла: изведена {pol} | заложена {man['politika']}")
        print(f"\nНАХОДКИ ({len(N.spisak)}):")
        for n in sorted(N.spisak, key=lambda x: (str(x["kade"]), x["id"])):
            kade = "файл" if n["kade"] == "fayl" else f"ред {n['kade']}"
            print(f"  [{kade:8}] {n['id']:26} {n['opisanie']}")
            if n["dalzhimo"] is not None:
                print(f"{'':13} начислено {n['nachisleno']} | дължимо {n['dalzhimo']}")
        if propusnati:
            print(f"\nПРОПУСНАТИ ({len(propusnati)}):")
            for kade, ident in sorted(propusnati, key=str):
                print(f"  {kade} · {ident} — {M.SCENARII.get(ident, ('', '?'))[1]}")
        if izlishni:
            print(f"\nФАЛШИВИ ПОЛОЖИТЕЛНИ ({len(izlishni)}):")
            for kade, ident in sorted(izlishni, key=str):
                print(f"  {kade} · {ident}")
        print(f"\nвкарани {len(ochakvani)} · намерени {len(ochakvani & nameren)} · "
              f"пропуснати {len(propusnati)} · излишни {len(izlishni)}")

    return dict(vkarani=len(ochakvani), nameren=len(ochakvani & nameren),
                propusnati=sorted(f"{a}:{b}" for a, b in propusnati),
                izlishni=sorted(f"{a}:{b}" for a, b in izlishni),
                nahodki=len(N.spisak))


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--seed", type=int, default=1)
    ap.add_argument("--mesec", type=int, default=None, choices=[6, 7, 8, 9, 10, 11])
    ap.add_argument("--tiho", action="store_true")
    a = ap.parse_args()

    sys.path.insert(0, HERE)
    import generate_struktura as G

    xlsx, mpath, _ = G.generirai(a.seed, a.mesec)
    rez = proveri(xlsx, mpath, tiho=a.tiho)
    sys.exit(0 if not rez["propusnati"] and not rez["izlishni"] else 1)
