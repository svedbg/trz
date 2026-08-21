# -*- coding: utf-8 -*-
"""Генератор на тестови ведомости с произволни данни и нарочно вкарани дефекти.

    python test/generate_struktura.py --seed 7

Строи `test/tmp/struktura_<seed>.xlsx` и манифест `..._manifest.json` с точния
списък на вкараното. Всичко е измислено и се преизчислява от семето: имена,
фирма, заплати, дни, цени на придобивките, месец, процент ТЗПБ. Нито едно число
не идва от истинска ведомост.

Дефектите се вкарват като мутация върху коректно изчислен ред, като зависимите
клетки се преизчисляват така, както би ги преизчислил и сгрешилият файл. Целта е
всеки дефект да поражда точно определен набор находки — иначе тестът не може да
различи пропусната находка от каскаден ефект.

Ред на вкарване: първо дефектите на нивото на файла (процент ТЗПБ, приложен
таван), защото те се отразяват на всички редове, и едва след това дефектите по
редове. Обратният ред би затрил вече вкарана мутация.

Еднозначност. Проверката за състава на осигурителния доход и на данъчната основа
работи, като търси кое подмножество от елементите обяснява обявеното число.
Затова генераторът гарантира, че сумите на всички подмножества от
{доход в натура, превишение, болнични, обезщетение} се различават достатъчно —
иначе задачата има повече от едно решение и находката не може да се локализира.
Това не е компромис на теста: в истински файл същата двусмислица прави извода
несигурен и скилът е длъжен да го каже, вместо да гадае.
"""
import argparse
import itertools
import json
import os
import random

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import trz_model as M
from trz_model import r2

HERE = os.path.dirname(os.path.abspath(__file__))
TMP = os.path.join(HERE, "tmp")
RAZDELIMOST = 0.50      # минимална разлика между сумите на две подмножества

# --- измислени имена. Нарочно различни от кой да е реален списък. ---
IMENA_M = ["Борислав", "Дарин", "Захари", "Ивайло", "Камен", "Лъчезар", "Никифор",
           "Орлин", "Тихомир", "Юлиан", "Явор", "Емил"]
IMENA_F = ["Ана", "Ваня", "Галина", "Емилия", "Жанета", "Красимира", "Магдалена",
           "Пенка", "Симона", "Христина", "Здравка", "Теодора"]
FAMILII_M = ["Аврамов", "Бакалов", "Влахов", "Гошев", "Даскалов", "Еленков", "Жеков",
             "Зидаров", "Кожухаров", "Лозанов", "Мутафчиев", "Ненов", "Орешков",
             "Пенчев", "Радулов", "Сивков", "Тошков", "Узунов", "Фандъков", "Хаджиев",
             "Цветков", "Чакъров", "Шивачев", "Янев"]
OTDELI = ["Разработка", "Внедряване", "Поддръжка", "Операции", "Администрация"]
FIRMI = ["Тестова Дигитал", "Пробна Софтуер", "Примерна Интеграции",
         "Демонстрационна Системи", "Условна Технолоджи"]
MESETSI_BG = {6: "юни", 7: "юли", 8: "август", 9: "септември", 10: "октомври", 11: "ноември"}


def _ime(rnd, izpolzvani):
    while True:
        if rnd.random() < 0.45:
            ime = f"{rnd.choice(IMENA_F)} {rnd.choice(FAMILII_M)}а"
        else:
            ime = f"{rnd.choice(IMENA_M)} {rnd.choice(FAMILII_M)}"
        if ime not in izpolzvani:
            izpolzvani.add(ime)
            return ime


def _razdelimi(stoynosti):
    """Всички подмножествени суми различават ли се достатъчно?"""
    st = [v for v in stoynosti if v]
    sumi = []
    for k in range(len(st) + 1):
        for c in itertools.combinations(st, k):
            sumi.append(r2(sum(c)))
    sumi.sort()
    return all(b - a > RAZDELIMOST for a, b in zip(sumi, sumi[1:]))


def _izpolzvaemi(lica, tavan_prilozhim, tavan_ef):
    """Колко реда позволяват практиката на файла да се извади за всеки елемент.

    Използваем е ред с начисления за труд, който не стои на таван — там сумата
    е една и съща за много комбинации и нищо не се вижда. Броят има значение:
    под три реда практиката не е установима, а дефект, който зависи от нея, не
    може да се локализира — нито от теста, нито от жив проверяващ.
    """
    broy = {"natura": 0, "prev": 0}
    for y in lica:
        red = y["red"]
        baza = r2(red["Основна за отработеното"] + red["Клас сума"]
                  + red["Бонус"] + red["Платен отпуск"])
        if baza <= 0:
            continue
        if any(abs(red["Осигурителен доход"] - c) <= M.TOL
               for c in (tavan_prilozhim, tavan_ef)):
            continue
        if red["Карта (за сметка на работодателя)"]:
            broy["natura"] += 1
        prem = red["Доброволно здравно осигуряване (премия)"]
        if prem and r2(prem - M.PRAG_SOTSIALNI_RAZHODI) > 0:
            broy["prev"] += 1
    return broy


def _nosi(red, el):
    if el == "natura":
        return red["Карта (за сметка на работодателя)"] > 0
    prem = red["Доброволно здравно осигуряване (премия)"]
    return bool(prem) and r2(prem - M.PRAG_SOTSIALNI_RAZHODI) > 0


def _elementi(red):
    premia = red["Доброволно здравно осигуряване (премия)"]
    return [red["Карта (за сметка на работодателя)"],
            r2(max(0.0, premia - M.PRAG_SOTSIALNI_RAZHODI)) if premia else 0.0,
            red["Болнични (работодател)"],
            red["Обезщетение чл. 224"]]


def vhodni_danni(rnd, norma, rezhim):
    """Случайно, но правдоподобно лице."""
    tip = rnd.random()
    if tip < 0.15:
        ms = r2(rnd.uniform(rezhim["mrz"], rezhim["mrz"] * 1.4))
    elif tip < 0.7:
        ms = r2(rnd.uniform(1200, 2600))
    else:
        ms = r2(rnd.uniform(2600, 9000))

    klas_pr = rnd.choice([0, 0, 0.6, 1.2, 1.8, 2.4, 3.0, 4.2, 4.8, 6.0, 7.2, 9.0])
    sd = rnd.choice([0] * 8 + [2, 3, 5])
    pl = 0 if sd else rnd.choice([0, 0, 1, 2, 3, 5, 8, 11])
    pl = min(pl, norma - sd)
    wd = norma - pl - sd

    karta = rnd.random() < 0.6
    premia_da = rnd.random() < 0.8
    return dict(
        mesechna_zaplata=ms, klas_pr=klas_pr,
        dni_rabota=wd, dni_otpusk=pl, dni_bolnichen=sd, dni_maychinstvo=0,
        bonus=r2(rnd.uniform(50, 400)) if rnd.random() < 0.18 else 0.0,
        obezsht_224=0.0,
        karta_er=r2(rnd.uniform(38, 72)) if karta else 0.0,
        karta_ee=r2(rnd.uniform(3, 12)) if karta else 0.0,
        premia=r2(rnd.uniform(32.9, 44.5)) if premia_da else 0.0,
        lichna_vnoska=r2(rnd.uniform(20, 120)) if rnd.random() < 0.15 else 0.0,
    )


def lice(rnd, norma, rezhim, tzpb, politika):
    """Връща (вход, чист ред) с гарантирана разделимост на елементите."""
    for _ in range(30):
        vhod = vhodni_danni(rnd, norma, rezhim)
        red = M.chist_red(vhod, rezhim, tzpb, politika, norma)
        if _razdelimi(_elementi(red)):
            red["_norma"] = norma
            return vhod, red
        # нагласяме придобивките, вместо да въртим цялото лице
        for _ in range(20):
            vhod["karta_er"] = r2(rnd.uniform(38, 72)) if vhod["karta_er"] else 0.0
            vhod["premia"] = r2(rnd.uniform(32.9, 44.5)) if vhod["premia"] else 0.0
            red = M.chist_red(vhod, rezhim, tzpb, politika, norma)
            if _razdelimi(_elementi(red)):
                red["_norma"] = norma
                return vhod, red
    vhod["karta_er"] = vhod["karta_ee"] = 0.0
    vhod["premia"] = 0.0
    red = M.chist_red(vhod, rezhim, tzpb, politika, norma)
    red["_norma"] = norma
    return vhod, red


# =====================================================================
#                             мутации
# Всяка връща (нов ред, множество очаквани id) или None, ако редът не е
# подходящ за този дефект.
# =====================================================================

def _preizchisli_sled_bruto(red, vhod, rezhim, tzpb, politika, *,
                            osig=None, dan_osnova=None):
    """Преизчислява вноски, данък, нето и разход след промяна в брутото/базите —
    както би го направил файл, чиито надолу вървящи колони са формули."""
    bruto = red["БРУТО"]
    natura = red["Карта (за сметка на работодателя)"]
    premia = red["Доброволно здравно осигуряване (премия)"]
    prev = r2(max(0.0, premia - M.PRAG_SOTSIALNI_RAZHODI)) if premia else 0.0
    baza = r2(red["Основна за отработеното"] + red["Клас сума"]
              + red["Бонус"] + red["Платен отпуск"])
    dobavki = 0.0 if baza <= 0 else (
        (natura if politika["natura_v_bazite"] else 0.0)
        + (prev if politika["previshenie_v_bazite"] else 0.0))

    if osig is None:
        osig = r2(min(rezhim["max_osig"], r2(baza + dobavki)))
    red["Осигурителен доход"] = osig

    for kolona, klyuch in (("ДОО пенсии", "pensii"), ("ДОО ОЗМ", "ozm"),
                           ("ДОО безработица", "bezrab"), ("ЗО лична", "zo"),
                           ("ДЗПО-УПФ лична", "upf")):
        red[kolona] = r2(osig * M.LICHNI[klyuch] / 100.0)
    lichni = r2(sum(red[k] for k in ("ДОО пенсии", "ДОО ОЗМ", "ДОО безработица",
                                     "ЗО лична", "ДЗПО-УПФ лична")))
    red["Лични вноски общо"] = lichni

    if dan_osnova is None:
        predi = r2(bruto + dobavki - lichni)
        limit = r2(predi * M.OBLEKCHENIE_LIMIT)
        oblek = r2(min(red["Удръжка доброволно осиг. (лична)"], limit)) \
            if red["Удръжка доброволно осиг. (лична)"] else 0.0
        dan_osnova = r2(predi - oblek)
    red["Данъчна основа"] = dan_osnova
    red["ДДФЛ"] = r2(dan_osnova * M.DDFL)

    red["НЕТО преди удръжки"] = r2(bruto - lichni - red["ДДФЛ"])
    red["НЕТО за изплащане"] = r2(red["НЕТО преди удръжки"]
                                  - red["Удръжка доброволно осиг. (лична)"]
                                  - red["Удръжка карта (лична част)"])
    red["Изплатено"] = red["НЕТО за изплащане"]
    red["Разлика"] = 0.0

    red["Вноски работодател ДОО+ТЗПБ"] = r2(osig * (M.RABOTODATEL_DOO + tzpb) / 100.0)
    red["ДЗПО-УПФ работодател"] = r2(osig * M.RABOTODATEL_UPF / 100.0)
    red["ЗО работодател"] = r2(osig * M.RABOTODATEL_ZO / 100.0)
    red["Вноски работодател общо"] = r2(red["Вноски работодател ДОО+ТЗПБ"]
                                       + red["ДЗПО-УПФ работодател"]
                                       + red["ЗО работодател"]
                                       + red["ЗО при болничен/майчинство"])
    red["Общ разход за труд"] = r2(bruto + red["Вноски работодател общо"] + natura + premia)
    return red


def m_K1(red, vhod, rez, tzpb, pol, rnd):
    """Брутото не включва колона за начисления (тук: обезщетението по чл. 224)."""
    red = dict(red)
    for _ in range(40):
        ob = r2(rnd.uniform(120, 900))
        proba = dict(red, **{"Обезщетение чл. 224": ob})
        if _razdelimi(_elementi(proba)):
            red = proba
            break
    else:
        return None
    red["БРУТО"] = r2(red["Основна за отработеното"] + red["Клас сума"] + red["Бонус"]
                      + red["Платен отпуск"] + red["Болнични (работодател)"])
    return _preizchisli_sled_bruto(red, vhod, rez, tzpb, pol), {"K1_sbor_izpuska_kolona"}


def m_K2(red, vhod, rez, tzpb, pol, rnd):
    """Сумата на болничния, въведена в колоната за дни болничен."""
    if not vhod["dni_bolnichen"] or not red["Болнични (работодател)"]:
        return None
    red = dict(red)
    red["Дни болничен"] = red["Болнични (работодател)"]
    return red, {"K2_suma_v_kolona_dni", "I5_dni_ne_se_vrazvat"}


def m_K3(red, vhod, rez, tzpb, pol, rnd):
    """Вноските са твърди стойности от предходен период — изостанали са."""
    red = dict(red)
    star = r2(red["Осигурителен доход"] * rnd.uniform(0.86, 0.95))
    if red["Осигурителен доход"] - star < 5:
        return None
    for kolona, klyuch in (("ДОО пенсии", "pensii"), ("ДОО ОЗМ", "ozm"),
                           ("ДОО безработица", "bezrab"), ("ЗО лична", "zo"),
                           ("ДЗПО-УПФ лична", "upf")):
        red[kolona] = r2(star * M.LICHNI[klyuch] / 100.0)
    lichni = r2(sum(red[k] for k in ("ДОО пенсии", "ДОО ОЗМ", "ДОО безработица",
                                     "ЗО лична", "ДЗПО-УПФ лична")))
    red["Лични вноски общо"] = lichni
    natura = red["Карта (за сметка на работодателя)"]
    premia = red["Доброволно здравно осигуряване (премия)"]
    prev = r2(max(0.0, premia - M.PRAG_SOTSIALNI_RAZHODI)) if premia else 0.0
    baza = r2(red["Основна за отработеното"] + red["Клас сума"] + red["Бонус"]
              + red["Платен отпуск"])
    dobavki = 0.0 if baza <= 0 else (
        (natura if pol["natura_v_bazite"] else 0.0)
        + (prev if pol["previshenie_v_bazite"] else 0.0))
    predi = r2(red["БРУТО"] + dobavki - lichni)
    limit = r2(predi * M.OBLEKCHENIE_LIMIT)
    oblek = r2(min(red["Удръжка доброволно осиг. (лична)"], limit)) \
        if red["Удръжка доброволно осиг. (лична)"] else 0.0
    red["Данъчна основа"] = r2(predi - oblek)
    red["ДДФЛ"] = r2(red["Данъчна основа"] * M.DDFL)
    red["НЕТО преди удръжки"] = r2(red["БРУТО"] - lichni - red["ДДФЛ"])
    red["НЕТО за изплащане"] = r2(red["НЕТО преди удръжки"]
                                  - red["Удръжка доброволно осиг. (лична)"]
                                  - red["Удръжка карта (лична част)"])
    red["Изплатено"] = red["НЕТО за изплащане"]
    return red, {"K3_tvardi_stoynosti"}


def m_K4(red, vhod, rez, tzpb, pol, rnd):
    """Изплатено е под нетото, а контролната колона показва нула."""
    red = dict(red)
    red["Изплатено"] = r2(red["НЕТО за изплащане"] - rnd.choice([0.05, 0.13, 0.40, 1.00]))
    red["Разлика"] = 0.0
    return red, {"K4_kontrola_ne_hvashta"}


def m_K6(red, vhod, rez, tzpb, pol, rnd):
    """Класът е без закръгляване и незакръглената стойност влиза в брутото."""
    if not vhod["klas_pr"]:
        return None
    nezakr = red["Основна за отработеното"] * vhod["klas_pr"] / 100.0
    if abs(nezakr - r2(nezakr)) < 0.002:
        return None                     # няма видима опашка — нищо за откриване
    red = dict(red)
    red["Клас сума"] = nezakr
    red["БРУТО"] = (red["Основна за отработеното"] + red["Клас сума"] + red["Бонус"]
                    + red["Платен отпуск"] + red["Обезщетение чл. 224"]
                    + red["Болнични (работодател)"])
    red = _preizchisli_sled_bruto(red, vhod, rez, tzpb, pol)
    red["Изплатено"] = r2(red["НЕТО за изплащане"])
    red["Разлика"] = r2(red["НЕТО за изплащане"] - red["Изплатено"])
    return red, {"K6_nezakraglen"}


def m_K7(red, vhod, rez, tzpb, pol, rnd):
    """Общият разход е сметнат от нетото след удръжките."""
    udr = r2(red["Удръжка доброволно осиг. (лична)"] + red["Удръжка карта (лична част)"])
    if udr < 1.0:
        return None
    red = dict(red)
    red["Общ разход за труд"] = r2(red["Общ разход за труд"] - udr)
    return red, {"K7_razhod_ot_neto"}


def m_F9_v_osig(red, vhod, rez, tzpb, pol, rnd):
    """Обезщетението по чл. 40, ал. 5 КСО е включено в осигурителния доход."""
    if not red["Болнични (работодател)"]:
        return None
    red = dict(red)
    natura = red["Карта (за сметка на работодателя)"]
    premia = red["Доброволно здравно осигуряване (премия)"]
    prev = r2(max(0.0, premia - M.PRAG_SOTSIALNI_RAZHODI)) if premia else 0.0
    baza = r2(red["Основна за отработеното"] + red["Клас сума"] + red["Бонус"]
              + red["Платен отпуск"])
    dobavki = (natura if pol["natura_v_bazite"] else 0.0) \
            + (prev if pol["previshenie_v_bazite"] else 0.0)
    osig = r2(min(rez["max_osig"], r2(baza + dobavki + red["Болнични (работодател)"])))
    if abs(osig - red["Осигурителен доход"]) < 1.0:
        return None                      # таванът го поглъща — няма видим дефект
    if osig > rez["max_osig"] - 1.0:
        # резултатът опира тавана: съставът вече не е възстановим и находката не
        # може да се локализира — нито от теста, нито от жив проверяващ
        return None
    return _preizchisli_sled_bruto(red, vhod, rez, tzpb, pol, osig=osig), \
        {"F9_bolnichen_v_osig"}


def m_F9_bez_danak(red, vhod, rez, tzpb, pol, rnd):
    """Обезщетението по чл. 40, ал. 5 КСО е извадено от данъчната основа."""
    if not red["Болнични (работодател)"]:
        return None
    red = dict(red)
    natura = red["Карта (за сметка на работодателя)"]
    premia = red["Доброволно здравно осигуряване (премия)"]
    prev = r2(max(0.0, premia - M.PRAG_SOTSIALNI_RAZHODI)) if premia else 0.0
    baza = r2(red["Основна за отработеното"] + red["Клас сума"] + red["Бонус"]
              + red["Платен отпуск"])
    dobavki = 0.0 if baza <= 0 else (
        (natura if pol["natura_v_bazite"] else 0.0)
        + (prev if pol["previshenie_v_bazite"] else 0.0))
    predi = r2(red["БРУТО"] - red["Болнични (работодател)"] + dobavki - red["Лични вноски общо"])
    limit = r2(predi * M.OBLEKCHENIE_LIMIT)
    oblek = r2(min(red["Удръжка доброволно осиг. (лична)"], limit)) \
        if red["Удръжка доброволно осиг. (лична)"] else 0.0
    return _preizchisli_sled_bruto(red, vhod, rez, tzpb, pol,
                                   osig=red["Осигурителен доход"],
                                   dan_osnova=r2(predi - oblek)), \
        {"F9_bolnichen_bez_danak"}


def m_F9_bez_zo(red, vhod, rez, tzpb, pol, rnd):
    """Липсва здравната вноска по чл. 40, ал. 1, т. 5 ЗЗО."""
    if not red["ЗО при болничен/майчинство"]:
        return None
    red = dict(red)
    red["ЗО при болничен/майчинство"] = 0.0
    red["Вноски работодател общо"] = r2(red["Вноски работодател ДОО+ТЗПБ"]
                                       + red["ДЗПО-УПФ работодател"] + red["ЗО работодател"])
    red["Общ разход за труд"] = r2(red["БРУТО"] + red["Вноски работодател общо"]
                                   + red["Карта (за сметка на работодателя)"]
                                   + red["Доброволно здравно осигуряване (премия)"])
    return red, {"F9_bez_zo_bolnichen"}


def _asimetria(red, vhod, rez, tzpb, pol, element, ident, nositeli=99):
    """Прави елемента видим само в едната от двете бази.

    `nositeli` е броят лица във ведомостта с този елемент. При по-малко от три
    практиката на файла не е установима — при двама носители единият случай е
    точно толкова „правило", колкото и другият. Тогава дефектът не се вкарва,
    защото и жив проверяващ не би могъл да го локализира, а би попитал.
    """
    if nositeli < 4:
        return None
    natura = red["Карта (за сметка на работодателя)"]
    premia = red["Доброволно здравно осигуряване (премия)"]
    prev = r2(max(0.0, premia - M.PRAG_SOTSIALNI_RAZHODI)) if premia else 0.0
    stoynost = natura if element == "natura" else prev
    if stoynost < 1.0:
        return None
    baza = r2(red["Основна за отработеното"] + red["Клас сума"] + red["Бонус"]
              + red["Платен отпуск"])
    if baza <= 0:
        return None
    # ако таванът е стигнат, съставът на осигурителния доход не е разпознаваем
    if r2(baza + natura + prev) > rez["max_osig"] - 1.0:
        return None
    vkl = pol["natura_v_bazite"] if element == "natura" else pol["previshenie_v_bazite"]
    red = dict(red)
    drug = (prev if element == "natura" and pol["previshenie_v_bazite"] else 0.0) \
         + (natura if element == "previshenie" and pol["natura_v_bazite"] else 0.0)
    if vkl:
        osig = r2(min(rez["max_osig"], r2(baza + drug)))       # махнат само от осиг. доход
    else:
        osig = red["Осигурителен доход"]                        # добавен само в данъчната основа
    if vkl:
        red = _preizchisli_sled_bruto(red, vhod, rez, tzpb, pol, osig=osig)
    else:
        # елементът влиза само в данъчната основа; облекчението се ограничава
        # спрямо същата тази основа, иначе файлът би бил непоследователен и в
        # трето отношение, а тестът трябва да мери един дефект наведнъж
        dobavki_dan = r2(drug + stoynost)
        lichni_sled = r2(sum(r2(osig * M.LICHNI[k] / 100.0) for k in M.LICHNI))
        predi = r2(red["БРУТО"] + dobavki_dan - lichni_sled)
        limit = r2(predi * M.OBLEKCHENIE_LIMIT)
        vnoska = red["Удръжка доброволно осиг. (лична)"]
        oblek = r2(min(vnoska, limit)) if vnoska else 0.0
        red = _preizchisli_sled_bruto(red, vhod, rez, tzpb, pol, osig=osig,
                                      dan_osnova=r2(predi - oblek))
    return red, {ident}


def m_F10_natura(red, vhod, rez, tzpb, pol, rnd, nositeli=99):
    return _asimetria(red, vhod, rez, tzpb, pol, "natura", "F10_natura_asimetria", nositeli)


def m_F10_prev(red, vhod, rez, tzpb, pol, rnd, nositeli=99):
    return _asimetria(red, vhod, rez, tzpb, pol, "previshenie", "F10_previshenie_asim",
                      nositeli)


def m_F7(red, vhod, rez, tzpb, pol, rnd):
    """Данъчното облекчение е приложено над 10% от месечната данъчна основа."""
    red = dict(red)
    natura = red["Карта (за сметка на работодателя)"]
    premia = red["Доброволно здравно осигуряване (премия)"]
    prev = r2(max(0.0, premia - M.PRAG_SOTSIALNI_RAZHODI)) if premia else 0.0
    baza = r2(red["Основна за отработеното"] + red["Клас сума"] + red["Бонус"]
              + red["Платен отпуск"])
    if baza <= 0:
        return None
    dobavki = (natura if pol["natura_v_bazite"] else 0.0) \
            + (prev if pol["previshenie_v_bazite"] else 0.0)
    predi = r2(red["БРУТО"] + dobavki - red["Лични вноски общо"])
    vnoska = r2(predi * M.OBLEKCHENIE_LIMIT + rnd.uniform(30, 150))
    red["Удръжка доброволно осиг. (лична)"] = vnoska
    red["Данъчна основа"] = r2(predi - vnoska)
    red["ДДФЛ"] = r2(red["Данъчна основа"] * M.DDFL)
    red["НЕТО преди удръжки"] = r2(red["БРУТО"] - red["Лични вноски общо"] - red["ДДФЛ"])
    red["НЕТО за изплащане"] = r2(red["НЕТО преди удръжки"] - vnoska
                                  - red["Удръжка карта (лична част)"])
    red["Изплатено"] = red["НЕТО за изплащане"]
    return red, {"F7_oblekchenie_nad_limit"}


def m_C2(red, vhod, rez, tzpb, pol, rnd):
    """Класът е начислен върху брутото, не върху основната заплата."""
    kp = vhod["klas_pr"]
    if not kp:
        return None
    dobavka = r2((red["Платен отпуск"] + red["Бонус"]) * kp / 100.0)
    if dobavka < 0.10:                     # без отпуск и бонус разликата е нула
        return None
    red = dict(red)
    baza = r2(red["Основна за отработеното"] + red["Платен отпуск"] + red["Бонус"])
    red["Клас сума"] = r2(baza * kp / 100.0)
    red["БРУТО"] = r2(red["Основна за отработеното"] + red["Клас сума"] + red["Бонус"]
                      + red["Платен отпуск"] + red["Обезщетение чл. 224"]
                      + red["Болнични (работодател)"])
    return _preizchisli_sled_bruto(red, vhod, rez, tzpb, pol), {"C2_klas_varhu_bruto"}


def m_E3(red, vhod, rez, tzpb, pol, rnd):
    """Платеният отпуск е изчислен без включения клас."""
    if not vhod["klas_pr"] or not vhod["dni_otpusk"]:
        return None
    dneven = vhod["mesechna_zaplata"] / red["_norma"]
    bez = r2(dneven * vhod["dni_otpusk"])
    if abs(bez - red["Платен отпуск"]) < 0.10:
        return None
    red = dict(red)
    red["Платен отпуск"] = bez
    red["БРУТО"] = r2(red["Основна за отработеното"] + red["Клас сума"] + red["Бонус"]
                      + red["Платен отпуск"] + red["Обезщетение чл. 224"]
                      + red["Болнични (работодател)"])
    return _preizchisli_sled_bruto(red, vhod, rez, tzpb, pol), {"E3_otpusk_bez_klas"}


def m_I5(red, vhod, rez, tzpb, pol, rnd):
    """Дните болничен не са отразени, макар сумата да е начислена."""
    if not vhod["dni_bolnichen"]:
        return None
    red = dict(red)
    red["Дни болничен"] = 0
    return red, {"I5_dni_ne_se_vrazvat"}


MUTATSII_RED = [
    ("K1_sbor_izpuska_kolona", m_K1),
    ("K2_suma_v_kolona_dni", m_K2),
    ("K3_tvardi_stoynosti", m_K3),
    ("K4_kontrola_ne_hvashta", m_K4),
    ("K6_nezakraglen", m_K6),
    ("K7_razhod_ot_neto", m_K7),
    ("F9_bolnichen_v_osig", m_F9_v_osig),
    ("F9_bolnichen_bez_danak", m_F9_bez_danak),
    ("F9_bez_zo_bolnichen", m_F9_bez_zo),
    ("F10_natura_asimetria", m_F10_natura),
    ("F10_previshenie_asim", m_F10_prev),
    ("F7_oblekchenie_nad_limit", m_F7),
    ("C2_klas_varhu_bruto", m_C2),
    ("E3_otpusk_bez_klas", m_E3),
    ("I5_dni_ne_se_vrazvat", m_I5),
]


# =====================================================================


def generirai(seed, mesec=None):
    rnd = random.Random(seed)
    mesec = mesec or rnd.choice([6, 7, 8, 9, 10, 11])
    godina = 2026
    norma = M.rabotni_dni(godina, mesec)
    rez_id = M.rezhim_za(godina, mesec)
    rezhim = M.REZHIMI[rez_id]
    tzpb = rnd.choice([0.4, 0.5, 0.7, 1.1])
    politika = dict(natura_v_bazite=rnd.random() < 0.5,
                    previshenie_v_bazite=rnd.random() < 0.5)
    firma = rnd.choice(FIRMI)
    n = rnd.randint(9, 15)

    # --- дефекти на нивото на файла се решават ПРЕДИ редовете -------------
    fayl_defekti = []
    tzpb_ef = tzpb
    if rnd.random() < 0.45:
        kandidat = round(tzpb - rnd.choice([0.1, 0.2, 0.3]), 2)
        if kandidat >= 0.1:
            tzpb_ef = kandidat
            fayl_defekti.append("F5_tzpb_pod_dalzhimiya")

    tavan_ef = rezhim["max_osig"]
    if rnd.random() < 0.35:
        tavan_ef = M.REZHIMI["H2" if rez_id == "H1" else "H1"]["max_osig"]
        fayl_defekti.append("B4_taван_ot_drug_period")
    rezhim_ef = dict(rezhim, max_osig=tavan_ef)

    izpolzvani = set()
    lica = []
    for _ in range(n):
        vhod, red = lice(rnd, norma, rezhim_ef, tzpb_ef, politika)
        lica.append(dict(ime=_ime(rnd, izpolzvani), otdel=rnd.choice(OTDELI),
                         vhod=vhod, red=red))

    # едно лице цял месец в майчинство: нула начисления, само ЗО по ЗЗО
    if n >= 10:
        vhod = dict(mesechna_zaplata=r2(rnd.uniform(1300, 2400)),
                    klas_pr=rnd.choice([0, 1.2, 2.4]),
                    dni_rabota=0, dni_otpusk=0, dni_bolnichen=0, dni_maychinstvo=norma,
                    bonus=0.0, obezsht_224=0.0, karta_er=0.0, karta_ee=0.0,
                    premia=r2(rnd.uniform(32.9, 44.5)), lichna_vnoska=0.0)
        red = M.chist_red(vhod, rezhim_ef, tzpb_ef, politika, norma)
        red["_norma"] = norma
        lica.append(dict(ime=_ime(rnd, izpolzvani), otdel=rnd.choice(OTDELI),
                         vhod=vhod, red=red))
        rnd.shuffle(lica)

    # B4 е находка само ако сгрешеният таван реално се вижда някъде
    if "B4_taван_ot_drug_period" in fayl_defekti:
        vidim = any(l["red"]["Осигурителен доход"] > rezhim["max_osig"] + M.TOL
                    or (abs(l["red"]["Осигурителен доход"] - tavan_ef) < M.TOL
                        and tavan_ef < rezhim["max_osig"])
                    for l in lica)
        if not vidim:
            fayl_defekti.remove("B4_taван_ot_drug_period")

    # --- дефекти по редове -------------------------------------------------
    ochakvani = []
    svobodni = list(range(len(lica)))
    rnd.shuffle(svobodni)
    kandidati = MUTATSII_RED[:]
    rnd.shuffle(kandidati)
    kolko = rnd.randint(5, 9)
    vkarani = 0
    kvarni = {"natura": 0, "prev": 0}      # редове с вече обърнат глас
    # Дефекти, чиято локализация минава през практиката на файла за придобивките.
    ZAVISI_OT_PRAKTIKA = ("F9_bolnichen_v_osig", "F10_natura_asimetria",
                          "F10_previshenie_asim", "F9_bolnichen_bez_danak",
                          "F7_oblekchenie_nad_limit")
    # От тях само тези развалят извадката, от която практиката се извежда.
    KVARI_IZVADKATA = ("F9_bolnichen_v_osig", "F10_natura_asimetria",
                       "F10_previshenie_asim")
    for ident, fn in kandidati:
        if vkarani >= kolko or not svobodni:
            break
        # Броят използваеми редове се преизчислява при всеки дефект: всяка вече
        # вкарана мутация може да размести кой ред стои на таван и кой не.
        izpolzvaemi = _izpolzvaemi(lica, rezhim["max_osig"], tavan_ef)
        for idx in list(svobodni):
            l = lica[idx]
            if ident in ZAVISI_OT_PRAKTIKA:
                # Практиката по всеки елемент, който редът носи, трябва да остане
                # установима СЛЕД мутацията. Затова се иска четвърти използваем
                # ред: сгрешеният ред или отпада от извадката (когато сумата му
                # вече не съвпада с никоя чиста комбинация), или гласува
                # обратното. И в двата случая три чисти реда трябва да останат.
                if any(_nosi(l["red"], el) and izpolzvaemi[el] < 4 + kvarni[el]
                       for el in ("natura", "prev")):
                    continue
            if ident.startswith("F10_"):
                el = "natura" if "natura" in ident else "prev"
                rez_mut = fn(l["red"], l["vhod"], rezhim_ef, tzpb_ef, politika, rnd,
                             nositeli=izpolzvaemi[el])
            else:
                rez_mut = fn(l["red"], l["vhod"], rezhim_ef, tzpb_ef, politika, rnd)
            if rez_mut is None:
                continue
            nov, ids = rez_mut
            nov["_norma"] = norma
            if ident in KVARI_IZVADKATA:
                # редът вече не е чист свидетел за практиката: или отпада от
                # извадката, или гласува обратното. Намалява бюджета за следващи
                # дефекти, зависещи от същия елемент.
                for e in ("natura", "prev"):
                    if _nosi(l["red"], e):
                        kvarni[e] += 1
            l["red"] = nov
            l["defekt"] = sorted(ids)
            svobodni.remove(idx)
            ochakvani += [["red", idx, i] for i in sorted(ids)]
            vkarani += 1
            break

    # --- запис ---------------------------------------------------------------
    os.makedirs(TMP, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{mesec:02d}-{godina}"
    ws["A1"] = f'ВЕДОМОСТ ЗА РАБОТНИ ЗАПЛАТИ — "{firma}" ЕООД'
    ws["A2"] = (f"Месец: {MESETSI_BG[mesec]} {godina} г.  |  Работни дни: {norma}  |  "
                f"Валута: EUR  |  ЕИК: 000000000 (тестов)")
    ws["A3"] = f"Икономическа дейност: тестова  |  ТЗПБ по КИД: {tzpb}%"
    ws["A1"].font = Font(bold=True, size=12)

    HDR = 5
    for i, kolona in enumerate(M.KOLONI, start=1):
        c = ws.cell(row=HDR, column=i, value=kolona)
        c.font = Font(bold=True, size=8)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    for r_off, l in enumerate(lica):
        r = HDR + 1 + r_off
        ws.cell(row=r, column=M.K["№"], value=r_off + 1)
        ws.cell(row=r, column=M.K["Име"], value=l["ime"])
        ws.cell(row=r, column=M.K["Отдел"], value=l["otdel"])
        for kolona in M.KOLONI:
            if kolona in ("№", "Име", "Отдел"):
                continue
            v = l["red"].get(kolona, 0)
            ws.cell(row=r, column=M.K[kolona],
                    value=(v if v else (0 if kolona in M.KOLONI_DNI else None)))

    tot = HDR + 1 + len(lica)
    ws.cell(row=tot, column=M.K["Име"], value="ОБЩО").font = Font(bold=True)
    for kolona in M.KOLONI_SBOR:
        s = r2(sum(l["red"].get(kolona, 0) or 0 for l in lica))
        ws.cell(row=tot, column=M.K[kolona], value=s).font = Font(bold=True)

    if rnd.random() < 0.4:
        kolona = rnd.choice(["Карта (за сметка на работодателя)",
                             "Доброволно здравно осигуряване (премия)",
                             "Удръжка карта (лична част)"])
        s = r2(sum(l["red"].get(kolona, 0) or 0 for l in lica))
        if s > 0:
            fayl_defekti.append("K5_sbor_ne_e_sum")
            ws.cell(row=tot, column=M.K[kolona], value=r2(s + len(lica) * 0.004 + 0.02))

    for i, kolona in enumerate(M.KOLONI, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(9, min(15, len(kolona) // 2 + 6))

    put = os.path.join(TMP, f"struktura_{seed}.xlsx")
    wb.save(put)

    ochakvani += [["fayl", None, i] for i in fayl_defekti]

    manifest = dict(
        seed=seed, fayl=os.path.basename(put), list=ws.title,
        godina=godina, mesec=mesec, norma_dni=norma, rezhim=rez_id,
        max_osig_prilozhim=rezhim["max_osig"], mod_samoosig=rezhim["mod_samoosig"],
        tzpb_dalzhim=tzpb, politika=politika, hdr=HDR, red_obshto=tot,
        lica=[dict(red=HDR + 1 + i, ime=l["ime"], vhod=l["vhod"],
                   defekt=l.get("defekt", [])) for i, l in enumerate(lica)],
        ochakvani=ochakvani,
    )
    mpath = os.path.join(TMP, f"struktura_{seed}_manifest.json")
    with open(mpath, "w", encoding="utf8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=1)
    return put, mpath, manifest


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--seed", type=int, default=1)
    ap.add_argument("--mesec", type=int, default=None, choices=[6, 7, 8, 9, 10, 11])
    a = ap.parse_args()
    p, mp, man = generirai(a.seed, a.mesec)
    print(f"Записано: {p}")
    print(f"Манифест: {mp}")
    print(f"Месец {man['mesec']:02d}.{man['godina']} · режим {man['rezhim']} · "
          f"{man['norma_dni']} работни дни · {len(man['lica'])} лица · ТЗПБ {man['tzpb_dalzhim']}%")
    print(f"Политика: {man['politika']}")
    print(f"Вкарани дефекти ({len(man['ochakvani'])}):")
    for kade, idx, ident in man["ochakvani"]:
        kade_s = "файл" if kade == "fayl" else f"ред {man['hdr'] + 1 + idx}"
        print(f"  {kade_s:9} {ident:28} {M.SCENARII[ident][1]}")
