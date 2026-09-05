#!/usr/bin/env python3
"""Tests for tools/preflight.py - the pre-flight check that runs before an audit.

Standalone, like skill_test.py: run it directly, not through run_tests.py, which owns
the five generated suites and says so in four places.

The proving standard is the repository's. A clean workbook must raise **no** signal at
all, and each planted shape defect must raise its own signal and **nothing else** - a
false positive fails exactly like a miss, the same rule the payroll suites are held to.
Asserting on signal ids rather than on the rendered Bulgarian is deliberate: prose can
match by coincidence.

Every fixture is built in memory from invented data. No real payroll is read, and
nothing is written outside a temporary directory.

    python test/preflight_test.py
"""

import json
import os
import shutil
import sys
import tempfile

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, os.path.join(ROOT, "tools"))
sys.path.insert(0, HERE)

import generate_shapes as G                                    # noqa: E402
import preflight as P                                          # noqa: E402
import trz_model as M                                          # noqa: E402

FAILURES = []


def check(condition, message):
    print(("  ok   " if condition else "  FAIL ") + message)
    if not condition:
        FAILURES.append(message)


def signals_for(path, mapping=None, kid="62", group="3", tzpb="0.4"):
    data = P.analyse(path, mapping, kid, group, tzpb)
    data["_path"] = path
    return P.all_signals(data), data


def main():
    tmp = tempfile.mkdtemp(prefix="preflight-test-")
    try:
        print("Pre-flight check")
        print("=" * 78)

        # ---------------------------------------------------------- vocabulary drift
        # Every canonical header the generated payrolls use must still be recognised, or
        # a rename in trz_model.py silently sends real columns to the unknown list.
        wanted = {"Име", "Отраб. дни", "Основна за отработеното", "Клас сума", "БРУТО",
                  "Осигурителен доход", "Данъчна основа", "ДДФЛ", "Лични вноски общо",
                  "НЕТО за изплащане", "Изплатено", "Дни болничен",
                  "Болнични (работодател)", "Вноски работодател общо"}
        missing = sorted(w for w in wanted if w not in M.COLUMNS)
        check(not missing, f"the canonical names this test pins still exist in "
                           f"trz_model.COLUMNS{' - missing ' + str(missing) if missing else ''}")
        unrecognised = sorted(w for w in wanted if P.classify(w) is None)
        check(not unrecognised, f"preflight recognises every canonical column"
                                f"{' - not ' + str(unrecognised) if unrecognised else ''}")
        check(P.classify("Болнични от работодател") == "болнични",
              "„Болнични от работодател“ maps to the sick-pay concept")
        check(P.classify("Брутно възнаграждение") == "бруто",
              "„Брутно възнаграждение“ maps to gross")
        check(P.classify("Отдел") is None, "an unrelated header stays unknown")
        # Both from running the tool against test/vedomost_05_2026.xlsx, where a correct
        # real layout was reported as naming the same quantity twice.
        check(P.classify("Клас %") != P.classify("Клас сума"),
              "the class rate and the class amount are different concepts")
        check(P.classify("Извънр. часове (раб. дни)") != "отработени дни",
              "an overtime-hours column is not mistaken for days worked")
        check(P.classify("Отработени часове") != "отработени дни",
              "hours worked are not mistaken for days worked")

        # --------------------------------------------------- the clean file is silent
        good = os.path.join(tmp, "clean.xlsx")
        G.save(G.clean(), good)
        sig, data = signals_for(good)
        check(sig == set(), f"a clean workbook raises no signal at all, got {sorted(sig)}")
        text, stop = P.report(data)
        check(not stop, f"and does not block, got {stop}")
        check("не е променян" in text, "the report states the file was not modified")
        check("Лице 1" not in text, "no value from the name column reaches the report")

        # ------------------------------------------------- one planted defect at a time
        print("-" * 78)
        for shape_id, (_, expected, why) in G.SHAPES.items():
            path = G.build(shape_id, os.path.join(tmp, f"{shape_id}.xlsx"))
            sig, _ = signals_for(path)
            extra = sorted(sig - {expected})
            check(expected in sig and not extra,
                  f"{shape_id:22} -> {expected:18} ({why})"
                  + ("" if expected in sig else "  NOT RAISED")
                  + (f"  ALSO RAISED {extra}" if extra else ""))
        print("-" * 78)

        # --------------------------------------------------- missing inputs are signals
        sig, _ = signals_for(good, kid=None, group=None, tzpb=None)
        check(sig == {P.NO_KID, P.NO_TZPB},
              f"КИД and ТЗПБ absent are signals, not guesses, got {sorted(sig)}")

        # ------------------------------------------------------------ phase 2: mapping
        # A file whose columns are named the company's way: unknown without a mapping,
        # clean with one. This is the whole point of declaring the layout once.
        odd = os.path.join(tmp, "odd.xlsx")
        wb = G.clean()
        ws = wb.active
        for c in range(1, ws.max_column + 1):
            if ws.cell(G.HEADER_ROW, c).value == "Осигурителен доход":
                ws.cell(G.HEADER_ROW, c, "Дох.осиг.")
        G.save(wb, odd)
        sig, _ = signals_for(odd)
        check(P.MISSING_REQUIRED in sig,
              "a company-specific header is missing without a mapping")

        m = P.Mapping({"kid": "62", "group": 3, "tzpb": 0.4,
                       "columns": {"осиг. доход": "Дох.осиг."}})
        sig, data = signals_for(odd, mapping=m, kid=None, group=None, tzpb=None)
        check(sig == set(), f"the declared mapping resolves it, got {sorted(sig)}")
        check(data["kid"] == "62" and data["tzpb"] == 0.4,
              "the mapping supplies КИД and ТЗПБ so they need not be retyped monthly")

        typo = P.Mapping({"columns": {"осиг доходд": "Дох.осиг."}})
        sig, _ = signals_for(odd, mapping=typo)
        check(P.MAPPING_UNKNOWN_CONCEPT in sig,
              "a typo in a concept key blocks instead of silently doing nothing")

        stale = P.Mapping({"columns": {"осиг. доход": "Колона, която я няма"}})
        sig, _ = signals_for(odd, mapping=stale)
        check(P.MAPPING_COLUMN_ABSENT in sig,
              "a mapping pointing at an absent column is reported as stale")

        ign = P.Mapping({"columns": {"осиг. доход": "Дох.осиг."}, "ignore": ["Забележка"]})
        wb = G.clean()
        ws = wb.active
        ws.cell(G.HEADER_ROW, ws.max_column + 1, "Забележка")
        for c in range(1, ws.max_column + 1):
            if ws.cell(G.HEADER_ROW, c).value == "Осигурителен доход":
                ws.cell(G.HEADER_ROW, c, "Дох.осиг.")
        noted = os.path.join(tmp, "noted.xlsx")
        G.save(wb, noted)
        sig, _ = signals_for(noted, mapping=ign)
        check(P.UNKNOWN_COLUMNS not in sig,
              "a declared-ignored column stops being reported as unknown")

        # ------------------------------------------------------------ phase 3: extract
        out = os.path.join(tmp, "extract.json")
        sig, data = signals_for(good)
        doc = P.extract(data, out)
        check(os.path.exists(out), "the extract is written where asked")
        sheet = doc["sheets"][0]
        check(len(sheet["rows"]) == len(G.ROWS),
              f"every data row is extracted, got {len(sheet['rows'])}")
        check(all(r["row"] != sheet["totals_row"] for r in sheet["rows"]),
              "the totals row is not extracted as a person")
        first = sheet["rows"][0]["cells"]
        check("основна" in first
              and first["основна"]["ref"].endswith(str(sheet["rows"][0]["row"])),
              "each value carries the cell reference it came from")
        # „бруто" is the computed column: it arrives only because the fixture stores the
        # cached results, which is exactly what S10 withholds.
        check("бруто" in first and first["бруто"]["value"] == G.ROWS[0][G.HEADERS.index("БРУТО")],
              "a formula column is extracted from its cached result")
        check(doc["uncached_cells"] == 0,
              "the extract reports how many formulas had no stored result")
        raw = json.dumps(doc, ensure_ascii=False)
        check("име" not in sheet["columns"] and "Лице 1" not in raw,
              "no name is written to the extract, by column or by value")
        check(doc["kid"] == "62" and doc["tzpb"] == "0.4",
              "the extract carries the inputs the workbook does not")

        # ---------------------------------------------------------- it never writes
        before = os.path.getmtime(good), os.path.getsize(good)
        P.analyse(good)
        P.extract(data, os.path.join(tmp, "again.json"))
        check((os.path.getmtime(good), os.path.getsize(good)) == before,
              "neither scanning nor extracting touches the workbook")

        # ----------------------------------------------------- boundaries from stavki
        bounds = P.regime_boundaries()
        check(len(bounds) >= 2,
              f"period boundaries are read from stavki.md, got {len(bounds)}")
        check(any((a.month, a.day) != (1, 1) for a, _ in bounds),
              "at least one mid-year boundary is known, so K8 can be warned")

        print("=" * 78)
        if FAILURES:
            print(f"FAILED: {len(FAILURES)} check(s)")
            return 1
        print("OK: clean is silent, every planted shape defect is found exactly once, "
              "nothing is written")
        return 0
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
