# -*- coding: utf-8 -*-
"""Suite 4's checker: the formula layer, judged from the formulas alone.

Reads data_only=False and never a computed value - openpyxl stores none for a file it
wrote itself. That constraint is the honest shape of the suite: the checks below are
exactly the ones proverki.md group K states as structural ("обхватът на всяка сума",
"контролната колона проверява ли нещо", "константи, скрити във формула", "кое е
формула и кое - твърдо въведена стойност"), so no arithmetic is needed to make them.

Two of the checks are semantic (gross coverage, control tautology). The rest lean on
one generic property with real reach: FORMULA SHAPE UNIFORMITY. Normalise each
formula by stripping row digits; within one column, every row's shape must match the
column's majority shape. A literal where the others reference a parameter cell, a
stray extra term, a typed value in a formula column - all of them surface as the one
row whose shape differs, without the checker knowing a single rate.
"""
import argparse
import json
import os
import re
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

from findings import Findings                                    # noqa: E402
import generate_formula as GF                                    # noqa: E402


def shape(formula):
    """A formula with its row numbers removed: the column-invariant part."""
    return re.sub(r"(?<=[A-Z])\d+", "@", formula)


def check(xlsx, manifest, quiet=False):
    man = json.load(open(manifest, encoding="utf8")) \
        if not isinstance(manifest, dict) else manifest
    ws = openpyxl.load_workbook(xlsx, data_only=False).active
    HDR, TOTAL = man["hdr"], man["total_row"]
    rows = list(range(HDR + 1, TOTAL))
    col = {ws.cell(HDR, c).value: c for c in range(1, ws.max_column + 1)
           if ws.cell(HDR, c).value}
    F = Findings()

    def cellv(r, name):
        return ws.cell(r, col[name]).value

    day_letters = {GF._L(c) for c in GF.DAY_COLS}
    acc_letters = {GF._L(c) for c in GF.ACCRUALS}
    # (row, column) pairs a semantic check already explained: the uniformity pass
    # skips them, because one defect must surface as one finding, not two.
    explained = set()

    for r in rows:
        # --- the gross covers every accrual column, and only money ----------
        gross = cellv(r, "БРУТО")
        if isinstance(gross, str) and gross.startswith("="):
            refs = set(re.findall(r"([A-Z]+)\d+", gross))
            missing = acc_letters - refs
            if missing:
                names = [c for c in GF.ACCRUALS if GF._L(c) in missing]
                F.add("KF1_sum_omits_column", r,
                      f"БРУТО does not reference {', '.join(names)} - the column "
                      f"exists, the sum ignores it, and the defect waits for the "
                      f"first amount entered there")
                explained.add((r, "БРУТО"))
            days = refs & day_letters
            if days:
                names = [c for c in GF.DAY_COLS if GF._L(c) in days]
                F.add("KF2_days_in_money_sum", r,
                      f"БРУТО adds a day-count cell ({', '.join(names)}) - days into "
                      f"money, the first sick day becomes salary")
                explained.add((r, "БРУТО"))

        # --- the control column must be able to be non-zero -----------------
        diff = cellv(r, "Разлика")
        if isinstance(diff, str) and diff.startswith("="):
            terms = re.findall(r"([A-Z]+\d+)", diff)
            if len(terms) == 2:
                a, b = terms
                fb = ws[b].value
                if isinstance(fb, str) and re.fullmatch(rf"=\s*{a}", fb.strip()):
                    F.add("KF4_tautological_control", r,
                          f"Разлика = {a}-{b} while {b} is itself ={a} - "
                          f"algebraically zero on every row, a control that can "
                          f"detect nothing")

    # --- shape uniformity per formula column --------------------------------
    for name in GF.FORMULA_COLUMNS:
        shapes = {}
        for r in rows:
            if (r, name) in explained:
                continue
            v = cellv(r, name)
            key = shape(v) if isinstance(v, str) and v.startswith("=") \
                else f"[стойност] {v!r}"
            shapes.setdefault(key, []).append(r)
        if len(shapes) < 2:
            continue
        majority = max(shapes.values(), key=len)
        for key, where in shapes.items():
            if where is majority:
                continue
            for r in where:
                if key.startswith("[стойност]"):
                    F.add("KF3_hard_value_in_formula_column", r,
                          f"{name} is a typed value while {len(majority)} rows carry "
                          f"formulas - right today, silent and stale after the first "
                          f"change upstream")
                elif re.search(r"\d", key.replace("@", "")):
                    F.add("KF5_constant_in_formula", r,
                          f"{name} inlines a literal where the other rows reference "
                          f"the parameter cell - correct until the parameter moves, "
                          f"then wrong on exactly this row")
                else:
                    F.add("KF_shape_deviates", r,
                          f"{name} is shaped unlike the column's other rows")

    # -------------------------------------------------------------- scoring
    expected = {(HDR + 1 + idx, ident) for _, idx, ident in man["expected"]}
    found = F.keys()
    missed, extra = expected - found, found - expected
    if not quiet:
        print(f"=== {os.path.basename(xlsx)} · {man['people']} people ===")
        for f in sorted(F.items, key=lambda x: (str(x["where"]), x["id"])):
            print(f"  [row {f['where']:3}] {f['id']:32} {f['text'][:80]}")
        if missed:
            print("MISSED:", sorted(missed))
        if extra:
            print("FALSE POSITIVES:", sorted(extra))
    return dict(injected=len(expected), found=len(expected) - len(missed),
                findings=len(F.items), missed=sorted(missed), extra=sorted(extra))


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--seed", type=int, default=1)
    ap.add_argument("--quiet", action="store_true")
    a = ap.parse_args()
    xlsx, mpath, _ = GF.generate(a.seed)
    result = check(xlsx, mpath, quiet=a.quiet)
    print(f"injected {result['injected']} · found {result['found']} · "
          f"missed {len(result['missed'])} · extra {len(result['extra'])}")
    sys.exit(1 if result["missed"] or result["extra"] else 0)
