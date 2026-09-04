# -*- coding: utf-8 -*-
"""Generator of test payrolls with random data and deliberately injected defects.

    python test/generate_wide.py --seed 7

Writes `test/tmp/wide_<seed>.xlsx` and a manifest `..._manifest.json` listing
exactly what was broken. Everything is invented and derived from the seed:
names, company, salaries, days, benefit prices, month, accident-insurance rate.
Not a single figure comes from a real payroll.

Defects are injected as a mutation on a correctly computed row, with the
dependent cells recomputed the way the erring file would recompute them. The aim
is for each defect to produce one determinate set of findings - otherwise the
suite cannot tell a missed finding from a cascade.

Order of injection: file-level defects first (accident rate, applied cap),
because they touch every row, and only then the per-row ones. The reverse order
would wipe a mutation already made.

Separability. The check for the composition of the insurable income and the
taxable base works by solving which subset of the elements explains the stated
figure. The generator therefore guarantees that the subset sums of
{income in kind, threshold excess, sick pay, compensation} differ by enough -
otherwise the problem has more than one solution and the finding cannot be
located. That is not a concession by the suite: in a real file the same ambiguity
makes the conclusion unsafe, and the skill is required to say so rather than
guess.

Person and company names are Bulgarian because they are data. The rest is English.
"""
import argparse
import datetime
import itertools
import json
import os
import random
import zipfile

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import ExcelWriter

import trz_model as M
from trz_model import r2

HERE = os.path.dirname(os.path.abspath(__file__))
TMP = os.path.join(HERE, "tmp")
SEPARATION = 0.50      # minimum gap between two subset sums
# One person in five is a leaver with чл. 224 КТ compensation for unused leave. Higher
# than any real roster, on purpose: three scenarios can only land on such a row
# (F1_compensation_in_insurable, F6_compensation_out_of_taxable, and the stale-net shape
# of I1_vertical), and at one in ten F1 was injected on 17 of 300 seeds - a 25-seed run
# had a one-in-five chance of never testing it at all.
COMPENSATION_RATE = 0.20

# The same seed must give the same bytes. openpyxl stamps docProps/core.xml with the
# wall clock - save_workbook() overwrites properties.modified with now() on every call,
# so setting it beforehand changes nothing - and ZipFile dates every entry with the
# current time as well. Both are pinned to this instant instead.
FIXED_STAMP = datetime.datetime(2026, 1, 1)


class _FrozenZip(zipfile.ZipFile):
    """A ZipFile whose entries carry FIXED_STAMP instead of the time of writing.

    openpyxl adds most parts with writestr() - dated with the wall clock - and the
    worksheets with write() from a temporary file, dated with that file's mtime. Both
    routes are pinned, or the sheet part alone keeps the workbook from repeating.
    """

    def _frozen_info(self, arcname, mode):
        zinfo = zipfile.ZipInfo(arcname, date_time=FIXED_STAMP.timetuple()[:6])
        zinfo.compress_type = self.compression
        zinfo.external_attr = mode << 16
        return zinfo

    def writestr(self, zinfo_or_arcname, data, *args, **kwargs):
        if not isinstance(zinfo_or_arcname, zipfile.ZipInfo):
            # 0o600 is what ZipFile gives a bare name itself
            zinfo_or_arcname = self._frozen_info(zinfo_or_arcname, 0o600)
        super().writestr(zinfo_or_arcname, data, *args, **kwargs)

    def write(self, filename, arcname=None, *args, **kwargs):
        with open(filename, "rb") as fh:
            data = fh.read()
        # 0o100600: a regular file, as ZipInfo.from_file would record it
        super().writestr(self._frozen_info(arcname or os.path.basename(filename),
                                           0o100600), data)


def save_frozen(wb, path):
    """Write `wb` to `path` so that two runs of one seed are byte-identical.

    Used by all three generators. Goes through ExcelWriter directly rather than
    wb.save(), because save_workbook() is where openpyxl restamps the modified date.
    """
    wb.properties.created = wb.properties.modified = FIXED_STAMP
    with _FrozenZip(path, "w", zipfile.ZIP_DEFLATED, allowZip64=True) as archive:
        ExcelWriter(wb, archive).save()

# --- invented names. Deliberately unlike any real roster. ---
FIRST_M = ["Борислав", "Дарин", "Захари", "Ивайло", "Камен", "Лъчезар", "Никифор",
           "Орлин", "Тихомир", "Юлиан", "Явор", "Емил"]
FIRST_F = ["Ана", "Ваня", "Галина", "Емилия", "Жанета", "Красимира", "Магдалена",
           "Пенка", "Симона", "Христина", "Здравка", "Теодора"]
LAST = ["Аврамов", "Бакалов", "Влахов", "Гошев", "Даскалов", "Еленков", "Жеков",
        "Зидаров", "Кожухаров", "Лозанов", "Мутафчиев", "Ненов", "Орешков",
        "Пенчев", "Радулов", "Сивков", "Тошков", "Узунов", "Фандъков", "Хаджиев",
        "Цветков", "Чакъров", "Шивачев", "Янев"]
DEPARTMENTS = ["Разработка", "Внедряване", "Поддръжка", "Операции", "Администрация"]
COMPANIES = ["Тестова Дигитал", "Пробна Софтуер", "Примерна Интеграции",
             "Демонстрационна Системи", "Условна Технолоджи"]
MONTHS_BG = {6: "юни", 7: "юли", 8: "август", 9: "септември", 10: "октомври",
             11: "ноември"}


def _name(rnd, used):
    while True:
        if rnd.random() < 0.45:
            name = f"{rnd.choice(FIRST_F)} {rnd.choice(LAST)}а"
        else:
            name = f"{rnd.choice(FIRST_M)} {rnd.choice(LAST)}"
        if name not in used:
            used.add(name)
            return name


def _separable(values):
    """Do all subset sums differ by more than SEPARATION?"""
    vs = [v for v in values if v]
    sums = []
    for k in range(len(vs) + 1):
        for combo in itertools.combinations(vs, k):
            sums.append(r2(sum(combo)))
    sums.sort()
    return all(b - a > SEPARATION for a, b in zip(sums, sums[1:]))


def _usable_rows(people, cap_applicable, cap_effective):
    """How many rows let the file's practice be inferred, per element.

    A row is usable when it has accruals for work and does not sit at a cap -
    at the cap the same figure follows from many combinations and nothing is
    visible. The count matters: below three rows the practice cannot be
    established, and a defect that depends on it cannot be located - neither by
    the suite nor by a live auditor.
    """
    count = {"in_kind": 0, "excess": 0}
    for p in people:
        row = p["row"]
        work_base = r2(row["Основна за отработеното"] + row["Клас сума"]
                       + row["Бонус"] + row["Платен отпуск"])
        if work_base <= 0:
            continue
        if any(abs(row["Осигурителен доход"] - c) <= M.TOL
               for c in (cap_applicable, cap_effective)):
            continue
        if row["Карта (за сметка на работодателя)"]:
            count["in_kind"] += 1
        premium = row["Доброволно здравно осигуряване (премия)"]
        if premium and r2(premium - M.SOCIAL_EXPENSE_THRESHOLD) > 0:
            count["excess"] += 1
    return count


def _carries(row, element):
    if element == "in_kind":
        return row["Карта (за сметка на работодателя)"] > 0
    premium = row["Доброволно здравно осигуряване (премия)"]
    return bool(premium) and r2(premium - M.SOCIAL_EXPENSE_THRESHOLD) > 0


def _taxable_alternatives(before, contribution, life, elements):
    """The taxable bases the checker's OTHER candidates would produce.

    A mirror of structural_test.resolve_taxable: each element of the row moved in or out
    of the base, under the lawful relief, no relief, or both groups squeezed under one
    limit. A mutation whose base lands within SEPARATION of any of these is ambiguous by
    construction and the checker is right to say so.
    """
    deltas = {0.0}
    for v in elements:
        if v:
            deltas |= {v, -v}
    out = []
    for delta in deltas:
        b = r2(before + delta)
        out.append(r2(b - M.relief_for(b, contribution, life)))
        out.append(b)
        out.append(r2(b - min(r2(contribution + life), r2(b * M.RELIEF_LIMIT))))
    return out


def _elements(row):
    premium = row["Доброволно здравно осигуряване (премия)"]
    return [row["Карта (за сметка на работодателя)"],
            r2(max(0.0, premium - M.SOCIAL_EXPENSE_THRESHOLD)) if premium else 0.0,
            row["Болнични (работодател)"],
            row["Обезщетение чл. 224"]]


def random_inputs(rnd, norm, regime):
    """One random but plausible person."""
    kind = rnd.random()
    if kind < 0.15:
        salary = r2(rnd.uniform(regime["min_wage"], regime["min_wage"] * 1.4))
    elif kind < 0.7:
        salary = r2(rnd.uniform(1200, 2600))
    else:
        salary = r2(rnd.uniform(2600, 9000))

    pct = rnd.choice([0, 0, 0.6, 1.2, 1.8, 2.4, 3.0, 4.2, 4.8, 6.0, 7.2, 9.0])
    sick = rnd.choice([0] * 8 + [2, 3, 5])
    leave = 0 if sick else rnd.choice([0, 0, 1, 2, 3, 5, 8, 11])
    leave = min(leave, norm - sick)
    worked = norm - leave - sick

    has_card = rnd.random() < 0.6
    has_premium = rnd.random() < 0.8
    has_pension = rnd.random() < 0.20
    has_life = rnd.random() < (0.55 if has_pension else 0.22)
    return dict(
        monthly_salary=salary, seniority_pct=pct,
        days_worked=worked, days_leave=leave, days_sick=sick, days_maternity=0,
        bonus=r2(rnd.uniform(50, 400)) if rnd.random() < 0.18 else 0.0,
        # A leaver with compensation for unused leave. Rare but real - and without any
        # row carrying one, F1_compensation_in_insurable was a detection with no
        # generator: a check that had never once been allowed to fail. The rate is
        # COMPENSATION_RATE, and the comment there says why it is what it is.
        # Whole days of unused leave at the contracted daily rate with the supplement -
        # the чл. 177 base of an unchanged contract, which is what чл. 224 pays. It was
        # a uniform amount once, and the first paid run of 2.7.0 said so on every seed:
        # „не съответства на цял брой дни по никоя база" - a correct observation about
        # the fixture, not about the payroll it stands for.
        compensation_224=(r2(rnd.randint(1, 10) * salary * (1 + pct / 100.0) / norm)
                          if rnd.random() < COMPENSATION_RATE else 0.0),
        card_employer=r2(rnd.uniform(38, 72)) if has_card else 0.0,
        card_employee=r2(rnd.uniform(3, 12)) if has_card else 0.0,
        premium=r2(rnd.uniform(32.9, 44.5)) if has_premium else 0.0,
        personal_contribution=r2(rnd.uniform(20, 120)) if has_pension else 0.0,
        # Group two of чл. 19, ал. 2: доброволно здравно осигуряване and застраховки
        # „Живот“. Its own 10%, independent of the group above. Correlated with the
        # first on purpose: a row carrying BOTH is the only one where a single shared
        # 10% differs from two separate ones, so the fixture has to produce them often
        # enough for the scenario to have teeth.
        life_premium_personal=r2(rnd.uniform(40, 220)) if has_life else 0.0,
    )


def make_person(rnd, norm, regime, tzpb, policy):
    """Return (inputs, clean row) with separable elements guaranteed."""
    for _ in range(30):
        inp = random_inputs(rnd, norm, regime)
        row = M.clean_row(inp, regime, tzpb, policy, norm)
        if _separable(_elements(row)):
            row["_norm"] = norm
            return inp, row
        # nudge the benefits rather than re-rolling the whole person
        for _ in range(20):
            inp["card_employer"] = r2(rnd.uniform(38, 72)) if inp["card_employer"] else 0.0
            inp["premium"] = r2(rnd.uniform(32.9, 44.5)) if inp["premium"] else 0.0
            row = M.clean_row(inp, regime, tzpb, policy, norm)
            if _separable(_elements(row)):
                row["_norm"] = norm
                return inp, row
    inp["card_employer"] = inp["card_employee"] = 0.0
    inp["premium"] = 0.0
    row = M.clean_row(inp, regime, tzpb, policy, norm)
    row["_norm"] = norm
    return inp, row


# =====================================================================
#                             mutations
# Each returns (new row, set of expected ids) or None if the row does not
# qualify for this defect.
# =====================================================================

def _recompute_downstream(row, inp, regime, tzpb, policy, *,
                          insurable=None, taxable=None):
    """Recompute contributions, tax, net and cost after the gross or the bases
    change - the way a file whose downstream columns are formulas would."""
    gross = row["БРУТО"]
    in_kind = row["Карта (за сметка на работодателя)"]
    premium = row["Доброволно здравно осигуряване (премия)"]
    excess = r2(max(0.0, premium - M.SOCIAL_EXPENSE_THRESHOLD)) if premium else 0.0
    work_base = r2(row["Основна за отработеното"] + row["Клас сума"]
                   + row["Бонус"] + row["Платен отпуск"])
    sick_pay = row["Болнични (работодател)"]
    add_insurable, add_taxable = M.additions_for(policy, in_kind, excess, work_base)

    if insurable is None:
        insurable = r2(min(regime["max_insurable"],
                           r2(work_base + sick_pay + add_insurable)))
    row["Осигурителен доход"] = insurable

    for column, key in M.EMPLOYEE_COLUMNS:
        row[column] = r2(insurable * M.EMPLOYEE[key] / 100.0)
    employee_total = r2(sum(row[c] for c, _ in M.EMPLOYEE_COLUMNS))
    row["Лични вноски общо"] = employee_total

    if taxable is None:
        # the sick pay is inside the gross and outside the taxable base
        before = r2(gross + add_taxable - sick_pay - employee_total)
        taxable = r2(before - M.relief_for(before,
                                           row["Удръжка доброволно осиг. (лична)"],
                                           row["Удръжка застраховка Живот (лична)"]))
    row["Данъчна основа"] = taxable
    row["ДДФЛ"] = r2(taxable * M.TAX_RATE)

    row["НЕТО преди удръжки"] = r2(gross - employee_total - row["ДДФЛ"])
    row["НЕТО за изплащане"] = r2(row["НЕТО преди удръжки"]
                                  - row["Удръжка доброволно осиг. (лична)"]
                                  - row["Удръжка застраховка Живот (лична)"]
                                  - row["Удръжка карта (лична част)"])
    row["Изплатено"] = row["НЕТО за изплащане"]
    row["Разлика"] = 0.0

    row["Вноски работодател ДОО+ТЗПБ"] = r2(insurable * (M.EMPLOYER_SOCIAL + tzpb) / 100.0)
    row["ДЗПО-УПФ работодател"] = r2(insurable * M.EMPLOYER_UPF / 100.0)
    row["ЗО работодател"] = r2(insurable * M.EMPLOYER_HEALTH / 100.0)
    row["Вноски работодател общо"] = r2(row["Вноски работодател ДОО+ТЗПБ"]
                                       + row["ДЗПО-УПФ работодател"]
                                       + row["ЗО работодател"]
                                       + row["ЗО при болничен/майчинство"])
    row["Общ разход за труд"] = r2(gross + row["Вноски работодател общо"]
                                   + in_kind + premium)
    return row


def m_sum_omits_column(row, inp, regime, tzpb, policy, rnd):
    """Gross leaves out an accrual column - here the чл. 224 compensation."""
    row = dict(row)
    for _ in range(40):
        amount = r2(rnd.uniform(120, 900))
        trial = dict(row, **{"Обезщетение чл. 224": amount})
        if _separable(_elements(trial)):
            row = trial
            break
    else:
        return None
    row["БРУТО"] = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                      + row["Платен отпуск"] + row["Болнични (работодател)"])
    return _recompute_downstream(row, inp, regime, tzpb, policy), \
        {"K1_sum_omits_column"}


def m_amount_in_day_column(row, inp, regime, tzpb, policy, rnd):
    """The sick-pay amount typed into the column meant for sick days."""
    if not inp["days_sick"] or not row["Болнични (работодател)"]:
        return None
    row = dict(row)
    row["Дни болничен"] = row["Болнични (работодател)"]
    return row, {"K2_amount_in_day_column", "I5_days_do_not_reconcile"}


def m_stale_contributions(row, inp, regime, tzpb, policy, rnd):
    """Contributions are hardcoded values from an earlier period."""
    row = dict(row)
    stale = r2(row["Осигурителен доход"] * rnd.uniform(0.86, 0.95))
    if row["Осигурителен доход"] - stale < 5:
        return None
    for column, key in M.EMPLOYEE_COLUMNS:
        row[column] = r2(stale * M.EMPLOYEE[key] / 100.0)
    employee_total = r2(sum(row[c] for c, _ in M.EMPLOYEE_COLUMNS))
    row["Лични вноски общо"] = employee_total

    in_kind = row["Карта (за сметка на работодателя)"]
    premium = row["Доброволно здравно осигуряване (премия)"]
    excess = r2(max(0.0, premium - M.SOCIAL_EXPENSE_THRESHOLD)) if premium else 0.0
    work_base = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                   + row["Платен отпуск"])
    _, add_taxable = M.additions_for(policy, in_kind, excess, work_base)
    before = r2(row["БРУТО"] + add_taxable - row["Болнични (работодател)"]
                - employee_total)
    relief = M.relief_for(before, row["Удръжка доброволно осиг. (лична)"],
                          row["Удръжка застраховка Живот (лична)"])
    row["Данъчна основа"] = r2(before - relief)
    row["ДДФЛ"] = r2(row["Данъчна основа"] * M.TAX_RATE)
    row["НЕТО преди удръжки"] = r2(row["БРУТО"] - employee_total - row["ДДФЛ"])
    row["НЕТО за изплащане"] = r2(row["НЕТО преди удръжки"]
                                  - row["Удръжка доброволно осиг. (лична)"]
                                  - row["Удръжка застраховка Живот (лична)"]
                                  - row["Удръжка карта (лична част)"])
    row["Изплатено"] = row["НЕТО за изплащане"]
    return row, {"K3_stale_contributions"}


def m_control_column_blind(row, inp, regime, tzpb, policy, rnd):
    """Paid is below net while the control column still reads zero."""
    row = dict(row)
    row["Изплатено"] = r2(row["НЕТО за изплащане"] - rnd.choice([0.05, 0.13, 0.40, 1.00]))
    row["Разлика"] = 0.0
    return row, {"K4_control_column_blind"}


def m_unrounded_accrual(row, inp, regime, tzpb, policy, rnd):
    """The seniority supplement is unrounded and flows into the gross."""
    if not inp["seniority_pct"]:
        return None
    raw = row["Основна за отработеното"] * inp["seniority_pct"] / 100.0
    if abs(raw - r2(raw)) < 0.002:
        return None                     # no visible tail - nothing to detect
    row = dict(row)
    row["Клас сума"] = raw
    row["БРУТО"] = (row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                    + row["Платен отпуск"] + row["Обезщетение чл. 224"]
                    + row["Болнични (работодател)"])
    row = _recompute_downstream(row, inp, regime, tzpb, policy)
    row["Изплатено"] = r2(row["НЕТО за изплащане"])
    row["Разлика"] = r2(row["НЕТО за изплащане"] - row["Изплатено"])
    return row, {"K6_unrounded_accrual"}


def m_cost_from_net(row, inp, regime, tzpb, policy, rnd):
    """The cost of labour is computed from net after deductions."""
    deductions = r2(row["Удръжка доброволно осиг. (лична)"]
                    + row["Удръжка застраховка Живот (лична)"]
                    + row["Удръжка карта (лична част)"])
    if deductions < 1.0:
        return None
    row = dict(row)
    row["Общ разход за труд"] = r2(row["Общ разход за труд"] - deductions)
    return row, {"K7_cost_from_net"}


def m_sick_pay_out_of_insurable(row, inp, regime, tzpb, policy, rnd):
    """The first-days sick pay is left out of the insurable income.

    чл. 3, ал. 1 НЕВДПОВ puts it in, so leaving it out understates the
    contributions and т. 21 of Декларация обр. 1 by 70% of the daily base per day.
    """
    if not row["Болнични (работодател)"]:
        return None
    row = dict(row)
    in_kind = row["Карта (за сметка на работодателя)"]
    premium = row["Доброволно здравно осигуряване (премия)"]
    excess = r2(max(0.0, premium - M.SOCIAL_EXPENSE_THRESHOLD)) if premium else 0.0
    work_base = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                   + row["Платен отпуск"])
    add_insurable, _ = M.additions_for(policy, in_kind, excess, work_base)
    if r2(work_base + add_insurable + row["Болнични (работодател)"]) \
            > regime["max_insurable"] - 1.0:
        # the correct figure touches the cap: the composition is no longer
        # recoverable and the finding cannot be located, by the suite or by a
        # live auditor
        return None
    insurable = r2(work_base + add_insurable)
    if abs(insurable - row["Осигурителен доход"]) < 1.0:
        return None                      # too small to tell from rounding
    return _recompute_downstream(row, inp, regime, tzpb, policy, insurable=insurable), \
        {"F9_sick_pay_out_of_insurable"}


def m_sick_pay_in_taxable(row, inp, regime, tzpb, policy, rnd):
    """The first-days sick pay is left inside the taxable base.

    чл. 24, ал. 2, т. 14 ЗДДФЛ keeps it out, so leaving it in overtaxes the person
    by 10% of the sick pay.
    """
    if not row["Болнични (работодател)"]:
        return None
    row = dict(row)
    in_kind = row["Карта (за сметка на работодателя)"]
    premium = row["Доброволно здравно осигуряване (премия)"]
    excess = r2(max(0.0, premium - M.SOCIAL_EXPENSE_THRESHOLD)) if premium else 0.0
    work_base = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                   + row["Платен отпуск"])
    _, add_taxable = M.additions_for(policy, in_kind, excess, work_base)
    # the correct base subtracts the sick pay here; this one does not
    before = r2(row["БРУТО"] + add_taxable - row["Лични вноски общо"])
    relief = M.relief_for(before, row["Удръжка доброволно осиг. (лична)"],
                          row["Удръжка застраховка Живот (лична)"])
    return _recompute_downstream(row, inp, regime, tzpb, policy,
                                 insurable=row["Осигурителен доход"],
                                 taxable=r2(before - relief)), \
        {"F9_sick_pay_in_taxable"}


def m_sick_pay_base_wrong_side(row, inp, regime, tzpb, policy, rnd):
    """Sick pay computed on the bonus's other side of чл. 17, ал. 1 НСОРЗ.

    чл. 40, ал. 5 КСО owes 70% of the month's average daily gross, and the gross it
    means is remuneration of permanent character. Whether an uncharacterised bonus is
    part of that has two lawful answers - a one-off is in none of the seven points, pay
    determined by an applied wage system is т. 2 - and the auditor's configured reading
    picks one. Whichever it picks, the defect is the payroll using the other:

    - configured **out** (the default): the bonus is spread over the days of incapacity
      and the payment comes out too high. The direction is why it survives - nobody
      queries a payment that came out large - and it is a finding all the same, because
      Декларация обр. 1 leaves the building with the wrong number;
    - configured **in** (т. 2): the bonus is left out and the worker is short.

    One mutation, both polarities, so neither reading is the only one with teeth.
    """
    if not inp["days_sick"] or not row["Болнични (работодател)"]:
        return None
    if not inp["bonus"]:
        return None                     # without a bonus the two bases coincide
    worked = inp["days_worked"]
    if not worked:
        return None
    employer_days = min(inp["days_sick"], M.SICK_DAYS_EMPLOYER)
    # The reading the file did NOT apply - which is what makes the row wrong.
    other = 0.0 if policy.get("bonus_in_base") else row["Бонус"]
    permanent = M.permanent_work_pay(row["Основна за отработеното"], row["Клас сума"],
                                     other)
    wrong = r2(M.sick_daily_base(inp["monthly_salary"], inp["seniority_pct"],
                                 row["_norm"], permanent, worked)
               * employer_days * M.SICK_RATE)
    if abs(wrong - row["Болнични (работодател)"]) < 0.10:
        return None                     # too small to tell from rounding
    row = dict(row)
    row["Болнични (работодател)"] = wrong
    row["БРУТО"] = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                      + row["Платен отпуск"] + row["Обезщетение чл. 224"] + wrong)
    return _recompute_downstream(row, inp, regime, tzpb, policy), \
        {"F9_sick_pay_amount"}


def m_compensation_in_insurable(row, inp, regime, tzpb, policy, rnd):
    """The чл. 224 КТ compensation pulled into the insurable income.

    чл. 1, ал. 8, т. 7 НЕВДПОВ is an exhaustive list of the sums no contributions are
    due on, and чл. 224 is in it — the same statute-settled footing as the sick pay,
    opposite direction. A file that makes this mistake overcharges both sides'
    contributions and misreports т. 21 of Декларация обр. 1. The scenario existed as
    a detection with no generator since the beginning, which is exactly the state
    CONTRIBUTING warns about: a check that has never failed has not been tested.
    """
    comp = row["Обезщетение чл. 224"]
    if not comp:
        return None
    if not _separable(_elements(row)):
        return None                      # comp collides with another element's value
    insurable = r2(row["Осигурителен доход"] + comp)
    # Strictly below EVERY cap of the year, not merely the applicable one: a pull
    # that lands on a cap is pinned there by min() and the checker rightly refuses
    # to solve the composition of a capped row - the defect becomes invisible and
    # the seed fails as a miss.
    if insurable >= min(v["max_insurable"] for v in M.REGIMES.values()) - 1.0:
        return None
    return _recompute_downstream(row, inp, regime, tzpb, policy,
                                 insurable=insurable), \
        {"F1_compensation_in_insurable"}


def m_health_on_sick_days(row, inp, regime, tzpb, policy, rnd):
    """The чл. 40, ал. 1, т. 5 ЗЗО contribution computed on the wrong days.

    The base counts the days the employer's чл. 40, ал. 5 pay does NOT cover - from
    the third working day of incapacity, plus maternity. Two shapes, one id:

    - the contribution due is missing entirely (the original scenario);
    - it is charged over ALL the days of incapacity, employer-paid ones included -
      the refuted rule, paying health twice for those days, since their pay is
      insurable income and carries the full contribution there. This model itself
      did that until the adversarial review held it against т. 17 of Декларация
      обр. 1, which stavki.md had been quoting all along.
    """
    sd, md = inp["days_sick"], inp["days_maternity"]
    due = row["ЗО при болничен/майчинство"]
    shapes = []
    if due:
        shapes.append(0.0)                              # missing entirely
    if sd:
        all_days = r2(regime["min_insurable_self"] * M.HEALTH_ON_INCAPACITY / 100.0
                      * (sd + md) / row["_norm"])
        if abs(all_days - due) > 0.10:
            shapes.append(all_days)                     # the employer days charged too
    if not shapes:
        return None
    row = dict(row)
    row["ЗО при болничен/майчинство"] = rnd.choice(shapes)
    row["Вноски работодател общо"] = r2(row["Вноски работодател ДОО+ТЗПБ"]
                                       + row["ДЗПО-УПФ работодател"]
                                       + row["ЗО работодател"])
    row["Общ разход за труд"] = r2(row["БРУТО"] + row["Вноски работодател общо"]
                                   + row["Карта (за сметка на работодателя)"]
                                   + row["Доброволно здравно осигуряване (премия)"])
    return row, {"F9_health_on_sick_days"}


def _asymmetry(row, inp, regime, tzpb, policy, element, ident, usable=99):
    """Make the element visible in only one of the two bases.

    `usable` is how many rows let the practice be inferred. Below four the
    practice stops being establishable once this row is spoiled, and a live
    auditor would ask rather than conclude - so the defect is not injected.
    """
    if usable < 4:
        return None
    in_kind = row["Карта (за сметка на работодателя)"]
    premium = row["Доброволно здравно осигуряване (премия)"]
    excess = r2(max(0.0, premium - M.SOCIAL_EXPENSE_THRESHOLD)) if premium else 0.0
    value = in_kind if element == "in_kind" else excess
    if value < 1.0:
        return None
    work_base = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                   + row["Платен отпуск"])
    if work_base <= 0:
        return None
    sick_pay = row["Болнични (работодател)"]
    # if the cap is reached, the composition of the insurable income is not
    # recognisable and the defect cannot be located. The sick pay counts towards
    # the cap too - without it a row that is in fact unusable looks usable.
    if r2(work_base + sick_pay + in_kind + excess) > regime["max_insurable"] - 1.0:
        return None
    # The defect is a departure from the file's own practice in exactly ONE base -
    # not an asymmetry between the two bases, which under reading В is correct.
    in_insurable = policy["in_kind_in_bases"] if element == "in_kind" \
        else policy["excess_in_insurable"]
    add_insurable, add_taxable = M.additions_for(policy, in_kind, excess, work_base)
    row = dict(row)
    if in_insurable:
        # taken out of the insurable income only; the taxable base stays as the file's
        # practice has it
        insurable = r2(min(regime["max_insurable"],
                           r2(work_base + sick_pay + add_insurable - value)))
        row = _recompute_downstream(row, inp, regime, tzpb, policy, insurable=insurable)
    else:
        # the element is in neither base for this file, so it is added to the taxable
        # base only. The relief is limited against that same base, so that the file is
        # not inconsistent in a third respect as well - the suite measures one defect
        # at a time.
        insurable = row["Осигурителен доход"]
        employee_total = r2(sum(r2(insurable * M.EMPLOYEE[k] / 100.0) for k in M.EMPLOYEE))
        before = r2(row["БРУТО"] + add_taxable + value - sick_pay - employee_total)
        relief = M.relief_for(before, row["Удръжка доброволно осиг. (лична)"],
                              row["Удръжка застраховка Живот (лична)"])
        row = _recompute_downstream(row, inp, regime, tzpb, policy,
                                    insurable=insurable, taxable=r2(before - relief))
    return row, {ident}


def m_in_kind_asymmetry(row, inp, regime, tzpb, policy, rnd, usable=99):
    return _asymmetry(row, inp, regime, tzpb, policy, "in_kind",
                      "F10_in_kind_asymmetry", usable)


def m_excess_asymmetry(row, inp, regime, tzpb, policy, rnd, usable=99):
    return _asymmetry(row, inp, regime, tzpb, policy, "excess",
                      "F10_excess_asymmetry", usable)


def m_relief_over_limit(row, inp, regime, tzpb, policy, rnd):
    """The tax relief is applied above the percentage limit."""
    row = dict(row)
    in_kind = row["Карта (за сметка на работодателя)"]
    premium = row["Доброволно здравно осигуряване (премия)"]
    excess = r2(max(0.0, premium - M.SOCIAL_EXPENSE_THRESHOLD)) if premium else 0.0
    work_base = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                   + row["Платен отпуск"])
    if work_base <= 0:
        return None
    _, add_taxable = M.additions_for(policy, in_kind, excess, work_base)
    before = r2(row["БРУТО"] + add_taxable - row["Болнични (работодател)"]
                - row["Лични вноски общо"])
    # The limit is ignored altogether: every withheld amount is deducted in full,
    # whichever group it belongs to. That is the defect - not the two-group split.
    life = row["Удръжка застраховка Живот (лична)"]
    for _ in range(20):
        contribution = r2(before * M.RELIEF_LIMIT + rnd.uniform(30, 150))
        applied = r2(contribution + life)
        # The base this defect produces must not coincide with the base another of the
        # checker's candidates would produce: at seed 2249 „the whole deduction without
        # the limit" and „the card outside the base plus the lawful relief" landed two
        # cents apart, the checker saw two explanations and reported
        # F6_taxable_unexplained instead.
        taxable = r2(before - applied)
        if all(abs(taxable - alt) > SEPARATION
               for alt in _taxable_alternatives(before, contribution, life, _elements(row))):
            break
    else:
        return None
    row["Удръжка доброволно осиг. (лична)"] = contribution
    row["Данъчна основа"] = r2(before - applied)
    row["ДДФЛ"] = r2(row["Данъчна основа"] * M.TAX_RATE)
    row["НЕТО преди удръжки"] = r2(row["БРУТО"] - row["Лични вноски общо"] - row["ДДФЛ"])
    row["НЕТО за изплащане"] = r2(row["НЕТО преди удръжки"] - contribution - life
                                  - row["Удръжка карта (лична част)"])
    row["Изплатено"] = row["НЕТО за изплащане"]
    return row, {"F7_relief_over_limit"}


def m_relief_not_applied(row, inp, regime, tzpb, policy, rnd):
    """A personal contribution is withheld, but it reduces no taxable base.

    чл. 19, ал. 2 във вр. с чл. 42, ал. 3 ЗДДФЛ reduces the monthly base by personal
    premiums the employer withholds. Withholding them and then taxing as though they
    had not been withheld overtaxes the person every month.

    This one is the opposite of most defects in the suite: it leaves the file perfectly
    self-consistent. The base is exactly „облагаем доход − лични вноски“, every row is
    treated the same way and no control column moves. Nothing contradicts anything -
    which is why it survives in real payrolls and why the checker has to know the
    relief was due rather than infer it from an inconsistency.
    """
    deduction = r2(row["Удръжка доброволно осиг. (лична)"]
                   + row["Удръжка застраховка Живот (лична)"])
    if not deduction:
        return None
    row = dict(row)
    in_kind = row["Карта (за сметка на работодателя)"]
    premium = row["Доброволно здравно осигуряване (премия)"]
    excess = r2(max(0.0, premium - M.SOCIAL_EXPENSE_THRESHOLD)) if premium else 0.0
    work_base = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                   + row["Платен отпуск"])
    if work_base <= 0:
        return None
    _, add_taxable = M.additions_for(policy, in_kind, excess, work_base)
    before = r2(row["БРУТО"] + add_taxable - row["Болнични (работодател)"]
                - row["Лични вноски общо"])
    applied = M.relief_for(before, row["Удръжка доброволно осиг. (лична)"],
                           row["Удръжка застраховка Живот (лична)"])
    if applied < 1.0:
        return None                      # too small to tell from rounding
    return _recompute_downstream(row, inp, regime, tzpb, policy,
                                 insurable=row["Осигурителен доход"],
                                 taxable=before), \
        {"F7_relief_not_applied"}


def m_relief_combined_limit(row, inp, regime, tzpb, policy, rnd):
    """Both groups of the чл. 19 relief capped against ONE shared 10%.

    Чл. 19, ал. 2 ЗДДФЛ gives two independent allowances - one for допълнително
    доброволно осигуряване, one for доброволно здравно осигуряване and премии по
    договори за застраховки „Живот“ - each up to 10% of the same чл. 42, ал. 2 base.
    A payroll that adds the two together and caps the total at a single 10% relieves
    less than it should and overtaxes the person.

    This is the error the reference file itself carried until 31.08.2026, which is why
    it is worth a scenario: it is the shape a reader arrives at from „до общо 10%“.
    """
    pension = row["Удръжка доброволно осиг. (лична)"]
    life = row["Удръжка застраховка Живот (лична)"]
    if not pension or not life:
        return None                      # needs both groups to be distinguishable
    row = dict(row)
    in_kind = row["Карта (за сметка на работодателя)"]
    premium = row["Доброволно здравно осигуряване (премия)"]
    excess = r2(max(0.0, premium - M.SOCIAL_EXPENSE_THRESHOLD)) if premium else 0.0
    work_base = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                   + row["Платен отпуск"])
    if work_base <= 0:
        return None
    _, add_taxable = M.additions_for(policy, in_kind, excess, work_base)
    before = r2(row["БРУТО"] + add_taxable - row["Болнични (работодател)"]
                - row["Лични вноски общо"])
    due = M.relief_for(before, pension, life)
    combined = r2(min(r2(pension + life), r2(before * M.RELIEF_LIMIT)))
    if due - combined < 1.0:
        return None                      # the shared cap does not bind - nothing to see
    return _recompute_downstream(row, inp, regime, tzpb, policy,
                                 insurable=row["Осигурителен доход"],
                                 taxable=r2(before - combined)), \
        {"F7_relief_combined_limit"}


def m_seniority_on_gross(row, inp, regime, tzpb, policy, rnd):
    """The seniority supplement is computed on a wider base than the salary."""
    pct = inp["seniority_pct"]
    if not pct:
        return None
    delta = r2((row["Платен отпуск"] + row["Бонус"]) * pct / 100.0)
    if delta < 0.10:                     # without leave and bonus there is no gap
        return None
    row = dict(row)
    wider = r2(row["Основна за отработеното"] + row["Платен отпуск"] + row["Бонус"])
    row["Клас сума"] = r2(wider * pct / 100.0)
    row["БРУТО"] = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                      + row["Платен отпуск"] + row["Обезщетение чл. 224"]
                      + row["Болнични (работодател)"])
    return _recompute_downstream(row, inp, regime, tzpb, policy), \
        {"C2_seniority_on_gross"}


def m_leave_without_seniority(row, inp, regime, tzpb, policy, rnd):
    """Paid leave is computed without the seniority uplift."""
    if not inp["seniority_pct"] or not inp["days_leave"]:
        return None
    daily = inp["monthly_salary"] / row["_norm"]
    without = r2(daily * inp["days_leave"])
    if abs(without - row["Платен отпуск"]) < 0.10:
        return None
    row = dict(row)
    row["Платен отпуск"] = without
    row["БРУТО"] = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                      + row["Платен отпуск"] + row["Обезщетение чл. 224"]
                      + row["Болнични (работодател)"])
    return _recompute_downstream(row, inp, regime, tzpb, policy), \
        {"E3_leave_without_seniority"}


def m_days_do_not_reconcile(row, inp, regime, tzpb, policy, rnd):
    """The sick days are not recorded although the amount is accrued."""
    if not inp["days_sick"]:
        return None
    row = dict(row)
    row["Дни болничен"] = 0
    return row, {"I5_days_do_not_reconcile"}


DEDUCTION_COLUMNS = ("Удръжка доброволно осиг. (лична)",
                     "Удръжка застраховка Живот (лична)", "Удръжка карта (лична част)")


def m_net_not_refreshed(row, inp, regime, tzpb, policy, rnd):
    """The net columns do not follow the rest of the row. Two shapes, one id (I1).

    - stale: a bonus or a чл. 224 compensation was added to the row late, the gross,
      the contributions and the tax were recomputed - and „НЕТО преди удръжки“ and
      everything below it are pasted values from before the addition;
    - unwired: a deduction column is withheld on the row but was never subtracted in
      the net formula, so „НЕТО за изплащане“ is high by exactly that column.

    Both are the shape of a payslip whose bottom half stopped following its top half,
    which is what I1 in proverki.md - the vertical reconciliation - exists to catch.
    Until this mutation existed the two I1 branches in structural_test had never fired.
    """
    late = ("bonus" if inp["bonus"] else
            "compensation_224" if inp["compensation_224"] else None)
    shapes = (["stale"] if late else []) \
        + (["unwired"] if any(row[c] >= 1.0 for c in DEDUCTION_COLUMNS) else [])
    if not shapes:
        return None
    row = dict(row)
    if rnd.choice(shapes) == "stale":
        earlier = M.clean_row(dict(inp, **{late: 0.0}), regime, tzpb, policy, row["_norm"])
        if abs(earlier["НЕТО преди удръжки"] - row["НЕТО преди удръжки"]) < 1.0:
            return None                  # too small to tell from rounding
        row["НЕТО преди удръжки"] = earlier["НЕТО преди удръжки"]
        row["НЕТО за изплащане"] = earlier["НЕТО за изплащане"]
    else:
        skipped = rnd.choice([c for c in DEDUCTION_COLUMNS if row[c] >= 1.0])
        row["НЕТО за изплащане"] = r2(row["НЕТО за изплащане"] + row[skipped])
    row["Изплатено"] = row["НЕТО за изплащане"]
    row["Разлика"] = 0.0
    return row, {"I1_vertical"}


def m_tax_not_from_base(row, inp, regime, tzpb, policy, rnd):
    """The tax does not follow the taxable base the same row states (F6).

    Two shapes, one id: the tax formula points at the base BEFORE the чл. 19 relief
    while the base column shows the relieved figure (the relief is granted in one
    column and taxed away in the next); or the tax is a pasted value from an earlier
    version of the row. The net follows the wrong tax, as a file's formulas would, so
    the vertical reconciliation holds and the only thing wrong is the tax itself.
    """
    taxable = row["Данъчна основа"]
    if taxable <= 0:
        return None
    in_kind = row["Карта (за сметка на работодателя)"]
    premium = row["Доброволно здравно осигуряване (премия)"]
    excess = r2(max(0.0, premium - M.SOCIAL_EXPENSE_THRESHOLD)) if premium else 0.0
    work_base = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                   + row["Платен отпуск"])
    _, add_taxable = M.additions_for(policy, in_kind, excess, work_base)
    before = r2(row["БРУТО"] + add_taxable - row["Болнични (работодател)"]
                - row["Лични вноски общо"])
    shapes = [r2(taxable * M.TAX_RATE * rnd.uniform(0.80, 0.93))]         # pasted
    if r2(before - taxable) * M.TAX_RATE >= 1.0:
        shapes.append(r2(before * M.TAX_RATE))                            # relief taxed
    wrong = rnd.choice(shapes)
    if abs(wrong - row["ДДФЛ"]) < 1.0:
        return None                      # too small to tell from rounding
    row = dict(row)
    row["ДДФЛ"] = wrong
    row["НЕТО преди удръжки"] = r2(row["БРУТО"] - row["Лични вноски общо"] - wrong)
    row["НЕТО за изплащане"] = r2(row["НЕТО преди удръжки"]
                                  - sum(row[c] for c in DEDUCTION_COLUMNS))
    row["Изплатено"] = row["НЕТО за изплащане"]
    row["Разлика"] = 0.0
    return row, {"F6_tax_amount"}


def m_base_from_other_salary(row, inp, regime, tzpb, policy, rnd):
    """The row is computed from a salary that is not the contracted one (A6).

    A stale salary before a raise, or a raise applied in the payroll before the annex
    exists - proverki.md counts both directions, the first as a violation, the second
    as a risk. The whole row follows the wrong salary consistently, so nothing else on
    it disagrees with itself: the only witness is the contract. Rows with leave or
    sick days are skipped, because there the same wrong salary would also move the
    leave and the sick pay, and one defect would be reported three times.
    """
    if inp["days_leave"] or inp["days_sick"] or not inp["days_worked"]:
        return None
    if rnd.random() < 0.5:
        other = r2(inp["monthly_salary"] * rnd.uniform(0.85, 0.95))
        if other < regime["min_wage"]:
            return None                  # would add a minimum-wage finding to this one
    else:
        other = r2(inp["monthly_salary"] * rnd.uniform(1.05, 1.15))
    new = M.clean_row(dict(inp, monthly_salary=other), regime, tzpb, policy,
                      row["_norm"])
    if abs(new["Основна за отработеното"] - row["Основна за отработеното"]) < 1.0:
        return None
    # A raise must not carry the row onto a cap: capped, its composition stops being
    # solvable, it leaves the sample the file's practice is inferred from, and a
    # defect already injected elsewhere can lose the voters it was gated on.
    if new["Осигурителен доход"] >= min(v["max_insurable"]
                                       for v in M.REGIMES.values()) - 1.0:
        return None
    return new, {"A6_base_vs_contract"}


def m_insurable_unexplained(row, inp, regime, tzpb, policy, rnd):
    """The insurable income is a figure no composition of the row reaches (F1).

    A pasted value from another month: the contributions on both sides follow it, so
    K3 and F5 hold, and the taxable base follows the contributions, so F6 holds. What
    does not hold is the composition itself - the gap to the accruals is not any one
    element, present or absent, and the checker must say so rather than pick the
    nearest element. The guard below keeps the gap away from every element's value,
    or the finding would carry that element's id instead.
    """
    in_kind = row["Карта (за сметка на работодателя)"]
    premium = row["Доброволно здравно осигуряване (премия)"]
    excess = r2(max(0.0, premium - M.SOCIAL_EXPENSE_THRESHOLD)) if premium else 0.0
    work_base = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                   + row["Платен отпуск"])
    if work_base <= 0:
        return None
    add_insurable, _ = M.additions_for(policy, in_kind, excess, work_base)
    expected = r2(work_base + row["Болнични (работодател)"] + add_insurable)
    caps = [v["max_insurable"] for v in M.REGIMES.values()]
    if expected >= min(caps) - 1.0:
        return None                      # at a cap the composition is not solvable
    for _ in range(20):
        stale = r2(expected * rnd.uniform(0.86, 0.95))
        gap = r2(expected - stale)
        if gap < 1.0 or any(abs(stale - c) < 1.0 for c in caps):
            continue
        if any(abs(gap - v) <= SEPARATION for v in _elements(row) if v):
            continue                     # one element would explain it - another id
        break
    else:
        return None
    return _recompute_downstream(row, inp, regime, tzpb, policy, insurable=stale), \
        {"F1_insurable_unexplained"}


def _taxable_explanations(row, policy):
    """Every taxable base a catalogued single deviation would give this row.

    The generator's side of the same determinacy guarantee _separable gives the
    insurable income: a mutation of the taxable base must land at least SEPARATION
    away from each of these, or the checker cannot tell which deviation it is looking
    at - and in a real file that same coincidence would make the conclusion unsafe.
    Keyed by the scenario id whose deviation produces the figure; „clean“ is the
    correct base.
    """
    in_kind = row["Карта (за сметка на работодателя)"]
    premium = row["Доброволно здравно осигуряване (премия)"]
    excess = r2(max(0.0, premium - M.SOCIAL_EXPENSE_THRESHOLD)) if premium else 0.0
    work_base = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                   + row["Платен отпуск"])
    sick_pay, comp = row["Болнични (работодател)"], row["Обезщетение чл. 224"]
    pension = row["Удръжка доброволно осиг. (лична)"]
    life = row["Удръжка застраховка Живот (лична)"]
    _, add_taxable = M.additions_for(policy, in_kind, excess, work_base)
    before = r2(row["БРУТО"] + add_taxable - sick_pay - row["Лични вноски общо"])

    def relieved(b):
        return r2(b - M.relief_for(b, pension, life))

    out = {"clean": relieved(before),
           "F7_relief_over_limit": r2(before - r2(pension + life)),
           "F7_relief_not_applied": before,
           "F7_relief_combined_limit": r2(before - min(r2(pension + life),
                                                       r2(before * M.RELIEF_LIMIT)))}
    if sick_pay:
        out["F9_sick_pay_in_taxable"] = relieved(r2(before + sick_pay))
    if comp:
        out["F6_compensation_out_of_taxable"] = relieved(r2(before - comp))
    # The contested elements flipped against the file's own practice for this base.
    if in_kind:
        sign = -1.0 if policy["in_kind_in_bases"] else 1.0
        out["F10_in_kind_asymmetry"] = relieved(r2(before + sign * in_kind))
    if excess:
        sign = -1.0 if policy["excess_in_taxable"] else 1.0
        out["F10_excess_asymmetry"] = relieved(r2(before + sign * excess))
    return out


def m_taxable_unexplained(row, inp, regime, tzpb, policy, rnd):
    """The taxable base is a figure no catalogued deviation reaches (F6).

    The base is a pasted value from another month; the tax and the net follow it.
    Nothing about the gross, the contributions or the relief explains the figure, and
    the honest finding is exactly that - „does not follow from the gross minus the
    contributions; none of the known deviations fits“ - rather than the nearest
    named defect.
    """
    work_base = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                   + row["Платен отпуск"])
    if work_base <= 0:
        return None
    explanations = _taxable_explanations(row, policy)
    clean = explanations.pop("clean")
    if clean <= 0:
        return None
    for _ in range(20):
        stale = r2(clean * rnd.uniform(0.86, 0.95))
        if clean - stale < 1.0:
            continue
        if any(abs(stale - v) <= SEPARATION for v in explanations.values()):
            continue                     # a named deviation would explain it instead
        break
    else:
        return None
    return _recompute_downstream(row, inp, regime, tzpb, policy,
                                 insurable=row["Осигурителен доход"], taxable=stale), \
        {"F6_taxable_unexplained"}


def m_compensation_out_of_taxable(row, inp, regime, tzpb, policy, rnd):
    """The чл. 224 КТ compensation left out of the taxable base (F6).

    The mirror of F1_compensation_in_insurable: a file that reads „обезщетение“ as
    exempt. чл. 24, ал. 2, т. 8 ЗДДФЛ lists the compensations that are, and чл. 224
    is not among them - the sum is taxable, and leaving it out undertaxes the person
    by 10% of it. Only a row that carries the compensation qualifies, and only when no
    other catalogued deviation lands on the same figure.
    """
    if not row["Обезщетение чл. 224"]:
        return None
    work_base = r2(row["Основна за отработеното"] + row["Клас сума"] + row["Бонус"]
                   + row["Платен отпуск"])
    if work_base <= 0:
        return None
    explanations = _taxable_explanations(row, policy)
    target = explanations.pop("F6_compensation_out_of_taxable")
    if any(abs(target - v) <= SEPARATION for v in explanations.values()):
        return None                      # another deviation reaches the same figure
    return _recompute_downstream(row, inp, regime, tzpb, policy,
                                 insurable=row["Осигурителен доход"], taxable=target), \
        {"F6_compensation_out_of_taxable"}


ROW_MUTATIONS = [
    ("K1_sum_omits_column", m_sum_omits_column),
    ("K2_amount_in_day_column", m_amount_in_day_column),
    ("K3_stale_contributions", m_stale_contributions),
    ("K4_control_column_blind", m_control_column_blind),
    ("K6_unrounded_accrual", m_unrounded_accrual),
    ("K7_cost_from_net", m_cost_from_net),
    ("F9_sick_pay_out_of_insurable", m_sick_pay_out_of_insurable),
    ("F9_sick_pay_in_taxable", m_sick_pay_in_taxable),
    ("F9_sick_pay_amount", m_sick_pay_base_wrong_side),
    ("F9_health_on_sick_days", m_health_on_sick_days),
    ("F1_compensation_in_insurable", m_compensation_in_insurable),
    ("F10_in_kind_asymmetry", m_in_kind_asymmetry),
    ("F10_excess_asymmetry", m_excess_asymmetry),
    ("F7_relief_over_limit", m_relief_over_limit),
    ("F7_relief_not_applied", m_relief_not_applied),
    ("F7_relief_combined_limit", m_relief_combined_limit),
    ("C2_seniority_on_gross", m_seniority_on_gross),
    ("E3_leave_without_seniority", m_leave_without_seniority),
    ("I5_days_do_not_reconcile", m_days_do_not_reconcile),
    ("I1_vertical", m_net_not_refreshed),
    ("F6_tax_amount", m_tax_not_from_base),
    ("A6_base_vs_contract", m_base_from_other_salary),
    ("F1_insurable_unexplained", m_insurable_unexplained),
    ("F6_taxable_unexplained", m_taxable_unexplained),
    ("F6_compensation_out_of_taxable", m_compensation_out_of_taxable),
]

# Defects whose localisation goes through the file's practice for the benefits.
NEEDS_PRACTICE = ("F9_sick_pay_out_of_insurable", "F1_compensation_in_insurable",
                  "F1_insurable_unexplained",
                  "F10_in_kind_asymmetry",
                  "F10_excess_asymmetry", "F9_sick_pay_in_taxable",
                  "F6_taxable_unexplained", "F6_compensation_out_of_taxable",
                  "F7_relief_over_limit", "F7_relief_not_applied",
                  "F7_relief_combined_limit")
# Of those, only these spoil the sample the practice is inferred from.
SPOILS_SAMPLE = ("F9_sick_pay_out_of_insurable", "F1_compensation_in_insurable",
                 "F1_insurable_unexplained",
                 "F10_in_kind_asymmetry",
                 "F10_excess_asymmetry")


# =====================================================================


def generate(seed, month=None, year=2026, bonus_in_base=None):
    """Build one payroll. `bonus_in_base` pins the configured reading of чл. 17, ал. 1
    for an uncharacterised bonus column instead of drawing it from the seed - the skill
    eval needs it pinned, because a cloned skill has no plugin setting and applies the
    documented default."""
    rnd = random.Random(seed)
    month = month or rnd.choice([6, 7, 8, 9, 10, 11])
    norm = M.working_days(year, month)
    # A year the reference file has no thresholds for is built the way a real payroll is
    # built every January: by copying last year's file forward, so it carries the last
    # published regime. Whether those thresholds still apply is exactly what the skill
    # cannot know, and the whole point of the fixture is that it must not guess.
    rates_known = year in M.RATES_KNOWN_YEARS
    regime_id = M.regime_for(year, month) if rates_known else "H2"
    regime = M.REGIMES[regime_id]
    tzpb = rnd.choice([0.4, 0.5, 0.7, 1.1])
    # One of the three readings of the excess, applied to the whole file. В is the
    # asymmetric one and must produce no finding when it is applied consistently.
    reading = rnd.choice(list(M.EXCESS_READINGS))
    policy = dict(bonus_in_base=(rnd.random() < 0.5 if bonus_in_base is None
                                 else bool(bonus_in_base)),
                  in_kind_in_bases=rnd.random() < 0.5,
                  excess_in_insurable=M.EXCESS_READINGS[reading]["insurable"],
                  excess_in_taxable=M.EXCESS_READINGS[reading]["taxable"],
                  excess_reading=reading)
    company = rnd.choice(COMPANIES)
    n = rnd.randint(9, 15)

    # --- file-level defects are decided BEFORE the rows ---------------------
    file_defects = []
    tzpb_effective = tzpb
    if rnd.random() < 0.45:
        candidate = round(tzpb - rnd.choice([0.1, 0.2, 0.3]), 2)
        if candidate >= 0.1:
            tzpb_effective = candidate
            file_defects.append("F5_tzpb_below_due")

    cap_effective = regime["max_insurable"]
    # Applying "the cap from the other half-year" presupposes a published cap for this
    # year to be wrong about. With no rates there is nothing to be wrong against, and
    # the finding the skill owes is that it cannot tell - not a violation.
    if rates_known and rnd.random() < 0.35:
        cap_effective = M.REGIMES[M.other_regime_id(regime_id)]["max_insurable"]
        file_defects.append("B4_cap_from_wrong_period")
    regime_effective = dict(regime, max_insurable=cap_effective)

    used = set()
    people = []
    for _ in range(n):
        inp, row = make_person(rnd, norm, regime_effective, tzpb_effective, policy)
        people.append(dict(name=_name(rnd, used), department=rnd.choice(DEPARTMENTS),
                           inputs=inp, row=row))

    # one person on maternity leave for the whole month: no accruals, health only
    if n >= 10:
        inp = dict(monthly_salary=r2(rnd.uniform(1300, 2400)),
                   seniority_pct=rnd.choice([0, 1.2, 2.4]),
                   days_worked=0, days_leave=0, days_sick=0, days_maternity=norm,
                   bonus=0.0, compensation_224=0.0, card_employer=0.0,
                   card_employee=0.0, premium=r2(rnd.uniform(32.9, 44.5)),
                   personal_contribution=0.0, life_premium_personal=0.0)
        row = M.clean_row(inp, regime_effective, tzpb_effective, policy, norm)
        row["_norm"] = norm
        people.append(dict(name=_name(rnd, used), department=rnd.choice(DEPARTMENTS),
                           inputs=inp, row=row))
        rnd.shuffle(people)

    # B4 is a finding only if the wrong cap is actually visible somewhere
    if "B4_cap_from_wrong_period" in file_defects:
        visible = any(p["row"]["Осигурителен доход"] > regime["max_insurable"] + M.TOL
                      or (abs(p["row"]["Осигурителен доход"] - cap_effective) < M.TOL
                          and cap_effective < regime["max_insurable"])
                      for p in people)
        if not visible:
            file_defects.remove("B4_cap_from_wrong_period")

    # --- per-row defects ----------------------------------------------------
    pristine = {i: dict(p["row"]) for i, p in enumerate(people)}
    expected = []
    free = list(range(len(people)))
    rnd.shuffle(free)
    candidates = ROW_MUTATIONS[:]
    rnd.shuffle(candidates)
    # Six to eleven defects per file. It was five to nine with nineteen candidate
    # mutations; six more candidates at the same draw would have cut every scenario's
    # rate by a quarter (K1 from 0.54 to 0.40 per seed, F7_relief_combined_limit from
    # 0.12 to 0.06). This keeps the existing scenarios about where they were and the
    # 300-seed suite green with zero false positives.
    how_many = rnd.randint(6, 11)
    injected = 0
    spoiled = {"in_kind": 0, "excess": 0}      # rows whose vote is already wrong
    for ident, fn in candidates:
        if injected >= how_many or not free:
            break
        # recomputed per defect: every mutation already made can change which
        # row sits at a cap and which does not
        usable = _usable_rows(people, regime["max_insurable"], cap_effective)
        for idx in list(free):
            person = people[idx]
            if ident in NEEDS_PRACTICE:
                # The practice for every element the row carries must stay
                # establishable AFTER the mutation. Hence a fourth usable row:
                # the spoiled row either drops out of the sample (its figure no
                # longer matches any clean combination) or votes the other way.
                # Either way three clean rows must remain.
                if any(_carries(person["row"], el) and usable[el] < 4 + spoiled[el]
                       for el in ("in_kind", "excess")):
                    continue
            if ident.startswith("F10_"):
                el = "in_kind" if "in_kind" in ident else "excess"
                result = fn(person["row"], person["inputs"], regime_effective,
                            tzpb_effective, policy, rnd, usable=usable[el])
            else:
                result = fn(person["row"], person["inputs"], regime_effective,
                            tzpb_effective, policy, rnd)
            if result is None:
                continue
            new_row, ids = result
            new_row["_norm"] = norm
            if ident in SPOILS_SAMPLE:
                for el in ("in_kind", "excess"):
                    if _carries(person["row"], el):
                        spoiled[el] += 1
            person["row"] = new_row
            person["defects"] = sorted(ids)
            free.remove(idx)
            expected += [["row", idx, i] for i in sorted(ids)]
            injected += 1
            break

    # A later mutation can pull the only row that made the wrong cap visible back
    # under the applicable cap - the file then shows B4 nowhere while the manifest
    # still expects it, and the seed fails as a MISSED B4 on whoever runs it: a latent
    # flake that blames an innocent change. The visibility gate above ran over CLEAN
    # rows; re-check over the mutated ones and, if visibility is gone, give it back by
    # reverting mutations (the file-level defect was decided first; the row defects
    # have every other seed to live in).
    if "B4_cap_from_wrong_period" in file_defects:
        def b4_visible():
            return any(
                p["row"]["Осигурителен доход"] > regime["max_insurable"] + M.TOL
                or (abs(p["row"]["Осигурителен доход"] - cap_effective) < M.TOL
                    and cap_effective < regime["max_insurable"])
                for p in people)
        mutated = [i for i, p in enumerate(people) if p.get("defects")]
        while not b4_visible() and mutated:
            i = mutated.pop()
            people[i]["row"] = dict(pristine[i], _norm=norm)
            people[i]["defects"] = []
            expected = [e for e in expected if not (e[0] == "row" and e[1] == i)]
        # Reverting every mutation returns the exact rows the first gate approved,
        # so visibility cannot still be missing here.
        assert b4_visible(), "B4 visibility lost even on pristine rows"

    # The practice gate in the loop ran at each mutation's turn, over rows that later
    # mutations could still change: a K1 that blanks a figure or an A6 that moves the
    # base takes a row out of the checker's sample without being counted as spoiling
    # it. At 3000 seeds two F1_insurable_unexplained rows were left with a practice the
    # checker could no longer infer - a missed finding that blamed nobody. Re-check over
    # the FINAL rows: a NEEDS_PRACTICE defect needs three clean usable rows per element
    # its row carries, beyond the rows already spoiled; otherwise it is reverted and the
    # scenario lives in another seed.
    # Which mutations take a row out of the checker's sample: the ones that move an
    # element between the bases (SPOILS_SAMPLE) and the two that leave the insurable
    # income unexplained by the row's own elements - a gross that skips a column, a base
    # computed from another salary. Every other mutation recomputes the row consistently,
    # so it still votes; excluding all mutated rows halved the coverage of every
    # NEEDS_PRACTICE scenario (F1_compensation_in_insurable 28 -> 5 at 300 seeds).
    BREAKS_INSURABLE = set(SPOILS_SAMPLE) | {"K1_sum_omits_column", "A6_base_vs_contract"}

    def unsupported():
        intact = [p for p in people if not (set(p.get("defects") or ()) & BREAKS_INSURABLE)]
        clean = _usable_rows(intact, regime["max_insurable"], cap_effective)
        for e in expected:
            if e[0] != "row" or e[2] not in NEEDS_PRACTICE:
                continue
            # Three intact rows, the same bar the loop's gate set (four usable including
            # the row about to be spoiled). `intact` already leaves the spoiled rows
            # out, so they are not counted again here.
            if any(_carries(people[e[1]]["row"], el) and clean[el] < 3
                   for el in ("in_kind", "excess")):
                return e[1]
        return None

    while (i := unsupported()) is not None:
        if any(d in SPOILS_SAMPLE for d in people[i]["defects"]):
            for el in ("in_kind", "excess"):
                if _carries(pristine[i], el):
                    spoiled[el] -= 1
        people[i]["row"] = dict(pristine[i], _norm=norm)
        people[i]["defects"] = []
        expected = [e for e in expected if not (e[0] == "row" and e[1] == i)]
        free.append(i)

    # --- write --------------------------------------------------------------
    os.makedirs(TMP, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month:02d}-{year}"
    ws["A1"] = f'ВЕДОМОСТ ЗА РАБОТНИ ЗАПЛАТИ — "{company}" ЕООД'
    ws["A2"] = (f"Месец: {MONTHS_BG[month]} {year} г.  |  Работни дни: {norm}  |  "
                f"Валута: EUR  |  ЕИК: 000000000 (тестов)")
    ws["A3"] = f"Икономическа дейност: тестова  |  ТЗПБ по КИД: {tzpb}%"
    ws["A1"].font = Font(bold=True, size=12)

    HDR = 5
    for i, column in enumerate(M.COLUMNS, start=1):
        cell = ws.cell(row=HDR, column=i, value=column)
        cell.font = Font(bold=True, size=8)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for offset, person in enumerate(people):
        r = HDR + 1 + offset
        ws.cell(row=r, column=M.COL["№"], value=offset + 1)
        ws.cell(row=r, column=M.COL["Име"], value=person["name"])
        ws.cell(row=r, column=M.COL["Отдел"], value=person["department"])
        for column in M.COLUMNS:
            if column in ("№", "Име", "Отдел"):
                continue
            v = person["row"].get(column, 0)
            ws.cell(row=r, column=M.COL[column],
                    value=(v if v else (0 if column in M.DAY_COLUMNS else None)))

    total_row = HDR + 1 + len(people)
    ws.cell(row=total_row, column=M.COL["Име"], value="ОБЩО").font = Font(bold=True)
    for column in M.SUMMED_COLUMNS:
        s = r2(sum(p["row"].get(column, 0) or 0 for p in people))
        ws.cell(row=total_row, column=M.COL[column], value=s).font = Font(bold=True)

    if rnd.random() < 0.4:
        column = rnd.choice(["Карта (за сметка на работодателя)",
                             "Доброволно здравно осигуряване (премия)",
                             "Удръжка карта (лична част)"])
        s = r2(sum(p["row"].get(column, 0) or 0 for p in people))
        if s > 0:
            file_defects.append("K5_total_not_sum")
            ws.cell(row=total_row, column=M.COL[column],
                    value=r2(s + len(people) * 0.004 + 0.02))

    for i, column in enumerate(M.COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(9, min(15, len(column) // 2 + 6))

    path = os.path.join(TMP, f"wide_{seed}.xlsx")
    save_frozen(wb, path)

    expected += [["file", None, i] for i in file_defects]

    manifest = dict(
        seed=seed, file=os.path.basename(path), sheet=ws.title,
        year=year, month=month, norm_days=norm, regime=regime_id,
        rates_known=rates_known,
        max_insurable=regime["max_insurable"],
        # The applicable rates an auditor legitimately has include both halves of the
        # year - the checker needs the other cap to recognise B4 without assuming the
        # regime table holds exactly two entries forever.
        other_max_insurable=M.REGIMES[M.other_regime_id(regime_id)]["max_insurable"],
        min_insurable_self=regime["min_insurable_self"],
        tzpb_due=tzpb, policy=policy, hdr=HDR, total_row=total_row,
        people=[dict(row=HDR + 1 + i, name=p["name"], inputs=p["inputs"],
                     defects=p.get("defects", [])) for i, p in enumerate(people)],
        expected=expected,
    )
    manifest_path = os.path.join(TMP, f"wide_{seed}_manifest.json")
    with open(manifest_path, "w", encoding="utf8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=1)
    return path, manifest_path, manifest


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--seed", type=int, default=1)
    ap.add_argument("--month", type=int, default=None, choices=[6, 7, 8, 9, 10, 11])
    ap.add_argument("--year", type=int, default=2026,
                    help="a year outside RATES_KNOWN_YEARS builds the refusal fixture")
    a = ap.parse_args()
    path, manifest_path, man = generate(a.seed, a.month, a.year)
    print(f"written:  {path}")
    print(f"manifest: {manifest_path}")
    print(f"{man['month']:02d}.{man['year']} · regime {man['regime']} · "
          f"{man['norm_days']} working days · {len(man['people'])} people · "
          f"accident rate {man['tzpb_due']}%")
    if not man["rates_known"]:
        print(f"rates: none the model can compute with for {man['year']} (the "
              f"reference may carry them in another denomination) - the file carries "
              f"the {man['regime']} thresholds rolled forward")
    print(f"policy: {man['policy']}")
    print(f"injected defects ({len(man['expected'])}):")
    for where, idx, ident in man["expected"]:
        loc = "file" if where == "file" else f"row {man['hdr'] + 1 + idx}"
        print(f"  {loc:9} {ident:28} {M.SCENARIOS[ident][1]}")
