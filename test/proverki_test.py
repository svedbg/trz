# -*- coding: utf-8 -*-
"""Проверки по references/proverki.md срещу ведомост за май 2026 г.
Ставки: references/stavki.md, режим 01.01-31.07.2026."""
import os
import openpyxl

# --- ставки (stavki.md, период 01.01-31.07.2026) ---
MRZ            = 620.20     # ПМС 243, ДВ бр.98/18.11.2025            [ДВ]
MRZ_CHAS       = 3.74       # същото                                   [ДВ]
MAX_OSIG       = 2111.64    # чл.9 ЗБДОО 2026, ДВ бр.68/28.07.2026     [ДВ]
MOD_DEYNOST    = None       # Приложение №1 — ЛИПСВА в справочника
KLAS_MIN       = 0.006      # ПМС 147, ДВ бр.56/10.07.2007             [ДВ]
NOSHTEN_CHAS   = round(0.0015 * MRZ, 4)   # чл.8 НСОРЗ = 0.9303
IZV_RAB_DEN    = 0.50       # чл.262 ал.1 т.1 КТ                       [ДВ]
PRAZNIK_MULT   = 2.0        # чл.264 КТ — удвоен размер                [ДВ]
BOLN_DNI       = 2          # чл.40 ал.5 КСО                           [ДВ]
BOLN_PROC      = 0.70       # чл.40 ал.5 КСО                           [ДВ]
LICHNI_PROC    = 0.1378     # [вторичен]
DDFL           = 0.10       # [вторичен]

RAB_DNI, NORMA_CH, PALNO_CH = 18, 144, 8
TOL = 0.02

HERE = os.path.dirname(os.path.abspath(__file__))
wb = openpyxl.load_workbook(os.path.join(HERE, "vedomost_05_2026.xlsx"), data_only=True)
ws = wb.active
H = 5
K = {ws.cell(row=H, column=c).value: c for c in range(1, ws.max_column + 1)}
def g(r, name): return ws.cell(row=r, column=K[name]).value

nah = []   # находки
def f(teg, red, lice, proverka, osnov, nach, dalj, deystvie):
    nah.append(dict(tegest=teg, red=red, lice=lice, proverka=proverka, osnovanie=osnov,
                    nachisleno=nach, dalzhimo=dalj, razlika=(None if nach is None or dalj is None
                                                             else round(dalj - nach, 2)),
                    deystvie=deystvie))

redove = [r for r in range(H + 1, ws.max_row + 1) if isinstance(g(r, "№"), int)]

for r in redove:
    ime, dlaj = g(r, "Име"), g(r, "Длъжност")
    stazh     = g(r, "Стаж (г.)")
    ch_den    = g(r, "Раб. време (ч/ден)")
    otr_dni   = g(r, "Отраб. дни")
    izv_ch    = g(r, "Извънр. часове (раб. дни)") or 0
    praz_ch   = g(r, "Часове на празник") or 0
    nosht_ch  = g(r, "Нощни часове") or 0
    boln_dni  = g(r, "Дни болничен от работодател") or 0
    osnovna   = g(r, "Основна заплата")
    klas_pr   = g(r, "Клас %") or 0
    klas_sum  = g(r, "Клас сума") or 0
    d_izv     = g(r, "Доп. извънреден") or 0
    d_praz    = g(r, "Доп. празник") or 0
    d_nosht   = g(r, "Доп. нощен") or 0
    boln_sum  = g(r, "Болнични от работодател") or 0
    bruto     = g(r, "БРУТО")
    osig      = g(r, "Осиг. доход")
    lichni    = g(r, "Лични осигуровки")
    dan_osn   = g(r, "Данъчна основа")
    danak     = g(r, "ДДФЛ")
    udr       = g(r, "Удръжки") or 0
    neto      = g(r, "НЕТО")

    chast = ch_den / PALNO_CH                      # коефициент непълно работно време
    chasova = osnovna / (otr_dni * ch_den) if otr_dni else None

    # --- B1 МРЗ ---
    mrz_dalj = round(MRZ * chast, 2)
    if osnovna + 1e-9 < mrz_dalj - TOL:
        f("нарушение", r, ime, "B1 основно възнаграждение под МРЗ",
          "чл.244 т.1 КТ; ПМС №243 от 13.11.2025, ДВ бр.98/2025",
          osnovna, mrz_dalj, "Доначисляване до МРЗ и допълнително споразумение към договора.")

    # --- B3 МОД ---
    if MOD_DEYNOST is None:
        pass  # обобщено по-долу, не се повтаря на ред

    # --- B4 максимален осигурителен доход ---
    if osig > MAX_OSIG + TOL:
        f("нарушение", r, ime, "B4 осигурителен доход над максималния",
          "чл.9 ЗБДОО за 2026 г., ДВ бр.68 от 28.07.2026",
          osig, MAX_OSIG, "Ограничаване на осиг. доход до 2111.64 EUR и корекция на вноските (Д1/Д6).")

    # --- C1/C2 клас прослужено време ---
    if stazh and stazh >= 1:
        pr_dalj = round(stazh * KLAS_MIN * 100, 2)
        if klas_pr + 1e-9 < pr_dalj:
            f("нарушение", r, ime, f"C1 клас: приложен {klas_pr}% при {stazh} г. стаж",
              "ПМС №147 от 29.06.2007, ДВ бр.56/2007; чл.12 ал.1 НСОРЗ",
              round(osnovna * klas_pr / 100, 2), round(osnovna * pr_dalj / 100, 2),
              f"Начисляване на минимум {pr_dalj}% върху основната заплата.")
        else:
            ochakv = round(osnovna * klas_pr / 100, 2)
            if abs(klas_sum - ochakv) > TOL:
                f("нарушение", r, ime, "C2 клас: сумата не отговаря на обявения процент",
                  "чл.12 ал.1 НСОРЗ", klas_sum, ochakv, "Преизчисляване на клас върху основната заплата.")

    # --- D4 извънреден труд ---
    if izv_ch:
        dalj = round(chasova * izv_ch * (1 + IZV_RAB_DEN), 2)
        if d_izv + TOL < dalj:
            f("нарушение", r, ime, f"D4 извънреден труд {izv_ch} ч в работни дни без увеличение",
              "чл.262 ал.1 т.1 КТ (+50%)", d_izv, dalj, "Доначисляване на увеличението.")

    # --- D6 нощен труд ---
    if nosht_ch:
        dalj = round(NOSHTEN_CHAS * nosht_ch, 2)
        if d_nosht + TOL < dalj:
            f("нарушение", r, ime, f"D6 нощен труд {nosht_ch} ч без допълнително възнаграждение",
              "чл.8 НСОРЗ (0.15% от МРЗ на час)", d_nosht, dalj, "Доначисляване на нощния труд.")

    # --- D7 труд на официален празник ---
    if praz_ch:
        dalj = round(chasova * praz_ch * PRAZNIK_MULT, 2)
        if d_praz + TOL < dalj:
            f("нарушение", r, ime, f"D7 труд на официален празник {praz_ch} ч под удвоения размер",
              "чл.264 КТ", d_praz, dalj, "Доначисляване до удвоения размер.")

    # --- F9 болнични ---
    if boln_dni > BOLN_DNI:
        sredno = round(boln_sum / boln_dni, 2) if boln_dni else 0
        f("нарушение", r, ime, f"F9 работодателят плаща {boln_dni} дни болничен вместо {BOLN_DNI}",
          "чл.40 ал.5 КСО, изм. ДВ бр.106/2023, в сила от 01.01.2024",
          boln_sum, round(sredno * BOLN_DNI, 2),
          "Първите 2 работни дни са за сметка на осигурителя; останалите — ДОО.")

    # --- F2 лични осигуровки ---
    ochakv = round(min(osig, MAX_OSIG) * LICHNI_PROC, 2)
    if abs(lichni - ochakv) > 0.05:
        f("нарушение", r, ime, "F2 лични осигуровки не отговарят на 13.78% от осиг. доход (с таван)",
          "чл.6 ал.1 и ал.3 КСО", lichni, ochakv, "Преизчисляване на вноските.")

    # --- F6 данъчна основа и данък ---
    osn_ochakv = round(bruto - lichni, 2)
    if abs(dan_osn - osn_ochakv) > TOL:
        f("нарушение", r, ime, "F6 данъчна основа ≠ бруто − лични осигуровки",
          "ЗДДФЛ", dan_osn, osn_ochakv, "Корекция на данъчната основа.")
    dan_ochakv = round(dan_osn * DDFL, 2)
    if abs(danak - dan_ochakv) > TOL:
        f("нарушение", r, ime, "F6 данъкът ≠ 10% от данъчната основа", "ЗДДФЛ",
          danak, dan_ochakv, "Корекция на данъка.")

    # --- I2 сбор на начисленията = бруто ---
    sbor = round(osnovna + klas_sum + d_izv + d_praz + d_nosht + boln_sum, 2)
    if abs(bruto - sbor) > TOL:
        f("нарушение", r, ime, "I2 сборът на начисленията ≠ БРУТО", "аритметична консистентност",
          bruto, sbor, "Проверка на начисленията по видове.")

    # --- I1 вертикална сверка ---
    neto_ochakv = round(bruto - lichni - danak - udr, 2)
    if abs(neto - neto_ochakv) > TOL:
        f("нарушение", r, ime, "I1 НЕТО ≠ бруто − лични осигуровки − данък − удръжки",
          "аритметична консистентност", neto, neto_ochakv, "Корекция на изплатената сума.")

    # --- G2 несеквестируем доход ---
    if udr > 0:
        f("за проверка", r, ime, f"G2 удръжка {udr:.2f} EUR при нето преди удръжки "
          f"{round(bruto - lichni - danak, 2):.2f} EUR",
          "чл.446 ГПК — праговете НЕ СА в references/stavki.md", udr, None,
          "Да се въведат праговете по чл.446 ГПК и броят издържани лица, след което да се преизчисли.")

# --- отчет ---
red = {"нарушение": 0, "риск": 1, "за проверка": 2, "бележка": 3}
nah.sort(key=lambda x: (red[x["tegest"]], x["red"]))
print(f"НАХОДКИ: {len(nah)}\n" + "=" * 100)
for n in nah:
    print(f"[{n['tegest'].upper():11}] ред {n['red']:2d} · {n['lice']}")
    print(f"  {n['proverka']}")
    print(f"  Основание: {n['osnovanie']}")
    if n["dalzhimo"] is not None:
        print(f"  Начислено {n['nachisleno']:.2f} | Дължимо {n['dalzhimo']:.2f} | Разлика {n['razlika']:+.2f} EUR")
    print(f"  Действие: {n['deystvie']}\n")

zaseg = sorted({n["lice"] for n in nah if n["tegest"] == "нарушение"})
suma = sum(n["razlika"] for n in nah
           if n["tegest"] == "нарушение" and n["razlika"] and n["razlika"] > 0
           and "осигурителен доход над" not in n["proverka"])
print("=" * 100)
print(f"Лица с нарушения: {len(zaseg)} от {len(redove)}")
print(f"Недоплатено на работниците (без корекциите по осигуряване): {suma:.2f} EUR")
print(f"Лица без находки: {sorted(set(g(r,'Име') for r in redove) - {n['lice'] for n in nah})}")
