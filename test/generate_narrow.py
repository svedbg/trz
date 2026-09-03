# -*- coding: utf-8 -*-
"""Builds the static test payroll for suite 1: `vedomost_05_2026.xlsx`.

Eleven people, May 2026, nine deliberately injected defects and two clean control
rows. The answer key is in `expected_findings.md`.

The second control row (row 16) is paid at exactly the statutory minimum for
overtime, night work and a public holiday, on the чл. 7 НСОРЗ base (basic salary
plus the length-of-service supplement). It exists because every other row with
those hours pays nothing above the single rate, so any positive rate in the checker
fired and a wrong rate was indistinguishable from the right one; a row paid exactly
at the minimum goes red the moment a rate or a base in the checker drifts upward.

The period is chosen on purpose: May 2026 falls in the 01.01-31.07.2026 regime
and its norm is 18 working days / 144 hours - 21 weekdays minus 1, 6 and 25 May,
because 24 May is a Sunday (чл. 154, ал. 2 КТ).

Everything here is invented. Column headers and names are Bulgarian because they
are data; the code is English.
"""
import os

import openpyxl
from openpyxl.styles import Alignment, Font

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ведомост 05.2026"

ws["A1"] = 'ВЕДОМОСТ ЗА РАБОТНИ ЗАПЛАТИ — "Примерна Логистика" ЕООД'
ws["A2"] = "Месец: май 2026 г.   |   Работни дни в месеца: 18   |   Норма часове: 144"
ws["A3"] = "Икономическа дейност: 52.10 Складиране и съхраняване на товари"
for r in (1, 2, 3):
    ws.cell(row=r, column=1).font = Font(bold=(r == 1), size=12 if r == 1 else 10)

COLUMNS = ["№", "Име", "Длъжност", "Дата постъпване", "Стаж (г.)",
           "Раб. време (ч/ден)", "Отраб. дни", "Отраб. часове",
           "Извънр. часове (раб. дни)", "Часове на празник", "Нощни часове",
           "Дни болничен от работодател",
           "Основна заплата", "Клас %", "Клас сума",
           "Доп. извънреден", "Доп. празник", "Доп. нощен",
           "Болнични от работодател", "БРУТО", "Осиг. доход",
           "Лични осигуровки", "Данъчна основа", "ДДФЛ", "Удръжки", "НЕТО"]

HDR = 5
for i, column in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=HDR, column=i, value=column)
    cell.font = Font(bold=True, size=9)
    cell.alignment = Alignment(wrap_text=True, vertical="top")

# №, name, position, hired, years, hours/day, days, hours, overtime, holiday hrs,
# night hrs, sick days, base pay, seniority %, seniority sum, overtime premium,
# holiday premium, night premium, sick pay, gross, insurable, employee
# contributions, taxable, tax, deductions, net
ROWS = [
    (1, "Иван Петров", "Чистач", "2023-03-01", 3, 8, 18, 144, 0, 0, 0, 0,
     610.00, 1.8, 10.98, 0, 0, 0, 0, 620.98, 620.98, 85.57, 535.41, 53.54, 0, 481.87),
    (2, "Мария Георгиева", "Специалист логистика", "2014-04-15", 12, 8, 18, 144, 0, 0, 0, 0,
     900.00, 0.0, 0.00, 0, 0, 0, 0, 900.00, 900.00, 124.02, 775.98, 77.60, 0, 698.38),
    (3, "Георги Иванов", "Техник поддръжка", "2018-02-01", 8, 8, 18, 154, 10, 0, 0, 0,
     800.00, 4.8, 38.40, 55.56, 0, 0, 0, 893.96, 893.96, 123.19, 770.77, 77.08, 0, 693.69),
    (4, "Елена Димитрова", "Управител операции", "2011-01-10", 15, 8, 18, 144, 0, 0, 0, 0,
     3500.00, 9.0, 315.00, 0, 0, 0, 0, 3815.00, 3815.00, 525.71, 3289.29, 328.93, 0, 2960.36),
    (5, "Петър Стоянов", "Оператор склад", "2021-06-01", 5, 8, 18, 144, 0, 0, 60, 0,
     700.00, 3.0, 21.00, 0, 0, 0.00, 0, 721.00, 721.00, 99.35, 621.65, 62.17, 0, 559.48),
    # The injected defect on this row is the THIRD employer-paid sick day (F9); the
    # two bases are correct and must stay so. The чл. 40, ал. 5 КСО sick pay is inside
    # the insurable income (чл. 3, ал. 1 НЕВДПОВ) and outside the taxable base
    # (чл. 24, ал. 2, т. 14 ЗДДФЛ) - the asymmetry is the statute's, not a defect.
    # Insurable 836.56 = the whole gross; taxable 618.55 = 836.56 - 102.73 - 115.28.
    (6, "Анна Тодорова", "Специалист доставки", "2020-05-04", 6, 8, 15, 120, 0, 0, 0, 3,
     733.83, 3.6, 0.00, 0, 0, 0, 102.73, 836.56, 836.56, 115.28, 618.55, 61.86, 0, 659.42),
    (7, "Димитър Николов", "Складов работник", "2024-09-02", 2, 8, 18, 144, 0, 0, 0, 0,
     700.00, 1.2, 8.40, 0, 0, 0, 0, 708.40, 708.40, 97.62, 610.78, 61.08, 0, 560.00),
    (8, "Стефка Ангелова", "Чистач", "2025-07-01", 1, 4, 18, 72, 0, 0, 0, 0,
     310.10, 0.6, 1.86, 0, 0, 0, 0, 311.96, 311.96, 42.99, 268.97, 26.90, 0, 242.07),
    (9, "Николай Христов", "Техник поддръжка", "2022-03-14", 4, 8, 18, 152, 0, 8, 0, 0,
     780.00, 2.4, 18.72, 0, 43.33, 0, 0, 842.05, 842.05, 116.03, 726.02, 72.60, 0, 653.42),
    (10, "Виктор Маринов", "Специалист транспорт", "2020-02-17", 6, 8, 18, 144, 0, 0, 0, 0,
     900.00, 3.6, 32.40, 0, 0, 0, 0, 932.40, 932.40, 128.48, 803.92, 80.39, 500.00, 223.53),
    # Second control row: everything at exactly the statutory minimum, nothing above.
    # Hourly base under чл. 7 НСОРЗ = (900.00 + 32.40) / 144 = 6.475. Overtime 8 h on
    # working days: 6.475 × 8 × 1.5 = 77.70. Public holiday 8 h at double: 6.475 × 8 × 2
    # = 103.60. Night 16 h at 0.15% of МРЗ: 0.9303 × 16 = 14.88 (0.93 × 16 gives the
    # same cents, so the per-hour rounding question does not decide this row). Gross
    # 1128.58; contributions 13.78% = 155.52; taxable 973.06; tax 97.31; net 875.75.
    # No rounding step here lands on a half cent.
    (11, "Росица Кънчева", "Оператор склад", "2019-10-01", 6, 8, 18, 160, 8, 8, 16, 0,
     900.00, 3.6, 32.40, 77.70, 103.60, 14.88, 0, 1128.58, 1128.58, 155.52, 973.06, 97.31, 0, 875.75),
]

for r, row in enumerate(ROWS, start=HDR + 1):
    for c, value in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=value)

TOTAL = HDR + len(ROWS) + 1
ws.cell(row=TOTAL, column=2, value="ОБЩО").font = Font(bold=True)
for c in (13, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26):
    total = sum(row[c - 1] for row in ROWS)
    cell = ws.cell(row=TOTAL, column=c, value=round(total, 2))
    cell.font = Font(bold=True)

ws.cell(row=TOTAL + 2, column=1,
        value="Валута: EUR. Изготвил: ТРЗ отдел. Дата: 05.06.2026 г.")

for i, width in enumerate([4, 20, 22, 13, 7, 8, 8, 9, 11, 10, 9, 12,
                           11, 7, 9, 11, 11, 10, 12, 10, 11, 11, 11, 9, 9, 10], start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "vedomost_05_2026.xlsx")
wb.save(out)
print("written:", out, "|", len(ROWS), "rows")
