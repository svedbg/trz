# -*- coding: utf-8 -*-
"""Generator of a two-month payroll: one workbook, two sheets, one roster.

    python test/generate_pair.py --seed 7

Writes `test/tmp/pair_<seed>.xlsx` and `..._manifest.json`. Everything is invented and
derived from the seed, as in `generate_wide.py`.

Why a second fixture exists at all. Three of the documented checks cannot be expressed in
one sheet, and until now none of them had a scenario:

  * **K8** - a sheet copied from the previous month keeps the previous month's
    thresholds. In a single sheet a stale threshold looks exactly like the threshold.
  * **I7** - a jump in someone's gross between adjacent months. One month has nothing to
    jump from.
  * **чл. 177 КТ** - paid leave is measured against the average daily gross of the
    preceding month, not against the contract. The base is in the month before.

The months are **July and August 2026**, and not by chance: the 2026 budget was adopted
late, so the thresholds change on 1 August. Two adjacent sheets therefore need different
figures, and copying the first is both the most natural thing for a person to do and
demonstrably wrong. That is the whole fixture in one sentence.

The roster is identical across the sheets - same people, same contracts - because that is
what makes the cross-month comparison meaningful, and what a real payroll looks like.

July carries no paid leave. It is the base month: its gross has to be unambiguous for
August's leave to be measured against it.
"""
import argparse
import json
import os
import random

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import trz_model as M
from trz_model import r2
import generate_wide as G

HERE = os.path.dirname(os.path.abspath(__file__))
TMP = os.path.join(HERE, "tmp")

MONTH_EARLY, MONTH_LATE = 7, 8          # H1 and H2: the thresholds change between them
YEAR = 2026


def _inputs_july(rnd, norm, regime):
    """A person's July: work, possibly sick days, never paid leave."""
    inp = G.random_inputs(rnd, norm, regime)
    inp["days_leave"] = 0
    inp["days_worked"] = norm - inp["days_sick"]
    return inp


def _inputs_august(rnd, norm, july):
    """The same person's August: same contract, and often some leave."""
    leave = rnd.choice([0, 0, 3, 5, 8, 10])
    sick = 0 if leave else rnd.choice([0] * 5 + [2, 3])
    return dict(july,
                days_leave=leave, days_sick=sick,
                days_worked=norm - leave - sick,
                bonus=r2(rnd.uniform(80, 500)) if rnd.random() < 0.35 else 0.0)


def _permanent_pay(row, policy):
    """The row's remuneration of permanent character — the base чл. 177 КТ measures.

    The bonus column is in or out according to the configured reading of чл. 17, ал. 1:
    out as a one-off (in none of the seven points), in as т. 2 pay determined by an
    applied wage system. See permanent_work_pay.
    """
    return M.permanent_work_pay(row["Основна за отработеното"], row["Клас сума"],
                                row["Бонус"] if policy.get("bonus_in_base") else 0.0)


def _write_sheet(wb, title, header, people, tzpb, rnd, first):
    ws = wb.create_sheet(title) if not first else wb.active
    if first:
        ws.title = title
    ws["A1"] = header[0]
    ws["A2"] = header[1]
    ws["A3"] = f"Икономическа дейност: тестова  |  ТЗПБ по КИД: {tzpb}%"
    ws["A1"].font = Font(bold=True, size=12)

    hdr = 5
    for i, column in enumerate(M.COLUMNS, start=1):
        cell = ws.cell(row=hdr, column=i, value=column)
        cell.font = Font(bold=True, size=8)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for offset, person in enumerate(people):
        r = hdr + 1 + offset
        ws.cell(row=r, column=M.COL["№"], value=offset + 1)
        ws.cell(row=r, column=M.COL["Име"], value=person["name"])
        ws.cell(row=r, column=M.COL["Отдел"], value=person["department"])
        for column in M.COLUMNS:
            if column in ("№", "Име", "Отдел"):
                continue
            v = person["row"].get(column, 0)
            ws.cell(row=r, column=M.COL[column],
                    value=(v if v else (0 if column in M.DAY_COLUMNS else None)))

    total_row = hdr + 1 + len(people)
    ws.cell(row=total_row, column=M.COL["Име"], value="ОБЩО").font = Font(bold=True)
    for column in M.SUMMED_COLUMNS:
        s = r2(sum(p["row"].get(column, 0) or 0 for p in people))
        ws.cell(row=total_row, column=M.COL[column], value=s).font = Font(bold=True)

    for i, column in enumerate(M.COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = \
            max(9, min(15, len(column) // 2 + 6))
    return ws, hdr, total_row


def generate(seed):
    rnd = random.Random(seed)
    norm_july = M.working_days(YEAR, MONTH_EARLY)
    norm_august = M.working_days(YEAR, MONTH_LATE)
    regime_july = M.REGIMES["H1"]
    regime_august = M.REGIMES["H2"]
    tzpb = rnd.choice([0.4, 0.5, 0.7, 1.1])
    reading = rnd.choice(list(M.EXCESS_READINGS))
    policy = dict(bonus_in_base=rnd.random() < 0.5,
                  in_kind_in_bases=rnd.random() < 0.5,
                  excess_in_insurable=M.EXCESS_READINGS[reading]["insurable"],
                  excess_in_taxable=M.EXCESS_READINGS[reading]["taxable"],
                  excess_reading=reading)
    company = rnd.choice(G.COMPANIES)
    n = rnd.randint(8, 12)

    # --- which cross-month defects this seed carries ------------------------
    stale = rnd.random() < 0.5           # August built on July's thresholds
    regime_for_august = regime_july if stale else regime_august
    norm_for_august = norm_july if stale else norm_august

    used = set()
    people = []
    for _ in range(n):
        july_inp = _inputs_july(rnd, norm_july, regime_july)
        july_row = M.clean_row(july_inp, regime_july, tzpb, policy, norm_july)
        july_row["_norm"] = norm_july

        august_inp = _inputs_august(rnd, norm_for_august, july_inp)
        # чл. 177: August's leave is measured against July's average daily gross.
        # чл. 18 НСОРЗ: July's permanent pay over July's worked days, then corrected
        # by the ratio of the two months' norms (ал. 2).
        base = M.leave_daily_base(_permanent_pay(july_row, policy),
                                  july_inp["days_worked"], norm_july, norm_for_august)
        august_row = M.clean_row(august_inp, regime_for_august, tzpb, policy,
                                 norm_for_august, leave_daily=base)
        august_row["_norm"] = norm_for_august
        people.append(dict(name=G._name(rnd, used),
                           department=rnd.choice(G.DEPARTMENTS),
                           july=dict(inputs=july_inp, row=july_row),
                           august=dict(inputs=august_inp, row=august_row),
                           leave_daily_due=base))

    cross = []
    if stale:
        cross.append(["file", None, "K8_stale_thresholds"])

    # --- paid leave computed on the bonus's other side of чл. 17, ал. 1 --------
    # Whichever reading the file applies, the defect is the base built the other way:
    # July's bonus carried into August's leave days when it should stay out, or left
    # out when the company pays it under a wage system (т. 2). Paying the leave from
    # the contract is NOT the defect it looks like - with the чл. 18, ал. 2 coefficient
    # the two norms cancel and the correct base lands on the leave month's contracted
    # daily rate.
    for idx, p in enumerate(people):
        if not p["august"]["inputs"]["days_leave"]:
            continue
        inp = p["august"]["inputs"]
        july_row, july_inp = p["july"]["row"], p["july"]["inputs"]
        if not july_row["Бонус"] or not july_inp["days_worked"]:
            continue
        other = dict(policy, bonus_in_base=not policy.get("bonus_in_base"))
        wrong = M.leave_daily_base(_permanent_pay(july_row, other),
                                   july_inp["days_worked"], norm_july, norm_for_august)
        if abs(wrong - p["leave_daily_due"]) * inp["days_leave"] < 1.0:
            continue                     # the two bases coincide; nothing to see
        broken = M.clean_row(inp, regime_for_august, tzpb, policy, norm_for_august,
                             leave_daily=wrong)
        broken["_norm"] = norm_for_august
        p["august"]["row"] = broken
        cross.append(["row", idx, "E3_leave_base"])
        break

    # --- an unexplained jump in the gross ------------------------------------
    for idx, p in enumerate(people):
        if any(c[1] == idx for c in cross if c[0] == "row"):
            continue
        if p["august"]["inputs"]["days_leave"] or p["august"]["inputs"]["days_sick"]:
            continue                     # a jump has to be unexplained to be a finding
        row = dict(p["august"]["row"])
        bump = r2(row["БРУТО"] * rnd.uniform(0.35, 0.6))
        row["Основна за отработеното"] = r2(row["Основна за отработеното"] + bump)
        row["БРУТО"] = r2(row["БРУТО"] + bump)
        p["august"]["row"] = G._recompute_downstream(
            row, p["august"]["inputs"], regime_for_august, tzpb, policy)
        p["august"]["row"]["_norm"] = norm_for_august
        cross.append(["row", idx, "I7_unexplained_jump"])
        break

    # --- write ---------------------------------------------------------------
    os.makedirs(TMP, exist_ok=True)
    wb = openpyxl.Workbook()
    sheets = []
    for label, month, norm, regime, key, first in (
            ("07-2026", MONTH_EARLY, norm_july, regime_july, "july", True),
            ("08-2026", MONTH_LATE, norm_for_august, regime_for_august, "august", False)):
        header = (f'ВЕДОМОСТ ЗА РАБОТНИ ЗАПЛАТИ — "{company}" ЕООД',
                  f"Месец: {G.MONTHS_BG[month]} {YEAR} г.  |  Работни дни: {norm}  |  "
                  f"Валута: EUR  |  ЕИК: 000000000 (тестов)")
        rows = [dict(name=p["name"], department=p["department"], row=p[key]["row"])
                for p in people]
        ws, hdr, total = _write_sheet(wb, label, header, rows, tzpb, rnd, first)
        sheets.append(dict(
            sheet=label, month=month, year=YEAR, norm_days=norm,
            regime="H1" if regime is regime_july else "H2",
            rates_known=True,
            max_insurable=regime["max_insurable"],
            min_insurable_self=regime["min_insurable_self"],
            tzpb_due=tzpb, policy=policy, hdr=hdr, total_row=total,
            people=[dict(row=hdr + 1 + i, name=p["name"], inputs=p[key]["inputs"],
                         defects=[]) for i, p in enumerate(people)],
            expected=[]))

    path = os.path.join(TMP, f"pair_{seed}.xlsx")
    wb.save(path)

    manifest = dict(
        seed=seed, file=os.path.basename(path), year=YEAR,
        # File-wide, and at the top level because the checker needs it before it
        # reaches either sheet: the configured reading of чл. 17, ал. 1 for an
        # uncharacterised bonus column.
        policy=policy,
        months=[MONTH_EARLY, MONTH_LATE],
        applicable=dict(early=dict(max_insurable=regime_july["max_insurable"],
                                   norm_days=norm_july),
                        late=dict(max_insurable=regime_august["max_insurable"],
                                  norm_days=norm_august)),
        stale=stale, sheets=sheets, cross_expected=cross,
        leave_daily_due={p["name"]: p["leave_daily_due"] for p in people},
    )
    manifest_path = os.path.join(TMP, f"pair_{seed}_manifest.json")
    with open(manifest_path, "w", encoding="utf8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=1)
    return path, manifest_path, manifest


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--seed", type=int, default=1)
    a = ap.parse_args()
    path, manifest_path, man = generate(a.seed)
    print(f"written:  {path}")
    print(f"manifest: {manifest_path}")
    print(f"{man['year']} · sheets {[s['sheet'] for s in man['sheets']]} · "
          f"{len(man['sheets'][0]['people'])} people · "
          f"August built on {'July' if man['stale'] else 'its own'} thresholds")
    print(f"injected cross-month defects ({len(man['cross_expected'])}):")
    for where, idx, ident in man["cross_expected"]:
        loc = "file" if where == "file" else f"row {man['sheets'][1]['hdr'] + 1 + idx}"
        print(f"  {loc:9} {ident:26} {M.PAIR_SCENARIOS[ident][1]}")
