# -*- coding: utf-8 -*-
"""Suite 4's generator: a payroll whose computed columns carry REAL formulas.

Every other fixture in this repository is a value-only export, so the skill's whole
"чети файла два пъти" guidance and the K-group's formula semantics had ZERO suite
coverage - while the first real audit this skill performed found every one of its
defects in the formula layer: a cap typed by hand on 13 of 24 rows, a days column
added into a money sum, a control column algebraically always zero, a premium
inlined as `=31.88*0.02+31.88` on every row. This generator writes those shapes on
purpose; `formula_test.py` must find them from the formulas alone.

openpyxl writes formulas WITHOUT cached values (nothing evaluates them), which is
exactly why this suite checks structure, not arithmetic - and why the paid eval does
not use this fixture yet: a live session opening it with data_only=True would see
None everywhere. A faithful eval fixture needs Excel-produced caches; that is
recorded in scenarios.md as the known limit of this suite.

Headers are Bulgarian because they are data. Parameter cells live in row 3, named in
column A, and the CLEAN file references them absolutely - a formula that inlines the
figure instead is one of the defects.
"""
import argparse
import json
import os
import random

import openpyxl
from openpyxl.utils import get_column_letter

import trz_model as M
from trz_model import r2
import generate_wide as G

HERE = os.path.dirname(os.path.abspath(__file__))
TMP = os.path.join(HERE, "tmp")

# The sheet is August 2026, so its parameters are the H2 regime's. From the model, not
# typed: the cap, the employee control sum and the tax rate below used to be literals,
# which is the one rule this repository does not bend on.
YEAR, MONTH = 2026, 8
CAP = M.REGIMES[M.regime_for(YEAR, MONTH)]["max_insurable"]

# A deliberately narrow layout: enough columns to carry every formula shape the
# suite tests, few enough that the shapes stay readable in a diff.
COLUMNS = ["№", "Име",
           "Отраб. дни", "Дни болничен",
           "Основна за отработеното", "Клас сума", "Бонус", "Платен отпуск",
           "БРУТО",
           "Осигурителен доход",
           "Лични вноски общо", "ДДФЛ", "НЕТО за изплащане",
           "Изплатено", "Разлика"]
COL = {name: i + 1 for i, name in enumerate(COLUMNS)}
ACCRUALS = ["Основна за отработеното", "Клас сума", "Бонус", "Платен отпуск"]
DAY_COLS = ["Отраб. дни", "Дни болничен"]
# The columns the clean file computes with formulas (everything downstream of entry).
FORMULA_COLUMNS = ["БРУТО", "Осигурителен доход", "Лични вноски общо", "ДДФЛ",
                   "НЕТО за изплащане", "Разлика"]

HDR = 5
P_CAP = "$D$3"            # parameter cells the clean formulas point at
P_RATE = "$F$3"           # the employee-contribution control sum, 13.78
P_TAX = "$H$3"


def _L(name):
    return get_column_letter(COL[name])


def _clean_formulas(r):
    """The formulas a sound row carries. Structure is the contract here."""
    acc = "+".join(f"{_L(c)}{r}" for c in ACCRUALS)
    return {
        "БРУТО": f"={acc}",
        "Осигурителен доход": f"=MIN({_L('БРУТО')}{r},{P_CAP})",
        "Лични вноски общо": f"=ROUND({_L('Осигурителен доход')}{r}*{P_RATE}/100,2)",
        "ДДФЛ": f"=ROUND(({_L('БРУТО')}{r}-{_L('Лични вноски общо')}{r})*{P_TAX},2)",
        "НЕТО за изплащане": (f"={_L('БРУТО')}{r}-{_L('Лични вноски общо')}{r}"
                              f"-{_L('ДДФЛ')}{r}"),
        "Разлика": f"={_L('НЕТО за изплащане')}{r}-{_L('Изплатено')}{r}",
    }


# ------------------------------------------------------------------- mutations
# Each takes the dict row-formulas for row r and returns (new_formulas, extra_cells,
# ident) or None. extra_cells lets a mutation overwrite a non-formula cell too.

def m_sum_omits_column(fm, r, rnd):
    """One accrual cell dropped from the gross - the defect with a delay fuse."""
    victim = rnd.choice(ACCRUALS[1:])           # never the base salary; too loud
    keep = [c for c in ACCRUALS if c != victim]
    fm = dict(fm, **{"БРУТО": "=" + "+".join(f"{_L(c)}{r}" for c in keep)})
    return fm, {}, "KF1_sum_omits_column"


def m_days_in_money_sum(fm, r, rnd):
    """A day-count cell added into the gross - money plus days, as seen live."""
    acc = "+".join(f"{_L(c)}{r}" for c in ACCRUALS)
    fm = dict(fm, **{"БРУТО": f"={acc}+{_L('Дни болничен')}{r}"})
    return fm, {}, "KF2_days_in_money_sum"


def m_hard_value_in_formula_column(fm, r, rnd):
    """The insurable income typed as a literal in a column of formulas.

    The live audit's exact shape: 13 of 24 rows carried the cap as a typed value.
    The number can even be RIGHT today - the defect is that it stops following."""
    fm = dict(fm)
    del fm["Осигурителен доход"]
    return fm, {"Осигурителен доход": CAP}, "KF3_hard_value_in_formula_column"


def m_tautological_control(fm, r, rnd):
    """Изплатено becomes =НЕТО, so Разлика is algebraically always zero."""
    return fm, {"Изплатено": f"={_L('НЕТО за изплащане')}{r}"}, \
        "KF4_tautological_control"


def m_constant_in_formula(fm, r, rnd):
    """One row inlines the parameter as a literal instead of referencing it."""
    which = rnd.choice(["cap", "rate"])
    fm = dict(fm)
    if which == "cap":
        fm["Осигурителен доход"] = f"=MIN({_L('БРУТО')}{r},{CAP:g})"
    else:
        fm["Лични вноски общо"] = \
            f"=ROUND({_L('Осигурителен доход')}{r}*{M.EMPLOYEE_TOTAL}/100,2)"
    return fm, {}, "KF5_constant_in_formula"


def m_shape_deviates(fm, r, rnd):
    """One row's formula has a different shape and no literal in it.

    The shape-uniformity pass sorts a deviating row into three bins: a typed value
    (KF3), a literal where the others reference a parameter (KF5), and everything else
    - a formula that is simply not the column's formula. Two such shapes, both seen in
    real files: an insurable income that skips the MIN against the cap, and a net that
    forgets to subtract the tax. Until this mutation existed that third bin had never
    been reached.
    """
    fm = dict(fm)
    if rnd.random() < 0.5:
        fm["Осигурителен доход"] = f"={_L('БРУТО')}{r}"                  # no cap
    else:
        fm["НЕТО за изплащане"] = (f"={_L('БРУТО')}{r}"
                                   f"-{_L('Лични вноски общо')}{r}")        # no tax
    return fm, {}, "KF_shape_deviates"


MUTATIONS = [m_sum_omits_column, m_days_in_money_sum, m_hard_value_in_formula_column,
             m_tautological_control, m_constant_in_formula, m_shape_deviates]


def generate(seed):
    rnd = random.Random(10_000_000 + seed)
    n = rnd.randint(8, 14)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{MONTH:02d}-{YEAR}"
    ws["A1"] = 'ВЕДОМОСТ ЗА РАБОТНИ ЗАПЛАТИ — "Формулен тест" ЕООД (изцяло измислена)'
    ws["A2"] = f"Месец: {G.MONTHS_BG[MONTH]} {YEAR} г.  |  Валута: EUR"
    ws["C3"], ws["D3"] = "Таван:", CAP
    ws["E3"], ws["F3"] = "Лични %:", M.EMPLOYEE_TOTAL
    ws["G3"], ws["H3"] = "Данък:", M.TAX_RATE
    for i, c in enumerate(COLUMNS, start=1):
        ws.cell(row=HDR, column=i, value=c)

    how_many = rnd.randint(2, 4)
    victims = rnd.sample(range(n), how_many)
    muts = rnd.sample(MUTATIONS, how_many)
    expected = []
    norm = M.working_days(YEAR, MONTH)

    for idx in range(n):
        r = HDR + 1 + idx
        ws.cell(row=r, column=COL["№"], value=idx + 1)
        ws.cell(row=r, column=COL["Име"], value=f"Лице {idx + 1:02d} (измислено)")
        worked = rnd.randint(15, norm)
        sick = rnd.choice([0, 0, 0, 2])
        # Days worked plus sick days cannot exceed the month. Clamped after both draws
        # rather than drawn within the room left, so that the rows this never touched
        # keep the values every existing seed already had.
        worked = min(worked, norm - sick)
        base = r2(rnd.uniform(900, 4200))
        ws.cell(row=r, column=COL["Отраб. дни"], value=worked)
        ws.cell(row=r, column=COL["Дни болничен"], value=sick)
        ws.cell(row=r, column=COL["Основна за отработеното"], value=base)
        ws.cell(row=r, column=COL["Клас сума"], value=r2(base * 0.024))
        ws.cell(row=r, column=COL["Бонус"],
                value=r2(rnd.uniform(60, 300)) if rnd.random() < 0.25 else 0)
        ws.cell(row=r, column=COL["Платен отпуск"], value=0)
        ws.cell(row=r, column=COL["Изплатено"], value=0)   # entry cell by design

        fm = _clean_formulas(r)
        extra = {}
        if idx in victims:
            fn = muts[victims.index(idx)]
            out = fn(fm, r, rnd)
            if out is not None:
                fm, extra, ident = out
                expected.append(["row", idx, ident])
        for name, formula in fm.items():
            ws.cell(row=r, column=COL[name], value=formula)
        for name, value in extra.items():
            ws.cell(row=r, column=COL[name], value=value)

    total = HDR + 1 + n
    ws.cell(row=total, column=COL["Име"], value="ОБЩО")
    for name in FORMULA_COLUMNS + ACCRUALS:
        cletter = _L(name)
        ws.cell(row=total, column=COL[name],
                value=f"=SUM({cletter}{HDR + 1}:{cletter}{total - 1})")

    os.makedirs(TMP, exist_ok=True)
    path = os.path.join(TMP, f"formula_{seed}.xlsx")
    G.save_frozen(wb, path)
    man = dict(seed=seed, file=os.path.basename(path), hdr=HDR, total_row=total,
               people=n, expected=expected)
    mpath = os.path.join(TMP, f"formula_{seed}_manifest.json")
    with open(mpath, "w", encoding="utf8") as f:
        json.dump(man, f, ensure_ascii=False, indent=1)
    return path, mpath, man


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--seed", type=int, default=1)
    a = ap.parse_args()
    p, mp, man = generate(a.seed)
    print(f"written:  {p}\nmanifest: {mp}")
    for where, idx, ident in man["expected"]:
        print(f"  row {man['hdr'] + 1 + idx}  {ident}")
