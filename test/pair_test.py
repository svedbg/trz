# -*- coding: utf-8 -*-
"""Suite 3: the cross-month checks, over a two-sheet payroll.

    python test/pair_test.py --seed 7
    python test/pair_test.py --seed 7 --quiet

`generate_pair.py` builds July and August 2026 into one workbook with one roster, and
breaks the things that can only be broken across months. This file has to find exactly
those and nothing else.

People are identified by row, never by name, as in `structural_test.py`. The names in
these fixtures are invented, but the report is the part a person pastes into an issue, and
a row number carries the same information.

The I7 finding carries no figures at all - not in the report and not in the finding it
returns. Its evidence is a pay figure rather than a discrepancy in one, and writing that
anywhere is clear-text logging of the kind CodeQL flags, whatever the data behind it
happens to be. Guarding the print was not enough: a runtime condition is invisible to the
scanner, and the value had already entered the structure being printed. So it is not
computed into the finding. The suite scores on identity and location, which is all it ever
used; a case is reproduced from its seed.

Every *finding* here is derived from the two sheets, never from the manifest, and the
manifest is read to score the run. That is not fastidiousness: the checks are meant to be
the ones a person could perform with the file alone, and a check that quietly reads the
answer key proves nothing about the file.

One value is read from the manifest before the sheets, and it is not a finding:
`policy.bonus_in_base`, the configured reading of чл. 17, ал. 1 for a bonus column the
file does not characterise. That is the auditor's own setting - the plugin asks it at
install time - so a checker that did not know it would be auditing against a rule nobody
chose. Nothing else may join it: if a future value is a property of the file, infer it.

Two of the three checks turn on a distinction worth stating. The norm of a sheet - how
many days it was built on - is not the same as the norm of its month. When a sheet is
copied forward, the first stops following the second, and every row then reconciles
perfectly against a month that is not its own. So the norm is *derived* from the day
sums and then compared with the calendar; deriving it from the calendar instead would
turn one file-level defect into a row-level finding against every person on the sheet.
"""
import argparse
import json
import os
import sys
from collections import Counter

import openpyxl

import trz_model as M
from trz_model import r2

HERE = os.path.dirname(os.path.abspath(__file__))
TOL = 0.02
SALARY_TOL = 1.00        # a month's rounding moves an implied salary by a few cents


from findings import Findings


def read_sheet(ws):
    """Every row of one sheet, by column name, plus the header and total row numbers."""
    hdr = next(r for r in range(1, 12)
               if ws.cell(row=r, column=1).value == "№")
    col = {ws.cell(row=hdr, column=c).value: c for c in range(1, ws.max_column + 1)}
    rows = []
    r = hdr + 1
    while ws.cell(row=r, column=col["Име"]).value not in (None, "", "ОБЩО"):
        v = {}
        for name, c in col.items():
            if name is None:
                continue
            value = ws.cell(row=r, column=c).value
            v[name] = 0.0 if value in (None, "") and name not in ("Име", "Отдел") else value
        v["_row"] = r
        rows.append(v)
        r += 1
    return hdr, r, rows


def derived_norm(rows):
    """The number of days the sheet was actually built on.

    The day columns of a correct sheet sum to the month's norm on every row. Taking the
    most common sum recovers that figure from the file, which is the only way to notice
    that it belongs to a different month.
    """
    sums = Counter(r2(sum(row[c] for c in M.DAY_COLUMNS)) for row in rows)
    return sums.most_common(1)[0][0]


def implied_salary(row, sheet_norm):
    """The monthly salary the row implies: daily rate scaled back up to the month."""
    worked = row["Отраб. дни"]
    return None if not worked else row["Основна за отработеното"] / worked * sheet_norm


def selftest_leave_base():
    """Pin чл. 18 НСОРЗ against arithmetic, not against the generator.

    The suites cannot see this one. The fixture and the checker both call
    M.leave_daily_base, so dropping the ал. 2 coefficient from it changes both sides
    equally and 300 seeds stay green - the round-trip blindness scenarios.md describes,
    which has now hidden three defects in this same area.

    So the property is asserted directly, derived from the statute rather than from the
    model: for an unchanged monthly contract, чл. 18, ал. 1 divides the base month's
    permanent pay by the days worked in it, and ал. 2 multiplies by (base norm / leave
    norm). The base month's norm cancels, and what is left is the salary over the LEAVE
    month's norm - the contracted daily rate of the month the leave falls in. That is
    why a payroll paying leave from the contract is right, and the reason must hold for
    two months whose norms differ, or the claim is empty.
    """
    salary, pct = 1740.0, 7.2
    uplift = 1 + pct / 100.0
    cases = 0
    for base_norm, leave_norm in ((23, 21), (21, 23), (20, 20)):
        for sick in (0, 3):
            cases += 1
            worked = base_norm - sick
            daily = salary / base_norm
            permanent = M.permanent_work_pay(r2(daily * worked),
                                             r2(r2(daily * worked) * pct / 100.0), 0.0)
            got = M.leave_daily_base(permanent, worked, base_norm, leave_norm)
            want = salary * uplift / leave_norm
            if abs(got - want) > 0.005:
                raise AssertionError(
                    f"чл. 18 НСОРЗ: base month {base_norm} days ({worked} worked), "
                    f"leave month {leave_norm} days - corrected daily {got:.4f}, "
                    f"contracted daily of the leave month {want:.4f}")
    # чл. 18, ал. 1, изр. второ: the agreed fallback divides by the year's average
    # monthly working days. 2026: 21+20+21+20+18+22+23+21+20+22+21+20 = 249 days,
    # average 20.75 - the same table the static fixture's May = 18 anchors.
    cases += 1
    got = M.leave_daily_base_agreed(salary, pct, 2026)
    want = salary * uplift / 20.75
    if abs(got - want) > 0.005:
        raise AssertionError(f"чл. 18, ал. 1, изр. второ: agreed daily {got:.4f}, "
                             f"expected {want:.4f} over the 2026 average of 20.75")
    return cases


def check(xlsx, manifest, quiet=False):
    if isinstance(manifest, dict):
        cfg = manifest
    else:
        with open(manifest, encoding="utf8") as fh:
            cfg = json.load(fh)
    # Told, not inferred - see the note in structural_test.check.
    bonus_in_base = bool((cfg.get("policy") or {}).get("bonus_in_base"))
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    early_name, late_name = wb.sheetnames[0], wb.sheetnames[1]
    _, _, early = read_sheet(wb[early_name])
    late_hdr, _, late = read_sheet(wb[late_name])

    early_month, early_year = (int(x) for x in early_name.split("-"))
    late_month, late_year = (int(x) for x in late_name.split("-"))
    F = Findings()

    # ---------------------------------------------- K8: was the sheet copied?
    early_norm, late_norm = derived_norm(early), derived_norm(late)
    calendar = {early_name: M.working_days(early_year, early_month),
                late_name: M.working_days(late_year, late_month)}
    for name, built_on in ((early_name, early_norm), (late_name, late_norm)):
        if abs(built_on - calendar[name]) < 1e-9:
            continue
        other = calendar[early_name if name == late_name else late_name]
        why = " - the norm of the other sheet's month, so this one was copied forward" \
            if abs(built_on - other) < 1e-9 else ""
        F.add("K8_stale_thresholds", "file",
              f"sheet „{name}“ is built on {built_on:g} working days while its month has "
              f"{calendar[name]:g}{why}", built_on, calendar[name])

    # The thresholds carried with it. The cap of the wrong period is visible whenever a
    # row sits on it, which is the case the copy actually costs money.
    applicable = {name: M.REGIMES[M.regime_for(year, month)]["max_insurable"]
                  for name, year, month in ((early_name, early_year, early_month),
                                            (late_name, late_year, late_month))}
    for name, rows in ((early_name, early), (late_name, late)):
        due = applicable[name]
        other = applicable[early_name if name == late_name else late_name]
        if other == due:
            continue
        # Two guards against reading coincidence as staleness. The cent must be
        # exact, and the row's own gross must exceed the stated insurable income -
        # a genuinely capped row had something cut off, a clean row that merely SUMS
        # to a number near the other period's cap has nothing cut off and its
        # insurable sits at or above its gross (the benefits add in).
        capped = [r for r in rows
                  if abs(r["Осигурителен доход"] - other) < 0.005
                  and r["БРУТО"] > r["Осигурителен доход"] + TOL]
        if capped:
            F.add("K8_stale_thresholds", "file",
                  f"in sheet „{name}“ {len(capped)} row(s) are capped at {other:.2f} - "
                  f"the maximum insurable income of the other period - instead of "
                  f"{due:.2f}", other, due)

    # ------------------------------------- the roster, matched across the sheets
    by_name = {r["Име"]: r for r in early}
    pairs = [(by_name[r["Име"]], r) for r in late if r["Име"] in by_name]

    for before, after in pairs:
        row_no = after["_row"]

        # --- I7: a jump with nothing to explain it -------------------------
        # The comparison is the implied monthly salary, not the gross. A gross may
        # legitimately move with a bonus, with leave, with sick days, or simply because
        # the two months hold a different number of working days. The salary behind it
        # may not, unless something was agreed - and an annex would be in the file.
        was, now = (implied_salary(before, early_norm), implied_salary(after, late_norm))
        if was and now and abs(now - was) > SALARY_TOL:
            F.add("I7_unexplained_jump", row_no,
                  f"the row implies a different monthly base in „{late_name}“ than in "
                  f"„{early_name}“, with the same contract and nothing in the file to "
                  f"explain the change")

        # --- чл. 177 КТ: the base for paid leave ---------------------------
        days_leave = after["Дни платен отпуск"]
        if not days_leave:
            continue
        worked_before = before["Отраб. дни"]
        permanent_before = M.permanent_pay_of(before, bonus_in_base)
        if worked_before >= 10:
            due_daily = M.leave_daily_base(permanent_before, worked_before,
                                           early_norm, late_norm)
        else:
            # чл. 18, ал. 1, изр. второ: no usable base month, so the agreed salary
            # over the year's average monthly working days. The contract is implied
            # from the LEAVE month's own row - its salary-for-days-worked is clean by
            # A6 for exactly this comparison.
            worked_after = after["Отраб. дни"]
            if not worked_after:
                continue
            salary = implied_salary(after, late_norm)
            due_daily = M.leave_daily_base_agreed(salary, after["Клас %"], late_year)
        due = r2(due_daily * days_leave)
        stated = after["Платен отпуск"]
        if abs(stated - due) <= TOL:
            continue
        why = ""
        if worked_before < 10:
            misapplied = M.leave_daily_base(permanent_before, worked_before,
                                            early_norm, late_norm)
            if abs(stated - r2(misapplied * days_leave)) <= TOL:
                why = (f" - from „{early_name}“'s {worked_before:g} worked days, when "
                       f"чл. 18, ал. 1, изр. второ sends a month under 10 days to the "
                       f"agreed salary over the year's average monthly working days")
        elif before["Бонус"]:
            other = M.leave_daily_base(
                r2(permanent_before + (-before["Бонус"] if bonus_in_base
                                       else before["Бонус"])),
                worked_before, early_norm, late_norm)
            if abs(stated - r2(other * days_leave)) <= TOL:
                why = (f" - on a base carrying the bonus of „{early_name}“, which is "
                       f"in none of the seven points of чл. 17, ал. 1 НСОРЗ"
                       if not bonus_in_base else
                       f" - on a base without the bonus of „{early_name}“, which this "
                       f"file pays under a wage system (чл. 17, ал. 1, т. 2 НСОРЗ)")
        F.add("E3_leave_base", row_no,
              f"paid leave for {days_leave:g} days{why}; чл. 177 КТ measures it "
              f"against the average daily gross of „{early_name}“ ({due_daily:.4f})",
              stated, due)

    # ----------------------------------------------------------------- scoring
    man = cfg
    expected = {("file" if where == "file" else late_hdr + 1 + idx, ident)
                for where, idx, ident in man["cross_expected"]}
    found = F.keys()
    missed, extra = expected - found, found - expected

    if not quiet:
        print(f"=== {os.path.basename(xlsx)} · „{early_name}“ + „{late_name}“ · "
              f"{len(pairs)} people · built on {early_norm:g} and {late_norm:g} days "
              f"(the calendar says {calendar[early_name]:g} and "
              f"{calendar[late_name]:g}) ===")
        print(f"\nFINDINGS ({len(F.items)}):")
        for f in sorted(F.items, key=lambda x: (str(x["where"]), x["id"])):
            where = "file" if f["where"] == "file" else f"row {f['where']}"
            print(f"  [{where:8}] {f['id']:24} {f['text']}")
            if f["due"] is not None:
                print(f"{'':13} stated {f['stated']} | due {f['due']}")
        if missed:
            print(f"\nMISSED ({len(missed)}):")
            for where, ident in sorted(missed, key=str):
                print(f"  {where} · {ident} — {M.PAIR_SCENARIOS.get(ident, ('', '?'))[1]}")
        if extra:
            print(f"\nFALSE POSITIVES ({len(extra)}):")
            for where, ident in sorted(extra, key=str):
                print(f"  {where} · {ident}")
        print(f"\ninjected {len(expected)} · found {len(expected & found)} · "
              f"missed {len(missed)} · extra {len(extra)}")

    return dict(injected=len(expected), found=len(expected & found),
                missed=sorted(f"{a}:{b}" for a, b in missed),
                extra=sorted(f"{a}:{b}" for a, b in extra),
                findings=len(F.items))


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--seed", type=int, default=1)
    ap.add_argument("--quiet", action="store_true")
    a = ap.parse_args()

    sys.path.insert(0, HERE)
    import generate_pair as P

    xlsx, manifest_path, _ = P.generate(a.seed)
    result = check(xlsx, manifest_path, quiet=a.quiet)
    sys.exit(0 if not result["missed"] and not result["extra"] else 1)
