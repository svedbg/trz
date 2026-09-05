#!/usr/bin/env python3
"""Payroll workbooks broken by *shape*, one defect each, for the pre-flight suite.

The five generated suites break the numbers. These break the file: a renamed header, a
merged cell, an export that lost its formulas, a tab with no month on it. They are the
reasons a real audit stalls before it reaches a single figure, and each one has to be
caught by tools/preflight.py exactly once with nothing else raised.

Every value here is invented. No real payroll is read, and nothing personal is written:
the name column holds „Лице 1", „Лице 2" and so on, which is also what the suite uses to
prove no cell value from that column reaches the report.

    python test/generate_shapes.py --list
    python test/generate_shapes.py S4_merged_in_data --out /tmp/broken.xlsx
"""

import argparse
import os
import re
import shutil
import sys
import zipfile

import openpyxl

# The clean layout every shape starts from: one recognised header per required concept,
# plus the optional ones the DEPENDS map mentions, so a clean file raises no signal at
# all. That silence is what makes "and nothing else" meaningful.
HEADERS = ["Име", "Отраб. дни", "Основна за отработеното", "Клас сума", "БРУТО",
           "Осигурителен доход", "Данъчна основа", "ДДФЛ", "Лични вноски общо",
           "НЕТО за изплащане", "Изплатено", "Дни болничен", "Болнични (работодател)",
           "Вноски работодател общо", "Дни платен отпуск"]

ROWS = [
    ["Лице 1", 21, 1000.00, 50.00, 1050.00, 1050.00, 945.00, 94.50, 105.00, 850.50,
     850.50, 0, 0.00, 200.00, 0],
    ["Лице 2", 18, 1200.00, 96.00, 1296.00, 1296.00, 1166.40, 116.64, 129.60, 1049.76,
     1049.76, 3, 84.00, 240.00, 0],
    ["Лице 3", 21, 900.00, 18.00, 918.00, 918.00, 826.20, 82.62, 91.80, 743.58,
     743.58, 0, 0.00, 175.00, 0],
]

CLEAN_SHEET = "05.2026"          # inside a regime that starts on 1 January
HEADER_ROW = 3                   # a title above it, the way real files are laid out


def cache_formula_values(path, cached, sheets=("xl/worksheets/sheet1.xml",)):
    """Fill in the cached results Excel stores next to each formula.

    openpyxl writes `<f>C4+D4</f><v />` - the formula with an empty cached value - so a
    workbook it produced reads as None behind every formula with data_only=True. Excel
    itself always stores the last computed result there. Patching it in makes the
    fixture behave like a real file, which matters because the difference between the
    two *is* one of the shapes under test: S10 is this same workbook with the patch
    withheld.

    `cached` maps a cell reference to the value to store, e.g. {"E4": 1050.0}. `sheets`
    names the worksheet parts to patch - more than one when a shape adds a second sheet
    that also carries formulas, since a sheet left unpatched raises NO_CACHED_VALUES and
    the shape would then be testing two things.
    """
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            blob = zin.read(item.filename)
            if item.filename in sheets:
                xml = blob.decode("utf8")
                for ref, value in cached.items():
                    # The cell element for this ref, up to its closing tag; only the
                    # empty <v/> inside it is replaced, so nothing else can be touched.
                    pattern = re.compile(r'(<c r="%s"[^>]*>.*?)<v\s*/>(.*?</c>)' % ref)
                    xml = pattern.sub(lambda m: f"{m.group(1)}<v>{value}</v>{m.group(2)}",
                                      xml, count=1)
                blob = xml.encode("utf8")
            zout.writestr(item, blob)
    shutil.move(tmp, path)
    return path


def clean(sheet=CLEAN_SHEET, header_row=HEADER_ROW, formulas=True):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.cell(1, 1, "ВЕДОМОСТ ЗА РАБОТНИ ЗАПЛАТИ — Измислено ЕООД")
    for c, h in enumerate(HEADERS, start=1):
        ws.cell(header_row, c, h)
    for i, row in enumerate(ROWS):
        for c, v in enumerate(row, start=1):
            ws.cell(header_row + 1 + i, c, v)
    if formulas:
        # БРУТО computed, so formula coverage is non-zero on a clean file.
        gross = HEADERS.index("БРУТО") + 1
        base = HEADERS.index("Основна за отработеното") + 1
        klas = HEADERS.index("Клас сума") + 1
        for i in range(len(ROWS)):
            r = header_row + 1 + i
            ws.cell(r, gross,
                    f"={openpyxl.utils.get_column_letter(base)}{r}"
                    f"+{openpyxl.utils.get_column_letter(klas)}{r}")
    ws.cell(header_row + 1 + len(ROWS), 1, "Общо")
    return wb


def _rename(wb, old, new):
    ws = wb.active
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(HEADER_ROW, c).value or "").strip() == old:
            ws.cell(HEADER_ROW, c, new)
            return
    raise AssertionError(f"{old!r} is not in the clean layout")


def s_no_header(wb):
    ws = wb.active
    for c in range(1, ws.max_column + 1):
        ws.cell(HEADER_ROW, c, f"кол{c}")


def s_no_period(wb):
    wb.active.title = "Лист1"
    wb.active.cell(1, 1, "ВЕДОМОСТ ЗА РАБОТНИ ЗАПЛАТИ")


def s_values_only(wb):
    ws = wb.active
    gross = HEADERS.index("БРУТО") + 1
    for i, row in enumerate(ROWS):
        ws.cell(HEADER_ROW + 1 + i, gross, row[gross - 1])


def s_merged(wb):
    wb.active.merge_cells(start_row=HEADER_ROW + 1, start_column=1,
                          end_row=HEADER_ROW + 1, end_column=2)


def s_no_totals(wb):
    # .value = None, not cell(r, c, None): openpyxl reads a None third argument as
    # "no value passed" and leaves the cell alone, so the row survived and this shape
    # quietly tested nothing until the suite caught it.
    wb.active.cell(HEADER_ROW + 1 + len(ROWS), 1).value = None


def s_missing_required(wb):
    # Deleted, not renamed: a rename leaves an unrecognised header behind and would
    # raise UNKNOWN_COLUMNS too, so the shape would no longer test one thing.
    ws = wb.active
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(HEADER_ROW, c).value or "").strip() == "Осигурителен доход":
            ws.delete_cols(c)
            return
    raise AssertionError("the clean layout has no 'Осигурителен доход' to delete")


def s_duplicate(wb):
    # A second column that means gross as well - the file cannot say which one is read.
    _rename(wb, "Изплатено", "Брутно възнаграждение")


def s_unknown_columns(wb):
    _rename(wb, "Вноски работодател общо", "Кол. 14 (стара)")


def s_mid_year(wb):
    wb.active.title = "08.2026"


def s_error_cells(wb):
    """A formula that never resolved, the way a real export carries one.

    openpyxl cannot produce a cached error, so the value is written directly - which is
    also exactly what Excel stores: the error is a string beside the formula.
    """
    ws = wb.active
    col = HEADERS.index("Осигурителен доход") + 1
    ws.cell(HEADER_ROW + 2, col, "#N/A")


def s_hidden_rows(wb):
    """A row inside the data block that the printed sheet does not show."""
    wb.active.row_dimensions[HEADER_ROW + 2].hidden = True


def s_hidden_sheet(wb):
    """A second, well-formed month, hidden.

    Deliberately a complete sheet rather than a scratch tab: a hidden scratch tab has no
    header and raises NO_HEADER as well, and then the fixture would be testing two things
    at once. What is under test is that hiding, by itself, is reported - so everything
    else about the sheet is in order.
    """
    ws = wb.create_sheet("04.2026")
    ws.cell(1, 1, "ВЕДОМОСТ ЗА РАБОТНИ ЗАПЛАТИ — Измислено ЕООД")
    for c, h in enumerate(HEADERS, start=1):
        ws.cell(HEADER_ROW, c, h)
    for i, row in enumerate(ROWS):
        for c, v in enumerate(row, start=1):
            ws.cell(HEADER_ROW + 1 + i, c, v)
    gross = HEADERS.index("БРУТО") + 1
    base = HEADERS.index("Основна за отработеното") + 1
    klas = HEADERS.index("Клас сума") + 1
    for i in range(len(ROWS)):
        r = HEADER_ROW + 1 + i
        ws.cell(r, gross, f"={openpyxl.utils.get_column_letter(base)}{r}"
                          f"+{openpyxl.utils.get_column_letter(klas)}{r}")
    ws.cell(HEADER_ROW + 1 + len(ROWS), 1, "Общо")
    ws.sheet_state = "hidden"


def s_hidden_columns(wb):
    """A recognised column hidden from the printed sheet."""
    col = openpyxl.utils.get_column_letter(HEADERS.index("Осигурителен доход") + 1)
    wb.active.column_dimensions[col].hidden = True


def s_numbers_as_text(wb):
    """A money column stored as text: prints like a number, sums to zero.

    Deliberately not БРУТО - that is the one computed column in the clean layout, and
    overwriting it with text removes the file's only formulas, so the shape would raise
    NO_FORMULAS as well and test two things.
    """
    ws = wb.active
    col = HEADERS.index("Данъчна основа") + 1
    for i, row in enumerate(ROWS):
        ws.cell(HEADER_ROW + 1 + i, col, f"{row[col - 1]:.2f}")


def s_no_cached_values(wb):
    # Nothing to change in the workbook: this shape is the clean file with the cached
    # values never patched in - what a script that writes .xlsx produces and Excel
    # never does. The formulas are there; the numbers behind them are not.
    pass


SHAPES = {
    "S1_no_header":         (s_no_header,      "NO_HEADER",
                             "the header row is unrecognisable"),
    "S2_no_period":         (s_no_period,      "NO_PERIOD",
                             "the sheet does not say which month it is"),
    "S3_values_only":       (s_values_only,    "NO_FORMULAS",
                             "an export that lost every formula"),
    "S4_merged_in_data":    (s_merged,         "MERGED_IN_DATA",
                             "merged cells inside the data block"),
    "S5_no_totals":         (s_no_totals,      "NO_TOTALS",
                             "no totals row to reconcile against"),
    "S6_missing_required":  (s_missing_required, "MISSING_REQUIRED",
                             "a required column is not there under any name"),
    "S7_duplicate_concept": (s_duplicate,      "DUPLICATE_CONCEPT",
                             "two columns mean the same quantity"),
    "S8_unknown_columns":   (s_unknown_columns, "UNKNOWN_COLUMNS",
                             "a column nothing recognises"),
    "S9_mid_year":          (s_mid_year,       "MID_YEAR_BOUNDARY",
                             "a month inside a regime that starts mid-year"),
    "S10_no_cached_values": (s_no_cached_values, "NO_CACHED_VALUES",
                             "formulas whose computed results were never stored"),
    "S11_error_cells":      (s_error_cells,     "ERROR_CELLS",
                             "a cell holding #N/A where money should be"),
    "S12_hidden_rows":      (s_hidden_rows,     "HIDDEN_ROWS",
                             "a row inside the data block that does not print"),
    "S13_hidden_sheet":     (s_hidden_sheet,    "HIDDEN_SHEET",
                             "a hidden sheet in the workbook"),
    "S14_hidden_columns":   (s_hidden_columns,  "HIDDEN_COLUMNS",
                             "a recognised column hidden from the printed sheet"),
    "S15_numbers_as_text":  (s_numbers_as_text, "NUMBERS_AS_TEXT",
                             "a money column stored as text"),
}


def gross_cache(header_row=HEADER_ROW):
    """{cell ref: value} for the computed БРУТО column of the clean layout."""
    col = openpyxl.utils.get_column_letter(HEADERS.index("БРУТО") + 1)
    return {f"{col}{header_row + 1 + i}": row[HEADERS.index("БРУТО")]
            for i, row in enumerate(ROWS)}


def save(wb, path, cached=True):
    """Save a workbook the way Excel would: formulas with their computed results.

    Every fixture goes through here, so „clean" means clean in the one respect that
    S10 exists to test as well.
    """
    wb.save(path)
    if cached:
        cache_formula_values(path, gross_cache())
    return path


def build(shape_id, path, cached=True):
    """Write the clean workbook with exactly one shape defect applied."""
    if shape_id not in SHAPES:
        raise KeyError(shape_id)
    wb = clean()
    SHAPES[shape_id][0](wb)
    wb.save(path)
    # S10 is defined by the absence of the cached values, and S1/S3 rewrite or bury the
    # column they would be written to, so the patch is skipped for those.
    if cached and shape_id not in ("S10_no_cached_values", "S3_values_only",
                                   "S1_no_header"):
        parts = ("xl/worksheets/sheet1.xml", "xl/worksheets/sheet2.xml") \
            if shape_id == "S13_hidden_sheet" else ("xl/worksheets/sheet1.xml",)
        cache_formula_values(path, gross_cache(), parts)
    return path


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("shape", nargs="?")
    ap.add_argument("--list", action="store_true")
    ap.add_argument("--out", default="broken.xlsx")
    a = ap.parse_args(argv)
    if a.list or not a.shape:
        for k, (_, signal, why) in SHAPES.items():
            print(f"{k:24} -> {signal:20} {why}")
        return 0
    build(a.shape, a.out)
    print(f"{a.shape} -> {os.path.abspath(a.out)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
