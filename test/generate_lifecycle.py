#!/usr/bin/env python3
"""Five months of payroll for the same people, with the documents behind the events.

`I11` — one person's timeline across months — is the last check `proverki.md` describes
that no fixture could reach. The комплект proves the chain *below* one month; this proves
the chain *along* the months, which is a different axis and needs a different fixture:
several sheets for the same people, plus the events that are supposed to explain what
changes between them.

    вписване · анекс · заповед за отпуск · болничен лист · заповед за прекратяване
                                  ↓
        05.2026 → 06.2026 → 07.2026 → 08.2026 → 09.2026

Every break here is a **timeline** break, not an arithmetic one: each month is internally
correct and reconciles with itself, and the only thing that disagrees is the sequence of
events or the document that should stand behind a change. That is exactly why I11 exists
and why nothing that checks a single sheet can find these.

The breaks map one-to-one onto the bullets I11 lists in `proverki.md`, so a break that
stops corresponding to a sentence in the skill is a break that should be deleted:

    заплатата се променя без допълнително споразумение   -> I11_salary_change_without_annex
    обезщетение, а лицето получава заплата и след това   -> I11_pay_after_termination
    обезщетение при прекратяване без самото прекратяване -> I11_severance_without_termination
    болничен, който започва отново от първия ден         -> I11_sick_days_restart
    клас, който се вдига без навършена година            -> I11_class_raised_early
    клас, който не се вдига, когато е навършена          -> I11_class_not_raised

Nothing here is personal data: „Лице 1"…, служебни номера `СЛ-001`, invented salaries,
and every amount computed by `trz_model`, which mirrors `references/stavki.md`.

    python test/generate_lifecycle.py --list
    python test/generate_lifecycle.py --clean --out /tmp/lifecycle
"""

import argparse
import csv
import datetime
import os
import random
import sys

import openpyxl
from openpyxl.styles import Font

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

import trz_model as M                                           # noqa: E402

r2 = M.r2

YEAR = 2026
MONTHS = [5, 6, 7, 8, 9]
HDR = 4
TZPB = 0.5
POLICY = dict(bonus_in_base=False, in_kind_in_bases=False,
              excess_in_insurable=False, excess_in_taxable=False, excess_reading="А")

# Class rises on a completed year of service, so the anniversary is what the check
# measures against - not the calendar year and not the month the payroll happens to run.
CLASS_PER_YEAR = 0.6            # `stavki.md`, „клас"; the model carries the same figure


def seniority_pct(start, on):
    """Percentage due on a date: completed years of service times the yearly rate."""
    years = on.year - start.year - ((on.month, on.day) < (start.month, start.day))
    return round(max(0, years) * CLASS_PER_YEAR, 2)


def month_end(year, month):
    return datetime.date(year, month, 28) + datetime.timedelta(days=4 - 1) \
        if month == 2 else datetime.date(year, month,
                                         [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1])


def people_for(seed):
    """Five people, each with a timeline. The events are the documents; the payroll
    below is what those events should produce."""
    rnd = random.Random(seed)
    out = []
    for i in range(1, 6):
        salary = r2(rnd.uniform(1200, 2400))
        # Two people have an anniversary inside the window, and which two is fixed
        # rather than drawn: person 4 is the one who leaves at the end of August, so an
        # anniversary on them would fall in a month that no longer has a row and the
        # class breaks could never fire. Persons 2 and 3 carry them, in June and July,
        # so the raise lands in a month that exists and is not the last.
        anniversary_month = {2: 6, 3: 7}.get(i, 11)
        start = datetime.date(YEAR - rnd.randint(2, 6), anniversary_month,
                              rnd.randint(2, 20))
        p = dict(index=i, name=f"Лице {i}", ident=f"СЛ-{i:03d}",
                 hired=datetime.date(YEAR - 5, 3, 1), seniority_start=start,
                 salary=salary, annexes=[], leaves={}, sick={}, termination=None,
                 severance=0.0)
        out.append(p)

    # Person 1: a raise from 1 July, with the annex that says so.
    out[0]["annexes"].append(dict(date=datetime.date(YEAR, 7, 1),
                                  salary=r2(out[0]["salary"] * 1.15)))
    # Person 3: five days' paid leave in August, with the order behind it.
    out[2]["leaves"][8] = 5
    # Person 5: one spell of sick leave running from August into September. The first
    # two working days are the employer's and they fall in August only.
    out[4]["sick"][8] = 4
    out[4]["sick"][9] = 3
    # Person 4: terminated at the end of August, severance paid in that month.
    out[3]["termination"] = datetime.date(YEAR, 8, 31)
    out[3]["severance"] = r2(out[3]["salary"])
    return out


def salary_on(p, year, month):
    """The agreed salary for a month: the last annex effective on or before its first day."""
    first = datetime.date(year, month, 1)
    salary = p["salary"]
    for a in sorted(p["annexes"], key=lambda a: a["date"]):
        if a["date"] <= first:
            salary = a["salary"]
    return salary


def build_months(people, overrides=None):
    """Compute every month for every person. `overrides` breaks one thing on purpose."""
    overrides = overrides or {}
    sheets = []
    for month in MONTHS:
        norm = M.working_days(YEAR, month)
        regime = M.REGIMES[M.regime_for(YEAR, month)]
        rows = []
        for p in people:
            over = overrides.get((p["ident"], month), {})
            if p["termination"] and datetime.date(YEAR, month, 1) > p["termination"] \
                    and not over.get("keep_paying"):
                continue
            salary = over.get("salary", salary_on(p, YEAR, month))
            pct = over.get("seniority_pct",
                           seniority_pct(p["seniority_start"], datetime.date(YEAR, month, 1)))
            leave = p["leaves"].get(month, 0)
            sick = over.get("sick", p["sick"].get(month, 0))
            # Employer-paid days belong to the month the spell STARTS in. A spell
            # running on from the previous month is the fund's from its first day here,
            # and reading that from the person's own history rather than from a flag is
            # what makes the clean fixture clean: the first version paid the employer's
            # two days in both months and the suite caught its own generator.
            continuing = sick > 0 and p["sick"].get(month - 1, 0) > 0
            employer_days = min(2, sick) if (sick and not continuing) else 0
            if over.get("employer_days") is not None:
                employer_days = over["employer_days"]
            worked = norm - leave - sick
            inp = dict(monthly_salary=salary, seniority_pct=pct,
                       days_worked=worked, days_leave=leave, days_sick=employer_days,
                       days_maternity=0, bonus=0.0,
                       compensation_224=over.get("severance",
                                                 p["severance"] if p["termination"]
                                                 and p["termination"].month == month else 0.0),
                       card_employer=0.0, card_employee=0.0, premium=0.0,
                       personal_contribution=0.0, life_premium_personal=0.0)
            row = M.clean_row(inp, regime, TZPB, POLICY, norm)
            rows.append(dict(person=p, row=row, salary=salary, pct=pct,
                             sick_days=sick, employer_days=employer_days,
                             leave_days=leave, norm=norm,
                             severance=inp["compensation_224"]))
        sheets.append(dict(month=month, norm=norm, rows=rows))
    return sheets


# =====================================================================
#                              the breaks
# Each returns (overrides, events-mutation), so a break can move a figure
# in the payroll, remove the document that explains it, or both.
# =====================================================================

def b_salary_change_without_annex(people):
    """Person 2's pay rises in August and no annex says so."""
    p = people[1]
    return {(p["ident"], m): {"salary": r2(p["salary"] * 1.2)} for m in (8, 9)}, None


def b_pay_after_termination(people):
    """Person 4 is terminated on 31 August and paid a salary in September anyway."""
    p = people[3]
    return {(p["ident"], 9): {"keep_paying": True}}, None


def b_severance_without_termination(people):
    """Person 1 is paid a чл. 224 compensation with no termination behind it."""
    p = people[0]
    return {(p["ident"], 9): {"severance": r2(p["salary"] * 0.5)}}, None


def b_sick_days_restart(people):
    """Person 5's continuing spell pays the employer's first days twice."""
    p = people[4]
    return {(p["ident"], 9): {"employer_days": 2}}, None


def b_class_raised_early(people):
    """Person 2's class rises a month before the anniversary."""
    p = people[1]
    early = seniority_pct(p["seniority_start"],
                          datetime.date(YEAR, p["seniority_start"].month, 1)) + CLASS_PER_YEAR
    month = max(MONTHS[0], p["seniority_start"].month - 1)
    return {(p["ident"], m): {"seniority_pct": round(early, 2)}
            for m in MONTHS if m >= month}, None


def b_class_not_raised(people):
    """Person 3's anniversary passes and the class stays where it was."""
    p = people[2]
    before = seniority_pct(p["seniority_start"], datetime.date(YEAR, MONTHS[0], 1))
    return {(p["ident"], m): {"seniority_pct": before} for m in MONTHS}, None


BREAKS = {
    "I11_salary_change_without_annex": (b_salary_change_without_annex,
                                        "pay rises between months with no annex"),
    "I11_pay_after_termination":       (b_pay_after_termination,
                                        "a salary in a month after the termination date"),
    "I11_severance_without_termination": (b_severance_without_termination,
                                          "чл. 224 compensation with no termination"),
    "I11_sick_days_restart":           (b_sick_days_restart,
                                        "a continuing spell pays the employer's days twice"),
    "I11_class_raised_early":          (b_class_raised_early,
                                        "the class rises before the anniversary"),
    "I11_class_not_raised":            (b_class_not_raised,
                                        "the anniversary passes and the class does not move"),
}

# Mutually exclusive members share a person or a figure; one break per group at a time.
GROUPS = [("I11_salary_change_without_annex", "I11_class_raised_early"),
          ("I11_pay_after_termination", "I11_class_not_raised"),   # persons 4 and 3
          ("I11_severance_without_termination",),
          ("I11_sick_days_restart",)]

assert set(b for g in GROUPS for b in g) == set(BREAKS)


def breaks_for_seed(seed, groups=3):
    """Which timeline breaks a seed plants. One function, two callers - as in the комплект."""
    rnd = random.Random(seed * 5077)
    return [rnd.choice(g) for g in rnd.sample(GROUPS, groups)]


def events_of(people):
    """The documents: what the employer can show for each change."""
    rows = []
    for p in people:
        rows.append([p["ident"], p["hired"].isoformat(), "постъпване",
                     f"основна заплата {p['salary']:.2f}"])
        rows.append([p["ident"], p["seniority_start"].isoformat(), "признат трудов стаж",
                     "начало на стажа за клас"])
        for a in p["annexes"]:
            rows.append([p["ident"], a["date"].isoformat(),
                         "допълнително споразумение", f"нова основна заплата {a['salary']:.2f}"])
        for month, days in sorted(p["leaves"].items()):
            rows.append([p["ident"], datetime.date(YEAR, month, 1).isoformat(),
                         "заповед за отпуск", f"{days} работни дни"])
        for month, days in sorted(p["sick"].items()):
            rows.append([p["ident"], datetime.date(YEAR, month, 1).isoformat(),
                         "болничен лист", f"{days} дни" +
                         (", продължение" if month != min(p["sick"]) else "")])
        if p["termination"]:
            rows.append([p["ident"], p["termination"].isoformat(),
                         "заповед за прекратяване", "чл. 71, ал. 1 КТ"])
    return sorted(rows, key=lambda r: (r[0], r[1]))


def lifecycle(seed=1, break_ids=()):
    people = people_for(seed)
    overrides = {}
    for ident in BREAKS:
        if ident in break_ids:
            over, _ = BREAKS[ident][0](people)
            for k, v in over.items():
                overrides.setdefault(k, {}).update(v)
    return dict(people=people, sheets=build_months(people, overrides),
                events=events_of(people), breaks=list(break_ids))


def write(lc, outdir):
    os.makedirs(outdir, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet in lc["sheets"]:
        ws = wb.create_sheet(f"{sheet['month']:02d}-{YEAR}")
        ws["A1"] = f'ВЕДОМОСТ {sheet["month"]:02d}.{YEAR} — "Измислено" ЕООД'
        ws["A2"] = (f"Работни дни: {sheet['norm']}  |  Валута: EUR  |  "
                    f"ЕИК: 000000000 (тестов)  |  ТЗПБ по КИД: {TZPB}%")
        ws["A1"].font = Font(bold=True, size=12)
        for i, column in enumerate(M.COLUMNS, start=1):
            ws.cell(row=HDR, column=i, value=column).font = Font(bold=True, size=8)
        for offset, item in enumerate(sheet["rows"]):
            r = HDR + 1 + offset
            ws.cell(row=r, column=M.COL["№"], value=offset + 1)
            ws.cell(row=r, column=M.COL["Име"], value=item["person"]["name"])
            ws.cell(row=r, column=M.COL["Отдел"], value=item["person"]["ident"])
            for column in M.COLUMNS:
                if column in ("№", "Име", "Отдел"):
                    continue
                v = item["row"].get(column, 0)
                ws.cell(row=r, column=M.COL[column],
                        value=(v if v else (0 if column in M.DAY_COLUMNS else None)))
        total = HDR + 1 + len(sheet["rows"])
        ws.cell(row=total, column=M.COL["Име"], value="ОБЩО").font = Font(bold=True)
        for column in M.SUMMED_COLUMNS:
            ws.cell(row=total, column=M.COL[column],
                    value=r2(sum(x["row"].get(column, 0) or 0 for x in sheet["rows"])))
    path = os.path.join(outdir, "vedomosti.xlsx")
    wb.save(path)

    with open(os.path.join(outdir, "sabitiya.csv"), "w", encoding="utf8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["идентификатор", "дата", "събитие", "детайл"])
        w.writerows(lc["events"])
    with open(os.path.join(outdir, "dogovori.csv"), "w", encoding="utf8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["идентификатор", "име", "основна заплата по договор",
                    "начало на признатия трудов стаж"])
        for p in lc["people"]:
            w.writerow([p["ident"], p["name"], f"{p['salary']:.2f}",
                        p["seniority_start"].isoformat()])
    return dict(dir=outdir, months=[s["month"] for s in lc["sheets"]], year=YEAR,
                header_row=HDR, tzpb=TZPB, people=len(lc["people"]),
                breaks=lc["breaks"])


def build(break_ids, outdir, seed=1):
    ids = [] if break_ids is None else ([break_ids] if isinstance(break_ids, str)
                                        else list(break_ids))
    unknown = [i for i in ids if i not in BREAKS]
    if unknown:
        raise KeyError(", ".join(unknown))
    lc = lifecycle(seed, ids)
    return lc, write(lc, outdir)


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("break_id", nargs="?")
    ap.add_argument("--list", action="store_true")
    ap.add_argument("--clean", action="store_true")
    ap.add_argument("--seed", type=int, default=1)
    ap.add_argument("--out", default="lifecycle")
    a = ap.parse_args(argv)
    if a.list or not (a.break_id or a.clean):
        for k, (_, why) in BREAKS.items():
            print(f"{k:36} {why}")
        return 0
    _, man = build(None if a.clean else a.break_id, a.out, a.seed)
    print(f"{man['breaks'] or 'clean'} -> {os.path.abspath(a.out)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
