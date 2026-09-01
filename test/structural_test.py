# -*- coding: utf-8 -*-
"""Checks over a generated payroll (the structural suite).

    python test/structural_test.py --seed 7            # generate and check
    python test/structural_test.py --seed 7 --quiet    # result only

The suite is a round trip: `generate_wide.py` builds a correct payroll and breaks
it in known places, and this file must find exactly what was broken - no less (a
missed finding) and no more (a false positive). False positives fail the suite
like misses: a skill that sees violations everywhere is as useless as one that
sees none.

What the suite deliberately does not do: settle the contested questions. Whether
a benefit in kind belongs in the insurable income has more than one defensible
reading (proverki.md, F10). So the check is for **consistency** - the file's
practice is inferred from the file itself and the rows deviating from it are
sought. Same for the excess over the social-expense threshold.

The method for F1/F6/F10 is "solving the composition": which subset of the
accruals and benefits explains the stated insurable income, and which explains
the stated taxable base. If the same element is inside one base and outside the
other, at least one of the two is wrong - whatever the reading. That is the
finding which needs no ruling on the contested question.

Findings text keeps Bulgarian column names and statutory citations, because both
are quoted from the domain. Everything else is English.
"""
import argparse
import json
import os
import sys
from collections import Counter

import openpyxl

import trz_model as M
from trz_model import r2

HERE = os.path.dirname(os.path.abspath(__file__))


def _subsets(elements):
    """All subsets of [(name, value)] as [(frozenset, sum)]."""
    out = [(frozenset(), 0.0)]
    for name, v in elements:
        if not v:
            continue
        out += [(frozenset(m | {name}), r2(s + v)) for m, s in out]
    return out


class Findings:
    def __init__(self):
        self.items = []
        self._seen = set()

    def add(self, ident, where, text, stated=None, due=None):
        if (where, ident) in self._seen:
            return                     # one finding per (location, kind)
        self._seen.add((where, ident))
        self.items.append(dict(id=ident, where=where, text=text, stated=stated, due=due))

    def keys(self):
        return {(f["where"], f["id"]) for f in self.items}


def check(xlsx, manifest, quiet=False):
    man = json.load(open(manifest, encoding="utf8"))
    # The configured reading of чл. 17, ал. 1 for an uncharacterised bonus column.
    # The auditor is told this one; only the file's own practices are inferred.
    bonus_in_base = bool((man.get("policy") or {}).get("bonus_in_base"))
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[man["sheet"]]
    HDR, TOTAL = man["hdr"], man["total_row"]
    norm = man["norm_days"]
    max_insurable = man["max_insurable"]
    min_insurable_self = man["min_insurable_self"]
    tzpb_due = man["tzpb_due"]
    other_cap = [v["max_insurable"] for v in M.REGIMES.values()
                 if v["max_insurable"] != max_insurable][0]

    col = {ws.cell(row=HDR, column=c).value: c for c in range(1, ws.max_column + 1)}

    def get(r, name):
        v = ws.cell(row=r, column=col[name]).value
        return 0.0 if v in (None, "") else v

    rows = list(range(HDR + 1, TOTAL))
    contracts = {p["row"]: p["inputs"] for p in man["people"]}
    F = Findings()
    data = []

    # ==================================================== pass 1: row by row
    for r in rows:
        v = {k: get(r, k) for k in M.COLUMNS}
        wd, pl, sd, md = (v[k] for k in M.DAY_COLUMNS)
        base = v["Основна за отработеното"]
        pct, seniority = v["Клас %"], v["Клас сума"]
        bonus, leave = v["Бонус"], v["Платен отпуск"]
        comp_224, sick_pay = v["Обезщетение чл. 224"], v["Болнични (работодател)"]
        gross, employee_total = v["БРУТО"], v["Лични вноски общо"]
        insurable, taxable, tax = (v["Осигурителен доход"], v["Данъчна основа"],
                                   v["ДДФЛ"])
        deduction = v["Удръжка доброволно осиг. (лична)"]
        deduction_life = v["Удръжка застраховка Живот (лична)"]
        card_deduction = v["Удръжка карта (лична част)"]
        net_before, net = v["НЕТО преди удръжки"], v["НЕТО за изплащане"]
        paid, control = v["Изплатено"], v["Разлика"]
        er_social = v["Вноски работодател ДОО+ТЗПБ"]
        er_health_sick = v["ЗО при болничен/майчинство"]
        er_total = v["Вноски работодател общо"]
        in_kind = v["Карта (за сметка на работодателя)"]
        premium = v["Доброволно здравно осигуряване (премия)"]
        cost = v["Общ разход за труд"]
        excess = r2(max(0.0, premium - M.SOCIAL_EXPENSE_THRESHOLD)) if premium else 0.0

        # --- K2: an amount in a column meant for days -------------------
        for day_col in M.DAY_COLUMNS:
            x = v[day_col]
            if x and (abs(x - round(x)) > 1e-9 or x > norm):
                F.add("K2_amount_in_day_column", r,
                      f"day column holds {x} against a norm of {norm} - that is an "
                      f"amount, not days", x, None)
                break

        # --- I5: the day counts ----------------------------------------
        day_sum = wd + pl + sd + md
        days_ok = abs(day_sum - norm) < 1e-9
        if not days_ok:
            F.add("I5_days_do_not_reconcile", r,
                  f"days {wd}+{pl}+{sd}+{md} = {day_sum} against a norm of {norm} "
                  f"working days", day_sum, norm)

        # --- K6: unrounded accruals -------------------------------------
        for accrual in M.ACCRUALS + ["БРУТО"]:
            if v[accrual] and abs(v[accrual] - r2(v[accrual])) > 1e-9:
                F.add("K6_unrounded_accrual", r,
                      f"accrual with more than two decimals: {v[accrual]!r}",
                      v[accrual], r2(v[accrual]))
                break

        # --- K1: gross = the sum of the accruals ------------------------
        accrual_sum = r2(sum(v[k] for k in M.ACCRUALS))
        if abs(gross - accrual_sum) > M.TOL:
            missing = [k for k in M.ACCRUALS
                       if v[k] and abs(r2(accrual_sum - v[k]) - gross) <= M.TOL]
            F.add("K1_sum_omits_column", r,
                  "БРУТО is not the sum of the accruals"
                  + (f"; the column left out is „{missing[0]}“" if missing else ""),
                  gross, accrual_sum)

        # --- K3: contributions as a percentage of the insurable income ---
        wrong = [(c, v[c], r2(insurable * M.EMPLOYEE[k] / 100.0))
                 for c, k in M.EMPLOYEE_COLUMNS
                 if abs(v[c] - r2(insurable * M.EMPLOYEE[k] / 100.0)) > M.TOL]
        if wrong:
            F.add("K3_stale_contributions", r,
                  f"contributions are not a percentage of the insurable income "
                  f"{insurable:.2f} (first mismatch: „{wrong[0][0]}“)",
                  wrong[0][1], wrong[0][2])
        else:
            # The five separately rounded contributions can drift up to 0.03 from
            # 13.78% of the insurable income - 0.005 per component plus the
            # rounding of the reference value itself. The exact control is the sum
            # of the components; the percentage is indicative.
            component_sum = r2(sum(v[c] for c, _ in M.EMPLOYEE_COLUMNS))
            if abs(employee_total - component_sum) > M.TOL_STRICT:
                F.add("K3_stale_contributions", r,
                      "„Лични вноски общо“ is not the sum of the five contributions",
                      employee_total, component_sum)
            elif abs(employee_total - r2(insurable * M.EMPLOYEE_TOTAL / 100.0)) > 0.03:
                F.add("K3_stale_contributions", r,
                      f"employee contributions are not {M.EMPLOYEE_TOTAL}% of "
                      f"{insurable:.2f}", employee_total,
                      r2(insurable * M.EMPLOYEE_TOTAL / 100.0))

        # --- I1 / F6: vertical reconciliation ---------------------------
        if abs(net_before - r2(gross - employee_total - tax)) > M.TOL:
            F.add("I1_vertical", r,
                  "net before deductions is not gross minus contributions minus tax",
                  net_before, r2(gross - employee_total - tax))
        withheld = r2(deduction + deduction_life + card_deduction)
        if abs(net - r2(net_before - withheld)) > M.TOL:
            F.add("I1_vertical", r,
                  "net payable is not net before deductions minus the deductions",
                  net, r2(net_before - withheld))
        if abs(tax - r2(taxable * M.TAX_RATE)) > M.TOL:
            F.add("F6_tax_amount", r,
                  f"tax is not {M.TAX_RATE:.0%} of the taxable base",
                  tax, r2(taxable * M.TAX_RATE))

        # --- K4: the control column -------------------------------------
        if abs(control - r2(net - paid)) > M.TOL_STRICT:
            F.add("K4_control_column_blind", r,
                  f"column „Разлика“ reads {control:.2f} while net minus paid is "
                  f"{r2(net - paid):.2f}; the control does not register the gap",
                  control, r2(net - paid))

        # --- K7: the cost of labour -------------------------------------
        due_cost = r2(gross + er_total + in_kind + premium)
        if abs(cost - due_cost) > M.TOL:
            withheld = r2(deduction + card_deduction)
            why = " - short by exactly what was withheld from the employee" \
                if withheld and abs(r2(due_cost - withheld) - cost) <= M.TOL else ""
            F.add("K7_cost_from_net", r,
                  "cost of labour is not gross plus employer contributions plus "
                  "benefits" + why, cost, due_cost)

        # --- checks that depend on the days and on the contract ----------
        if days_ok and r in contracts:
            c = contracts[r]
            daily = c["monthly_salary"] / norm
            uplift = 1 + pct / 100.0
            base_due = r2(daily * wd)
            leave_due = r2(daily * uplift * pl) if pl else 0.0
            if abs(base - base_due) > M.TOL:
                F.add("A6_base_vs_contract", r,
                      "the base for days worked does not match the contracted salary",
                      base, base_due)
            if pl:
                if abs(leave - leave_due) > M.TOL:
                    without = r2(daily * pl)
                    why = " - without the seniority uplift" \
                        if abs(leave - without) <= M.TOL else ""
                    F.add("E3_leave_without_seniority", r,
                          f"paid leave for {int(pl)} days{why}", leave, leave_due)
            if sd:
                employer_days = min(sd, M.SICK_DAYS_EMPLOYER)
                # Measured against what the contract says the month should have
                # accrued, not against the row's own figures: a defect injected into
                # the supplement or the leave would otherwise move the base the sick
                # pay is compared with, and one defect would be reported twice.
                # bonus_in_base is read from the manifest rather than inferred from
                # the file: it is the auditor's configured reading of чл. 17, ал. 1
                # (the plugin's install-time question), not a property of the payroll.
                permanent_due = M.permanent_work_pay(
                    base_due, r2(base_due * pct / 100.0),
                    c["bonus"] if bonus_in_base else 0.0)
                due = r2(M.sick_daily_base(c["monthly_salary"], pct, norm,
                                           permanent_due, wd)
                         * employer_days * M.SICK_RATE)
                if abs(sick_pay - due) > M.TOL:
                    other = r2(M.sick_daily_base(
                        c["monthly_salary"], pct, norm,
                        r2(permanent_due + (-c["bonus"] if bonus_in_base
                                            else c["bonus"])), wd)
                        * employer_days * M.SICK_RATE) if wd else 0.0
                    why = (" - on a base carrying the month's bonus, which is in none "
                           "of the seven points of чл. 17, ал. 1 НСОРЗ"
                           if not bonus_in_base else
                           " - on a base without the month's bonus, which this file "
                           "pays under a wage system (чл. 17, ал. 1, т. 2 НСОРЗ)") \
                        if bonus and abs(sick_pay - other) <= M.TOL else ""
                    F.add("F9_sick_pay_amount", r,
                          f"sick pay under чл. 40, ал. 5 КСО for {int(employer_days)} "
                          f"days{why}", sick_pay, due)
            if sd + md:
                due = r2(min_insurable_self * M.HEALTH_ON_INCAPACITY / 100.0
                         * (sd + md) / norm)
                if abs(er_health_sick - due) > M.TOL:
                    F.add("F9_missing_health_on_sick", r,
                          f"health contribution for {int(sd + md)} days of "
                          f"incapacity or maternity (чл. 40, ал. 1, т. 5 ЗЗО)",
                          er_health_sick, due)

        # --- C2: the base of the seniority supplement -------------------
        if pct:
            due = r2(base * pct / 100.0)
            if abs(seniority - due) > M.TOL:
                wider = r2((base + leave + bonus) * pct / 100.0)
                why = " - computed on a wider base than the contracted salary" \
                    if abs(seniority - wider) <= M.TOL else ""
                F.add("C2_seniority_on_gross", r,
                      f"the {pct}% supplement does not match the base salary{why}",
                      seniority, due)

        # --- work base and "which element is inside" ---------------------
        # The composition of the insurable income is solved in two passes: first
        # the file's practice is inferred from the rows that do not sit at a cap,
        # then every row is measured against it. Enumerating subsets alone cannot
        # tell capped rows apart - there the same figure follows from many
        # different combinations.
        work_base = r2(base + seniority + bonus + leave)
        elements = dict(in_kind=in_kind, excess=excess, sick_pay=sick_pay,
                        comp_224=comp_224)
        # The composition is solvable exactly when the row does NOT sit at a cap:
        # then the insurable income is the sum itself and the gap against the work
        # base points at the element.
        at_cap = any(abs(insurable - c) <= M.TOL for c in (max_insurable, other_cap))
        # A person with no accruals for work (a full month of maternity or unpaid
        # leave): the benefits alone do not create insurable income, because there
        # is no income from labour activity to attach them to (чл. 6, ал. 2 КСО).
        # For such a row the composition is not analysed.
        no_work = work_base <= 0
        inside_unique = None
        if not at_cap and not no_work:
            # The sick pay is part of the expectation, not one of the subsets: it is
            # inside the insurable income by чл. 3, ал. 1 НЕВДПОВ. Leaving it out
            # here drops every row carrying sick pay out of the sample, the practice
            # stops being establishable, and F10 findings turn into F1.
            matches = [mask for mask, s in _subsets([("in_kind", in_kind),
                                                     ("excess", excess)])
                       if abs(r2(work_base + sick_pay + s) - insurable) <= M.TOL]
            if len(matches) == 1:
                inside_unique = matches[0]

        # The same solve for the taxable base, separately. The two bases do not have
        # to agree: reading В of the excess (stavki.md) puts it inside the insurable
        # income and outside the taxable base, and a file applying В throughout is
        # correct. Inferring one practice and using it for both reports В as a defect.
        # There is no ceiling on the taxable base, so capped rows stay in this sample.
        inside_unique_tax = None
        if not no_work:
            t_matches = []
            for mask, subset_sum in _subsets([("in_kind", in_kind), ("excess", excess)]):
                before = r2(gross + subset_sum - sick_pay - employee_total)
                applied = M.relief_for(before, deduction, deduction_life)
                if abs(r2(before - applied) - taxable) <= M.TOL:
                    t_matches.append(mask)
            if len(t_matches) == 1:
                inside_unique_tax = t_matches[0]

        data.append(dict(row=r, work_base=work_base, elements=elements,
                         inside_unique=inside_unique, inside_unique_tax=inside_unique_tax,
                         at_cap=at_cap, no_work=no_work,
                         insurable=insurable, sick_pay=sick_pay, in_kind=in_kind,
                         excess=excess, gross=gross, employee_total=employee_total,
                         taxable=taxable, deduction=deduction,
                         deduction_life=deduction_life, er_social=er_social))

    # ------------------------- the file's practice for the two benefits
    # Whether the benefit in kind and the threshold excess enter the bases is a
    # contested question (proverki.md, F10). No answer is assumed: it is inferred
    # from the file itself - but only if the sample allows it: at least three
    # usable rows and a clear majority. A usable row is one that does not sit at a
    # cap (there the composition is indistinguishable) and has accruals for work.
    # If no practice can be established, no conclusion is drawn for that element -
    # exactly as a live auditor would ask instead of guessing.
    def practice_for(el, key):
        sample = [el in d[key] for d in data
                  if d[key] is not None and d["elements"][el]]
        if len(sample) < 3:
            return None, len(sample)
        counter = Counter(sample)
        value, count = counter.most_common(1)[0]
        if count / len(sample) < 2 / 3:
            return None, len(sample)
        return value, len(sample)

    # One practice per element PER BASE - see the note in pass 1.
    practice, sample_size = {}, {}
    practice_tax, sample_size_tax = {}, {}
    for el in ("in_kind", "excess"):
        practice[el], sample_size[el] = practice_for(el, "inside_unique")
        practice_tax[el], sample_size_tax[el] = practice_for(el, "inside_unique_tax")
        if not any(d["elements"][el] for d in data):
            continue
        for label, value, size in (("осигурителния доход", practice[el], sample_size[el]),
                                   ("данъчната основа", practice_tax[el],
                                    sample_size_tax[el])):
            if value is None:
                F.add("F10_practice_not_establishable", "file",
                      f"the file's practice for {el} in {label} cannot be inferred "
                      f"from the file itself ({size} usable rows) - the composition "
                      f"for those people does not support a conclusion without the "
                      f"internal rules")

    # ================== pass 2: composition, symmetry, file-level findings
    NAMES = dict(in_kind="the benefit in kind", excess="the threshold excess",
                 sick_pay="the чл. 40, ал. 5 КСО sick pay",
                 comp_224="the чл. 224 КТ compensation")
    ID_FOR = dict(sick_pay="F9_sick_pay_out_of_insurable",
                  comp_224="F1_compensation_in_insurable",
                  in_kind="F10_in_kind_asymmetry", excess="F10_excess_asymmetry")

    for d in data:
        r = d["row"]
        el = d["elements"]
        if d["no_work"]:
            continue                   # no income from labour activity, no composition
        # An unestablishable practice only stops conclusions about the composition
        # of the insurable income. The checks that do not depend on it - the cap,
        # the taxable base, the relief limit - carry on.
        # Gates the INSURABLE side only. An element whose place in the taxable base
        # cannot be inferred says nothing about the insurable one, and folding the two
        # together silenced every insurable check on a file where only the taxable
        # sample was too small - seed 165 lost four findings that way, three of which
        # had nothing to do with the unknown. The taxable side has no gate at all any
        # more: it enumerates the placements instead. See below.
        practice_clear = not any(el[k] and practice[k] is None
                                 for k in ("in_kind", "excess"))
        allowed = {k for k in ("in_kind", "excess") if practice[k] and el[k]}
        allowed_sum = r2(sum(el[k] for k in allowed))
        # What the file's practice puts in the TAXABLE base, which need not be the
        # same set - reading В of the excess is in one base only.
        allowed_tax = {k for k in ("in_kind", "excess") if practice_tax[k] and el[k]}
        allowed_tax_sum = r2(sum(el[k] for k in allowed_tax))
        # The чл. 40, ал. 5 КСО sick pay belongs to the expectation, not to the
        # candidate deviations from it (чл. 3, ал. 1 НЕВДПОВ). It can therefore only
        # be found missing, never found added.
        inside_expected = allowed | ({"sick_pay"} if el["sick_pay"] else set())
        expected_insurable = r2(d["work_base"] + el["sick_pay"] + allowed_sum)

        if d["at_cap"]:
            # only whether the cap is the right one for the period
            if abs(d["insurable"] - other_cap) <= M.TOL and other_cap < max_insurable \
                    and expected_insurable > other_cap + M.TOL:
                F.add("B4_cap_from_wrong_period", "file",
                      f"the insurable income is capped at {other_cap:.2f} - the cap "
                      f"of the other half-year - instead of the applicable "
                      f"{max_insurable:.2f}", other_cap, max_insurable)
        elif practice_clear and abs(d["insurable"] - expected_insurable) > M.TOL:
            added = [k for k, value in el.items() if value and k not in inside_expected
                     and abs(r2(expected_insurable + value) - d["insurable"]) <= M.TOL]
            removed = [k for k in inside_expected
                       if abs(r2(expected_insurable - el[k]) - d["insurable"]) <= M.TOL]
            if len(added) == 1:
                k = added[0]
                F.add(ID_FOR[k], r,
                      f"{NAMES[k]} ({el[k]:.2f}) is inside the insurable income while "
                      f"the other rows leave it out", d["insurable"], expected_insurable)
            elif len(removed) == 1:
                k = removed[0]
                F.add(ID_FOR[k], r,
                      f"{NAMES[k]} ({el[k]:.2f}) is outside the insurable income while "
                      f"the other rows include it", d["insurable"], expected_insurable)
            else:
                candidates = ", ".join(f"{NAMES[k]} {el[k]:.2f}"
                                       for k in (added or removed)) or "none"
                F.add("F1_insurable_unexplained", r,
                      f"the insurable income {d['insurable']:.2f} does not match the "
                      f"work base {d['work_base']:.2f} plus what the file's practice "
                      f"allows ({expected_insurable:.2f}); candidates for the "
                      f"difference: {candidates}", d["insurable"], expected_insurable)

        elif el["sick_pay"]:
            # The practice could not be inferred, so the composition as a whole is
            # not decidable - but the sick pay's place in it is not the contested
            # part. чл. 3, ал. 1 НЕВДПОВ names it, so ask a question that does not
            # need the practice: does ANY combination of the contested elements
            # reach the declared figure with the sick pay inside, and does one reach
            # it without? If only the second holds, the sick pay is out whatever the
            # practice turns out to be.
            #
            # Without this the check went silent on every row of a file whose
            # practice was unestablishable, and an injected defect went unfound for
            # 3000 seeds while the suite stayed green at 300.
            sums = [s for _, s in _subsets([("in_kind", el["in_kind"]),
                                            ("excess", el["excess"])])]
            with_sick = any(abs(r2(d["work_base"] + el["sick_pay"] + s)
                                - d["insurable"]) <= M.TOL for s in sums)
            without_sick = any(abs(r2(d["work_base"] + s) - d["insurable"]) <= M.TOL
                               for s in sums)
            if without_sick and not with_sick:
                F.add("F9_sick_pay_out_of_insurable", r,
                      f"{NAMES['sick_pay']} ({el['sick_pay']:.2f}) is outside the "
                      f"insurable income - no combination of the contested elements "
                      f"reaches {d['insurable']:.2f} with it inside, and one reaches "
                      f"it without (чл. 3, ал. 1 НЕВДПОВ)",
                      d["insurable"], r2(d["insurable"] + el["sick_pay"]))

        if d["insurable"] > max_insurable + M.TOL:
            F.add("B4_cap_from_wrong_period", "file",
                  f"insurable income {d['insurable']:.2f} above the maximum "
                  f"{max_insurable:.2f}", d["insurable"], max_insurable)

        # ---------------- composition of the taxable base ----------------
        # An element whose place in THIS base cannot be inferred from the file used to
        # end the row here. That went too far. Of the deviations below, the sick pay's
        # place is settled by чл. 24, ал. 2, т. 14 ЗДДФЛ and both relief scenarios are
        # about what was deducted rather than what the base contains - none of the
        # three needs the practice at all. So instead of going silent, enumerate the
        # placements the unknown elements could have and let the arithmetic choose:
        # the composition that explains the row with ONE known deviation is the
        # composition the file used. Seed 165 lost three findings to the old bail-out.
        unknown_tax = [k for k in ("in_kind", "excess")
                       if el[k] and practice_tax[k] is None]

        def resolve_taxable(inside, inside_sum):
            """The verdict for one assumed composition of the taxable base.

            (ident, text, stated, expected) for a single named deviation;
            F6_taxable_unexplained when none or several fit; None when the row already
            matches the hypothesis.
            """
            def taxable_for(delta, relief_mode):
                # The sick pay sits inside the gross and outside the taxable base
                # (чл. 24, ал. 2, т. 14 ЗДДФЛ), so it comes back out here.
                before = r2(d["gross"] + inside_sum - d["sick_pay"] + delta
                            - d["employee_total"])
                pension, life = d["deduction"], d["deduction_life"]
                if relief_mode == "limit":
                    # correct: one 10% per group of чл. 19, ал. 2
                    applied = M.relief_for(before, pension, life)
                elif relief_mode == "full":
                    applied = r2(pension + life)
                elif relief_mode == "combined":
                    # both groups squeezed under a single shared 10%
                    applied = r2(min(r2(pension + life), r2(before * M.RELIEF_LIMIT)))
                else:
                    applied = 0.0
                return r2(before - applied), applied, before

            hypothesis, _, _ = taxable_for(0.0, "limit")
            if abs(d["taxable"] - hypothesis) <= M.TOL:
                return None
            candidates = [(0.0, "full", "F7_relief_over_limit",
                           "the whole withheld amount was deducted, without the limit")]
            if d["deduction"] or d["deduction_life"]:
                # The relief was due and none of it was given. Without this candidate
                # the row falls through to F6_taxable_unexplained, which says the base
                # does not follow from the gross - true, but it does not name the
                # reason, and the reason is money the person overpaid.
                candidates.append((0.0, "none", "F7_relief_not_applied",
                                   "the withheld personal contribution reduced the "
                                   "taxable base by nothing (чл. 19, ал. 2 във вр. с "
                                   "чл. 42, ал. 3 ЗДДФЛ)"))
            if d["deduction"] and d["deduction_life"]:
                # Only distinguishable when the row carries both groups: with one
                # instrument a shared cap and a per-group cap are the same number.
                candidates.append((0.0, "combined", "F7_relief_combined_limit",
                                   "both groups of the чл. 19, ал. 2 relief were "
                                   "capped against one shared 10% - each has its own, "
                                   "against the same base"))
            if d["sick_pay"]:
                candidates.append((d["sick_pay"], "limit", "F9_sick_pay_in_taxable",
                                   "the чл. 40, ал. 5 КСО sick pay is inside the "
                                   "taxable base; it is not taxable income "
                                   "(чл. 24, ал. 2, т. 14 ЗДДФЛ)"))
            if el["comp_224"]:
                candidates.append((-el["comp_224"], "limit",
                                   "F6_compensation_out_of_taxable",
                                   "the чл. 224 КТ compensation is outside the taxable "
                                   "base"))
            for k in ("in_kind", "excess"):
                if not el[k] or k in unknown_tax:
                    # An element whose placement is being enumerated cannot also be a
                    # deviation from it: "while the other rows include it" is exactly
                    # what is not known. Offering it as a candidate turns the unknown
                    # into an asymmetry finding against a row that has none.
                    continue
                # Measured against what the OTHER ROWS do with this element in this
                # base - not against the insurable income. The two bases are allowed
                # to differ; rows are not.
                sign = -1.0 if k in inside else 1.0
                candidates.append((sign * el[k], "limit", ID_FOR[k],
                                   f"{NAMES[k]} ({el[k]:.2f}) is "
                                   + ("outside" if sign < 0 else "inside")
                                   + " the taxable base while the other rows "
                                   + ("include" if sign < 0 else "leave")
                                   + " it" + ("" if sign < 0 else " out")))
            found = []
            for delta, mode, ident, text in candidates:
                value, applied, before = taxable_for(delta, mode)
                if abs(d["taxable"] - value) <= M.TOL:
                    found.append((ident, text, applied, before))
            if len(found) == 1:
                ident, text, applied, before = found[0]
                if ident == "F7_relief_over_limit":
                    limit = r2(before * M.RELIEF_LIMIT)
                    if applied <= limit + M.TOL:
                        return None
                    return (ident,
                            f"{applied:.2f} was deducted against a "
                            f"{M.RELIEF_LIMIT:.0%} limit of {limit:.2f} "
                            f"(чл. 42, ал. 3 във вр. с чл. 19 ЗДДФЛ)", applied, limit)
                return (ident, text, d["taxable"], hypothesis)
            return ("F6_taxable_unexplained",
                    f"the taxable base {d['taxable']:.2f} does not follow from the "
                    f"gross {d['gross']:.2f} minus the employee contributions "
                    f"{d['employee_total']:.2f} (expected {hypothesis:.2f}); "
                    + (f"{len(found)} possible explanations" if found
                       else "none of the known deviations fits"),
                    d["taxable"], hypothesis)

        # With nothing unknown this is one pass over the inferred practice, exactly as
        # before. With an unknown element it is two, or four, and only a verdict the
        # arithmetic singles out survives.
        verdicts = [resolve_taxable(allowed_tax | set(mask), r2(allowed_tax_sum + extra))
                    for mask, extra in _subsets([(k, el[k]) for k in unknown_tax])]
        if any(v is None for v in verdicts):
            # An admissible composition reproduces the stated base to the cent. This
            # is tested FIRST and it outranks any named verdict from a different
            # placement: raising that one would be a finding against a row that is
            # explained, produced by the placement we happened to assume. A false
            # positive fails this suite exactly like a miss.
            continue
        named = {v[0]: v for v in verdicts if v[0] != "F6_taxable_unexplained"}
        if len(named) == 1:
            ident, text, stated, expected = next(iter(named.values()))
            F.add(ident, r, text, stated, expected)
        else:
            # Either no known deviation fits any placement, or the placements name
            # different ones and the file cannot settle which. Both are "the base does
            # not follow"; the second has no F6 verdict to borrow wording from, so it
            # says what it actually found.
            fallback = next((v for v in verdicts
                             if v[0] == "F6_taxable_unexplained"), None)
            if fallback is not None:
                F.add(fallback[0], r, fallback[1], fallback[2], fallback[3])
            else:
                F.add("F6_taxable_unexplained", r,
                      f"the taxable base {d['taxable']:.2f} is explained by a "
                      f"different single deviation under each admissible placement of "
                      f"the elements the file does not settle "
                      f"({', '.join(sorted(named))}) - which one is right cannot be "
                      f"told from the file", d["taxable"])

    # --- F5: the accident rate implied by the employer contributions -----
    implied = [r2(d["er_social"] / d["insurable"] * 100.0 - M.EMPLOYER_SOCIAL)
               for d in data if d["insurable"] > 0]
    if implied:
        rate = Counter(implied).most_common(1)[0][0]
        if rate < tzpb_due - 0.005:
            F.add("F5_tzpb_below_due", "file",
                  f"the employer contributions imply an accident rate of {rate}% "
                  f"against {tzpb_due}% due", rate, tzpb_due)

    # --- K5: the total row ------------------------------------------------
    for k in M.SUMMED_COLUMNS:
        s = r2(sum(ws.cell(row=r, column=col[k]).value or 0 for r in rows))
        stated = ws.cell(row=TOTAL, column=col[k]).value or 0
        # More than a cent and a half. One cent comes from rounding the sum itself
        # and from the order floating-point addition happens in - that is noise,
        # not a hand-typed total.
        if abs(stated - s) > 0.015:
            F.add("K5_total_not_sum", "file",
                  f"in the total row „{k}“ reads {stated} while the cells sum to {s}",
                  stated, s)

    # ------------------------------------------------------- comparison
    expected = {("file" if where == "file" else HDR + 1 + idx, ident)
                for where, idx, ident in man["expected"]}
    # Informational findings do not assert a defect; they describe how far the
    # data goes. They stay out of the comparison but are printed.
    INFORMATIONAL = {"F10_practice_not_establishable"}
    found = {k for k in F.keys() if k[1] not in INFORMATIONAL}
    missed, extra = expected - found, found - expected

    if not quiet:
        print(f"=== {os.path.basename(xlsx)} · {man['month']:02d}.{man['year']} · "
              f"regime {man['regime']} (cap {max_insurable}) · {norm} days · "
              f"{len(data)} people · accident rate {tzpb_due}% ===")
        print(f"practice: insurable {practice} · taxable {practice_tax} "
              f"| set {man['policy']}")
        print(f"\nFINDINGS ({len(F.items)}):")
        for f in sorted(F.items, key=lambda x: (str(x["where"]), x["id"])):
            where = "file" if f["where"] == "file" else f"row {f['where']}"
            print(f"  [{where:8}] {f['id']:30} {f['text']}")
            if f["due"] is not None:
                print(f"{'':13} stated {f['stated']} | due {f['due']}")
        if missed:
            print(f"\nMISSED ({len(missed)}):")
            for where, ident in sorted(missed, key=str):
                print(f"  {where} · {ident} — {M.SCENARIOS.get(ident, ('', '?'))[1]}")
        if extra:
            print(f"\nFALSE POSITIVES ({len(extra)}):")
            for where, ident in sorted(extra, key=str):
                print(f"  {where} · {ident}")
        print(f"\ninjected {len(expected)} · found {len(expected & found)} · "
              f"missed {len(missed)} · extra {len(extra)}")

    return dict(injected=len(expected), found=len(expected & found),
                missed=sorted(f"{a}:{b}" for a, b in missed),
                extra=sorted(f"{a}:{b}" for a, b in extra),
                findings=len(F.items))


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--seed", type=int, default=1)
    ap.add_argument("--month", type=int, default=None, choices=[6, 7, 8, 9, 10, 11])
    ap.add_argument("--quiet", action="store_true")
    a = ap.parse_args()

    sys.path.insert(0, HERE)
    import generate_wide as G

    xlsx, manifest_path, _ = G.generate(a.seed, a.month)
    result = check(xlsx, manifest_path, quiet=a.quiet)
    sys.exit(0 if not result["missed"] and not result["extra"] else 1)
