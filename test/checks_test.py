# -*- coding: utf-8 -*-
"""Suite 1: the checks of references/proverki.md against the May 2026 payroll.

Rates: references/stavki.md, regime 01.01-31.07.2026.

This suite tests knowledge of the rates and of the working-time regimes -
minimum wage, the length-of-service supplement, overtime, night work, public
holidays, employer-paid sick days, the cap, vertical arithmetic, an attachment.
The answer key is in expected_findings.md; suite 2 covers the construction of
the file instead.

Column names and statutory citations stay Bulgarian, because both are quoted
from the domain. The code is English.
"""
import os
import re

import openpyxl

import trz_model as M
from trz_model import r2

# --- rates -------------------------------------------------------------------
# Read from trz_model.py, which rates_test.py cross-checks against stavki.md line by
# line. This file used to keep its own transcription of the same figures. Nothing
# guarded it: rates_test reads only the model, and the fixture is static with a static
# answer key, so a rate that moved in the reference file left these stale and suite 1
# green. That is the failure PR #16 documented - generator and checker moving together -
# in a second place.
#
# The regime is 01.01-31.07.2026, because the fixture is May 2026. Percentages are
# converted to fractions here; the model states them the way stavki.md does.
_H1 = M.REGIMES["H1"]
MIN_WAGE = _H1["min_wage"]
MIN_WAGE_HOUR = _H1["min_wage_hour"]
MAX_INSURABLE = _H1["max_insurable"]
MIN_INSURABLE_ACTIVITY = None   # Приложение №1 - MISSING from the reference
SENIORITY_MIN = M.SENIORITY_RATE / 100.0
# чл. 8 НСОРЗ owes the higher of the two: 0.15% of the minimum wage, or the floor.
# For 2026 the percentage wins - 0.9303 against a floor of 0.51 - but the max() is the
# rule, not the arithmetic of one year.
NIGHT_HOUR = round(max(M.NIGHT_FACTOR * MIN_WAGE, M.NIGHT_FLOOR), 4)
OVERTIME_WORKDAY = M.OVERTIME_WORKDAY
HOLIDAY_MULTIPLIER = M.HOLIDAY_MULTIPLIER
SICK_DAYS_EMPLOYER = M.SICK_DAYS_EMPLOYER
SICK_RATE = M.SICK_RATE
EMPLOYEE_TOTAL = M.EMPLOYEE_TOTAL / 100.0
TAX_RATE = M.TAX_RATE

# May 2026 per the calendar the model computes and run_tests' selftest pins at 18
# - a literal here went stale the day the calendar code changed.
WORK_DAYS = M.working_days(2026, 5)
NORM_HOURS, FULL_DAY = WORK_DAYS * 8, 8
TOL = 0.02

HERE = os.path.dirname(os.path.abspath(__file__))
wb = openpyxl.load_workbook(os.path.join(HERE, "vedomost_05_2026.xlsx"), data_only=True)
ws = wb.active
HDR = 5
COL = {ws.cell(row=HDR, column=c).value: c for c in range(1, ws.max_column + 1)}


def get(r, name):
    return ws.cell(row=r, column=COL[name]).value


findings = []


def report(severity, row, person, check, basis, stated, due, action):
    findings.append(dict(
        severity=severity, row=row, person=person, check=check, basis=basis,
        stated=stated, due=due,
        difference=(None if stated is None or due is None else r2(due - stated)),
        action=action))


rows = [r for r in range(HDR + 1, ws.max_row + 1) if isinstance(get(r, "№"), int)]

for r in rows:
    name = get(r, "Име")
    years = get(r, "Стаж (г.)")
    hours_per_day = get(r, "Раб. време (ч/ден)")
    days_worked = get(r, "Отраб. дни")
    overtime_hours = get(r, "Извънр. часове (раб. дни)") or 0
    holiday_hours = get(r, "Часове на празник") or 0
    night_hours = get(r, "Нощни часове") or 0
    sick_days = get(r, "Дни болничен от работодател") or 0
    base_pay = get(r, "Основна заплата")
    seniority_pct = get(r, "Клас %") or 0
    seniority_sum = get(r, "Клас сума") or 0
    overtime_premium = get(r, "Доп. извънреден") or 0
    holiday_premium = get(r, "Доп. празник") or 0
    night_premium = get(r, "Доп. нощен") or 0
    sick_pay = get(r, "Болнични от работодател") or 0
    gross = get(r, "БРУТО")
    insurable = get(r, "Осиг. доход")
    employee_contributions = get(r, "Лични осигуровки")
    taxable = get(r, "Данъчна основа")
    tax = get(r, "ДДФЛ")
    deductions = get(r, "Удръжки") or 0
    net = get(r, "НЕТО")

    part_time = hours_per_day / FULL_DAY
    hours_worked = days_worked * hours_per_day
    hourly = base_pay / hours_worked if days_worked else None
    # чл. 7 НСОРЗ: the overtime increase is computed on the basic salary PLUS the
    # supplements of permanent character under the contract - the length-of-service
    # supplement is one (чл. 15, ал. 1) - unless the contract provides another base.
    # Until 2026-09-03 this file, the formula block in proverki.md and the skill's
    # prose all divided the basic salary alone, against the norm stavki.md quoted.
    hourly_perm = (base_pay + seniority_sum) / hours_worked if days_worked else None

    # --- B1 minimum wage ---
    # „Основна заплата“ holds what was accrued for the days actually worked, not the
    # contracted monthly salary, so the threshold is pro-rated the same way. Comparing
    # an accrued amount against a whole month's minimum reports a phantom violation for
    # anyone on the minimum who worked part of the month.
    due_min = r2(MIN_WAGE * part_time * days_worked / WORK_DAYS)
    if base_pay + 1e-9 < due_min - TOL:
        report("нарушение", r, name, "B1 base pay below the minimum wage",
               "чл.244 т.1 КТ; ПМС №243 от 13.11.2025, ДВ бр.98/2025",
               base_pay, due_min,
               "Top up to the minimum wage and sign an annex to the contract.")

    # --- B3 minimum insurable income by activity: not verifiable, see the summary ---

    # --- B4 maximum insurable income ---
    if insurable > MAX_INSURABLE + TOL:
        report("нарушение", r, name, "B4 insurable income above the maximum",
               "чл.9 ЗБДОО за 2026 г., ДВ бр.68 от 28.07.2026",
               insurable, MAX_INSURABLE,
               "Cap the insurable income at 2111.64 EUR and correct the "
               "contributions (Декларация обр. 1 и 6).")

    # --- C1/C2 length-of-service supplement ---
    if years and years >= 1:
        due_pct = r2(years * SENIORITY_MIN * 100)
        if seniority_pct + 1e-9 < due_pct:
            report("нарушение", r, name,
                   f"C1 supplement: {seniority_pct}% applied for {years} years of service",
                   "ПМС №147 от 29.06.2007, ДВ бр.56/2007; чл.12 ал.1 НСОРЗ",
                   r2(base_pay * seniority_pct / 100),
                   r2(base_pay * due_pct / 100),
                   f"Accrue at least {due_pct}% on the base salary.")
        else:
            expected = r2(base_pay * seniority_pct / 100)
            if abs(seniority_sum - expected) > TOL:
                report("нарушение", r, name,
                       "C2 supplement: the amount does not match the stated percentage",
                       "чл.12 ал.1 НСОРЗ", seniority_sum, expected,
                       "Recompute the supplement on the base salary.")

    # --- D4 overtime ---
    # Two tiers. Short of the increase on the basic salary alone is short under any
    # base: нарушение. Clearing that but not the чл. 7 НСОРЗ base is a finding that
    # depends on the contract's clause on the base, and this suite ships no contracts:
    # за проверка, naming what is missing.
    if overtime_hours:
        floor = r2(hourly * overtime_hours * (1 + OVERTIME_WORKDAY))
        due = r2(hourly_perm * overtime_hours * (1 + OVERTIME_WORKDAY))
        if overtime_premium + TOL < floor:
            report("нарушение", r, name,
                   f"D4 {overtime_hours} overtime hours on working days without the premium",
                   "чл.262 ал.1 т.1 КТ (+50%); чл.7 НСОРЗ (базата)", overtime_premium, due,
                   "Accrue the premium on the basic salary plus the permanent supplements.")
        elif overtime_premium + TOL < due:
            report("за проверка", r, name,
                   f"D4 {overtime_hours} overtime hours: the premium is computed on the "
                   f"basic salary alone",
                   "чл.7 НСОРЗ - the base is the basic salary plus the supplements of "
                   "permanent character unless the contract provides another base; the "
                   "contract is not supplied", overtime_premium, due,
                   "Supply the contract clause on the base; failing another base, "
                   "recompute on the basic salary plus the supplement.")

    # --- D6 night work ---
    if night_hours:
        due = r2(NIGHT_HOUR * night_hours)
        if night_premium + TOL < due:
            report("нарушение", r, name,
                   f"D6 {night_hours} night hours without the supplement",
                   "чл.8 НСОРЗ (0.15% of the minimum wage per hour)",
                   night_premium, due, "Accrue the night supplement.")

    # --- D7 work on a public holiday ---
    # Same two tiers. The base of the чл. 264 double pay has no norm like чл. 7 НСОРЗ
    # behind it and stands `за потвърждение` in stavki.md, so the second tier is capped
    # at за проверка by the status rule, not only by the missing contract.
    if holiday_hours:
        floor = r2(hourly * holiday_hours * HOLIDAY_MULTIPLIER)
        due = r2(hourly_perm * holiday_hours * HOLIDAY_MULTIPLIER)
        if holiday_premium + TOL < floor:
            report("нарушение", r, name,
                   f"D7 {holiday_hours} hours on a public holiday below the double rate",
                   "чл.264 КТ", holiday_premium, due, "Top up to the double rate.")
        elif holiday_premium + TOL < due:
            report("за проверка", r, name,
                   f"D7 {holiday_hours} hours on a public holiday doubled on the basic "
                   f"salary alone",
                   "чл.264 КТ; the composition of the base is `за потвърждение` in "
                   "references/stavki.md (чл.66 ал.1 т.7 КТ)", holiday_premium, due,
                   "Confirm the base of the double pay; on the basic salary plus the "
                   "supplement the difference is as stated.")

    # --- F9 sick days ---
    if sick_days > SICK_DAYS_EMPLOYER:
        per_day = r2(sick_pay / sick_days) if sick_days else 0
        report("нарушение", r, name,
               f"F9 the employer pays {sick_days} sick days instead of {SICK_DAYS_EMPLOYER}",
               "чл.40 ал.5 КСО, изм. ДВ бр.106/2023, в сила от 01.01.2024",
               sick_pay, r2(per_day * SICK_DAYS_EMPLOYER),
               "The first 2 working days are at the employer's expense; the rest "
               "are paid by the state fund.")

    # --- F1 composition of the insurable income ---
    # Every accrual in this file is insurable income, the чл. 40, ал. 5 КСО sick pay
    # included: чл. 3, ал. 1 НЕВДПОВ names „възнагражденията по чл. 40, ал. 5 от
    # КСО“ among the incomes the base covers, and т. 21 of Декларация обр. 1 declares
    # it inside. Nothing accrued here is on the exhaustive exclusion list of чл. 1,
    # ал. 8 НЕВДПОВ, so the base is the whole gross, capped. A stated base ABOVE the
    # ceiling is B4's finding, not this one - reporting both would count one defect
    # twice.
    expected_insurable = r2(min(gross, MAX_INSURABLE))
    if insurable <= MAX_INSURABLE + TOL and abs(insurable - expected_insurable) > TOL:
        why = (" - the sick pay for the first days is left out of it" if sick_pay
               and abs(insurable - r2(gross - sick_pay)) <= TOL else "")
        report("нарушение", r, name,
               f"F1 the insurable income is not the whole gross{why}",
               "чл.3 ал.1 НЕВДПОВ; чл.6 ал.2 КСО", insurable, expected_insurable,
               "Include every accrual in the insurable income, then recompute the "
               "contributions and Декларация обр. 1 (т. 21).")

    # --- F2 employee contributions ---
    expected = r2(min(insurable, MAX_INSURABLE) * EMPLOYEE_TOTAL)
    if abs(employee_contributions - expected) > 0.05:
        report("нарушение", r, name,
               "F2 employee contributions are not 13.78% of the insurable income "
               "(capped)", "чл.6 ал.1 и ал.3 КСО", employee_contributions, expected,
               "Recompute the contributions.")

    # --- F6 taxable base and tax ---
    # The base is the taxable income of чл. 24 ЗДДФЛ minus the personal contributions
    # (чл. 42, ал. 2 ЗДДФЛ) - not the whole gross. The чл. 40, ал. 5 КСО sick pay sits
    # inside the gross and outside the taxable income: чл. 24, ал. 2, т. 14 ЗДДФЛ
    # exempts the benefits under part one of the КСО, and the справка по чл. 73,
    # ал. 6 reports it under код 107. This is the mirror of F1 above - the same sum is
    # inside one base and outside the other, and both are correct.
    expected_base = r2(gross - sick_pay - employee_contributions)
    if abs(taxable - expected_base) > TOL:
        why = (" - the sick pay for the first days is left inside it" if sick_pay
               and abs(taxable - r2(gross - employee_contributions)) <= TOL else "")
        report("нарушение", r, name,
               f"F6 taxable base is not the taxable income minus the employee "
               f"contributions{why}",
               "чл.42 ал.2 във вр. с чл.24 ал.2 т.14 ЗДДФЛ",
               taxable, expected_base, "Correct the taxable base.")
    expected_tax = r2(taxable * TAX_RATE)
    if abs(tax - expected_tax) > TOL:
        report("нарушение", r, name, "F6 tax is not 10% of the taxable base",
               "чл.42 ал.4 ЗДДФЛ", tax, expected_tax, "Correct the tax.")

    # --- I2 the accruals add up to the gross ---
    accrual_sum = r2(base_pay + seniority_sum + overtime_premium + holiday_premium
                     + night_premium + sick_pay)
    if abs(gross - accrual_sum) > TOL:
        report("нарушение", r, name, "I2 the accruals do not add up to the gross",
               "arithmetic consistency", gross, accrual_sum,
               "Check the accruals by type.")

    # --- I1 vertical reconciliation ---
    expected_net = r2(gross - employee_contributions - tax - deductions)
    if abs(net - expected_net) > TOL:
        report("нарушение", r, name,
               "I1 net is not gross minus contributions minus tax minus deductions",
               "arithmetic consistency", net, expected_net,
               "Correct the amount paid.")

    # --- G2 protected minimum income ---
    if deductions > 0:
        report("за проверка", r, name,
               f"G2 deduction of {deductions:.2f} EUR against a net before deductions "
               f"of {r2(gross - employee_contributions - tax):.2f} EUR",
               "чл.446 ГПК - the thresholds are NOT in references/stavki.md",
               deductions, None,
               "Enter the чл.446 ГПК thresholds and the number of dependants, then "
               "recompute.")

# --- report ---
ORDER = {"нарушение": 0, "риск": 1, "за проверка": 2, "бележка": 3}
findings.sort(key=lambda f: (ORDER[f["severity"]], f["row"]))
print(f"FINDINGS: {len(findings)}\n" + "=" * 100)
for f in findings:
    print(f"[{f['severity'].upper():11}] row {f['row']:2d} · {f['person']}")
    print(f"  {f['check']}")
    print(f"  Basis: {f['basis']}")
    if f["due"] is not None:
        print(f"  Stated {f['stated']:.2f} | due {f['due']:.2f} | "
              f"difference {f['difference']:+.2f} EUR")
    print(f"  Action: {f['action']}\n")

affected = sorted({f["person"] for f in findings if f["severity"] == "нарушение"})
underpaid = sum(f["difference"] for f in findings
                if f["severity"] == "нарушение" and f["difference"]
                and f["difference"] > 0 and "above the maximum" not in f["check"])
print("=" * 100)
print(f"People with violations: {len(affected)} of {len(rows)}")
print(f"Underpaid to the workers (excluding insurance corrections): {underpaid:.2f} EUR")
print(f"People with no findings: "
      f"{sorted(set(get(r, 'Име') for r in rows) - {f['person'] for f in findings})}")
if MIN_INSURABLE_ACTIVITY is None:
    print("B3 minimum insurable income by economic activity: NOT APPLICABLE - "
          "Приложение №1 to the ЗБДОО is not in references/stavki.md")

# --- the answer key, asserted --------------------------------------------------
# expected_findings.md in prose is not a test: printing findings and exiting 0 passes
# just as happily when every one of them disappears. This is the same key, machine
# readable, and the run fails on a difference in either direction - a missed defect and
# a false positive are equally disqualifying.
#
# Nine defects were injected; two more follow from them and are documented as such in
# expected_findings.md. Row 13 is the control row and must stay clean.
EXPECTED = {
    6:  {"B1"},          # base pay below the minimum wage
    7:  {"C1"},          # no length-of-service supplement for 12 years
    8:  {"D4"},          # overtime without the premium
    9:  {"B4", "F2"},    # cap never applied - and so the contributions are wrong too
    10: {"D6"},          # night hours without the supplement
    11: {"C2", "F9"},    # 3 employer-paid sick days - and 3.6% stated, 0.00 accrued
    12: {"I1"},          # net does not reconcile
    14: {"D7"},          # public holiday at single rate
    15: {"G2"},          # attachment against an unknown protected minimum
}
CONTROL_ROW = 13
EXPECTED_SEVERITY = {"G2": "за проверка"}       # everything else: нарушение

# Every check string starts with its id by convention. Assert the convention rather
# than trusting it: a renamed check would otherwise silently drop out of the key.
def code_of(finding):
    code = finding["check"].split()[0]
    if not re.fullmatch(r"[A-K]\d+", code):
        raise SystemExit(f"check string does not begin with a check id: "
                         f"{finding['check']!r}")
    return code


found = {}
for f in findings:
    found.setdefault(f["row"], set()).add(code_of(f))

problems = []
for row in sorted(set(EXPECTED) | set(found) | {CONTROL_ROW}):
    due = EXPECTED.get(row, set())
    got = found.get(row, set())
    for code in sorted(due - got):
        problems.append(f"row {row}: MISSED {code}")
    for code in sorted(got - due):
        problems.append(f"row {row}: FALSE POSITIVE {code}")

for f in findings:
    want = EXPECTED_SEVERITY.get(code_of(f), "нарушение")
    if f["severity"] != want:
        problems.append(f"row {f['row']}: {code_of(f)} reported as "
                        f"„{f['severity']}“, expected „{want}“")

print("=" * 100)
if problems:
    for p in problems:
        print(f"  FAIL  {p}")
    print(f"FAILED: {len(problems)} difference(s) from test/expected_findings.md")
    raise SystemExit(1)
print(f"OK: {len(findings)} findings, exactly the answer key in expected_findings.md "
      f"(9 injected + 2 consequential), row {CONTROL_ROW} clean, "
      f"the attachment reported as „за проверка“")
