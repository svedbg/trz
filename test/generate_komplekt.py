#!/usr/bin/env python3
"""A complete set of documents for one month, with one link of the chain broken.

Every other fixture in this repository is a single payroll workbook. That is why I9,
I10 and the cross-document half of A9 had no test behind them: the checks reconcile
the payroll against the documents downstream of it, and those documents did not exist
anywhere under `test/`.

A комплект is the payroll plus what follows from it:

    ведомост -> Декларация обр. 1 -> Декларация обр. 6 -> внесено -> счетоводство
    (vedomost)     (deklaracia_1)      (deklaracia_6)     (plateni)  (oborotna)

plus `dogovori.csv`, which carries what the payroll cannot: the agreed salary, the
identifier and the bank account of each person.

**The chain is built forward, and a break stops the copying at exactly one link.**
Обр. 1 is compiled from the payroll, обр. 6 from обр. 1, the payments from обр. 6 and
from the nets - so a wrong figure in обр. 1 propagates into обр. 6 the way it does when
a real filing is compiled from the same wrong data, and only the transition where the
break was planted disagrees. Without that, one mutation would light up three checks and
"each break raises its own signal and nothing else" would be untestable.

What this fixture does NOT claim. Обр. 6 here has four kinds of obligation - ДОО, ЗО,
ДЗПО-УПФ and ДДФЛ - not the form's real code list, and обр. 1 carries the five fields
the chain needs, not all of its points. The suite tests whether the reconciliation
holds, not whether the CSV is a valid НАП submission.

Nothing here is personal data. The people are „Лице 1", „Лице 2", …, the identifiers
are служебни номера of the shape `СЛ-001` - never an ЕГН, not even an invented one -
and the IBANs are `BG00TEST…`, which no bank will ever issue. Every amount is computed
from `trz_model`, which mirrors `references/stavki.md`.

    python test/generate_komplekt.py --list
    python test/generate_komplekt.py --clean --out /tmp/komplekt
    python test/generate_komplekt.py I10_duplicate_payment --out /tmp/broken
"""

import argparse
import copy
import csv
import datetime
import os
import random
import sys

import openpyxl
from openpyxl.styles import Font

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

import trz_model as M                                           # noqa: E402

r2 = M.r2

YEAR, MONTH = 2026, 7
HDR = 5                       # a title block above the header, as in the wide fixture

# The four kinds of obligation обр. 6 is split into here, each with the payroll columns
# that make it up. Employee and employer parts of one fund are one obligation, which is
# how they are declared and how they are paid.
OBLIGATIONS = {
    "ДОО": ("ДОО пенсии", "ДОО ОЗМ", "ДОО безработица",
            "Вноски работодател ДОО+ТЗПБ"),
    "ЗО": ("ЗО лична", "ЗО работодател", "ЗО при болничен/майчинство"),
    "ДЗПО-УПФ": ("ДЗПО-УПФ лична", "ДЗПО-УПФ работодател"),
    "ДДФЛ": ("ДДФЛ",),
}


def working_days_from(year, month, day):
    """Working days in the month on or after `day` - the half an annex applies to."""
    off = M.days_off(year)
    d = datetime.date(year, month, day)
    n = 0
    while d.month == month:
        if d.weekday() < 5 and d not in off:
            n += 1
        d += datetime.timedelta(days=1)
    return n


def _iban(n):
    """An account that cannot collide with a real one: BG00 fails the IBAN checksum."""
    return f"BG00TEST{n:014d}"


def _person(rnd, norm, regime, tzpb, policy, index):
    """One correct row plus the master data the payroll does not carry."""
    salary = r2(rnd.uniform(1100, 2600))
    seniority = rnd.choice([0.0, 1.2, 2.4, 4.8, 7.2])
    # Two people always carry sick days, and not by luck: `b_sick_days_differ_in_d1`
    # needs a т. 16.А to disagree with, and person 3's row is the one deleted by
    # `b_person_missing_in_d1`, so a single sick person would leave that break with
    # nothing to break on the seeds where both are drawn.
    sick = 3 if index in (2, 4) else rnd.choice([0, 0, 0, 3])
    leave = rnd.choice([0, 0, 2, 5])
    worked = norm - sick - leave
    inp = dict(monthly_salary=salary, seniority_pct=seniority,
               days_worked=worked, days_leave=leave, days_sick=sick,
               days_maternity=0, bonus=0.0, compensation_224=0.0,
               card_employer=0.0, card_employee=0.0, premium=0.0,
               personal_contribution=0.0, life_premium_personal=0.0)
    row = M.clean_row(inp, regime, tzpb, policy, norm)
    return dict(index=index, name=f"Лице {index}", ident=f"СЛ-{index:03d}",
                iban=_iban(index), monthly_salary=salary, seniority_pct=seniority,
                inputs=inp, row=row)


def _add_annex(p, rnd, norm, regime, tzpb, policy, year, month):
    """Give this person a raise effective in the middle of the month, paid correctly.

    Correctly means pro rata: the old rate for the working days before the annex, the
    new one from the annex on. Because the base is the daily rate times the days, that
    is the same figure as one effective monthly salary of (old*days_before +
    new*days_after) / days_worked - which is how it is computed here, so the clean row
    goes through `trz_model` untouched like every other.

    The person is deliberately given no leave and no sick days: those bases are drawn
    from the monthly salary, and with two salaries in the month there would be two
    defensible answers - a question this fixture is not asking.
    """
    day = 16
    after = working_days_from(year, month, day)
    before = norm - after
    old = p["monthly_salary"]
    new = r2(old * rnd.choice([1.10, 1.15, 1.20]))
    p["inputs"].update(days_worked=norm, days_leave=0, days_sick=0, days_maternity=0)
    p["annex"] = dict(day=day, date=f"{day:02d}.{month:02d}.{year}", old=old, new=new,
                      days_before=before, days_after=after)
    p["inputs"]["monthly_salary"] = r2((old * before + new * after) / norm)
    p["row"] = M.clean_row(p["inputs"], regime, tzpb, policy, norm)
    # dogovori.csv shows the agreed salaries; the effective one is an artefact of the
    # arithmetic and is never written anywhere the model can read it.
    p["monthly_salary"] = old


def chain(people):
    """Обр. 1 from the payroll, обр. 6 from обр. 1, the payments from обр. 6 and the nets.

    Returned as plain lists of dicts so a mutation can reach one field of one row
    without rebuilding anything - see build().
    """
    d1 = []
    for p in people:
        row = p["row"]
        d1.append(dict(ident=p["ident"],
                       # т. 21 - the insurable income, including the чл. 40, ал. 5 pay
                       osig_dohod=r2(row["Осигурителен доход"]),
                       # т. 16.А - days in temporary incapacity
                       dni_nerabotosposobnost=int(row["Дни болничен"]),
                       ddfl=r2(row["ДДФЛ"]),
                       **{f"vnoski_{k}": r2(sum(row.get(c, 0) or 0 for c in cols))
                          for k, cols in _obligation_keys()}))
    return d1


def _obligation_keys():
    """(key usable in an identifier, columns) for each obligation."""
    return [("doo", OBLIGATIONS["ДОО"]), ("zo", OBLIGATIONS["ЗО"]),
            ("upf", OBLIGATIONS["ДЗПО-УПФ"]), ("ddfl", OBLIGATIONS["ДДФЛ"])]


_KEY_BY_NAME = {"ДОО": "doo", "ЗО": "zo", "ДЗПО-УПФ": "upf", "ДДФЛ": "ddfl"}


def d6_from_d1(d1):
    """Обр. 6 is a summary: one line per obligation, compiled from the обр. 1 rows."""
    return [dict(vid=name, dalzhimo=r2(sum(row[f"vnoski_{key}"] for row in d1)))
            for name, key in _KEY_BY_NAME.items()]


def ledger_from(d6, people):
    """The trial balance: what the books owe, by account.

    The last transition I9 names, and the one the fixture stopped short of until
    05.09.2026. It reconciles **liabilities**, not cost: the balance owed to НАП by kind
    of contribution and the balance owed to staff. The cost of labour is K7's business
    and is deliberately not recomputed here - checking it twice is how one defect gets
    reported as two.
    """
    rows = [dict(smetka=f"461 {line['vid']}", opisanie=f"задължение към НАП — {line['vid']}",
                 salddo=line["dalzhimo"]) for line in d6]
    rows.append(dict(smetka="421 Персонал", opisanie="задължение към персонала",
                     salddo=r2(sum(p["row"]["НЕТО за изплащане"] for p in people))))
    return rows


def payments_from(d6, people):
    """What left the account: one salary order per person, one НАП order per obligation."""
    out = [dict(vid="заплата", poluchatel=p["ident"], iban=p["iban"],
                suma=r2(p["row"]["НЕТО за изплащане"])) for p in people]
    out += [dict(vid=f"НАП {line['vid']}", poluchatel="НАП", iban=_iban(0),
                 suma=line["dalzhimo"]) for line in d6]
    return out


# =====================================================================
#                              the breaks
# Each takes the whole комплект and breaks exactly one link. The chain is
# already built at this point, so a mutation to обр. 1 does NOT propagate:
# that is the difference between a filing compiled from wrong data (the
# clean chain, rebuilt below) and a link that stopped following the one
# before it (these).
# =====================================================================

def b_person_missing_in_d1(k):
    del k["d1"][2]
    k["d6"] = d6_from_d1(k["d1"])
    k["ledger"] = ledger_from(k["d6"], k["people"])
    k["payments"] = payments_from(k["d6"], k["people"])
    return 2


def b_extra_person_in_d1(k):
    ghost = copy.deepcopy(k["d1"][0])
    ghost["ident"] = "СЛ-900"
    k["d1"].append(ghost)
    k["d6"] = d6_from_d1(k["d1"])
    k["ledger"] = ledger_from(k["d6"], k["people"])
    k["payments"] = payments_from(k["d6"], k["people"])
    return None            # the ghost has no row in the payroll to point at


def b_insurable_differs_in_d1(k):
    # Only т. 21 moves. The dues were computed from the payroll, so обр. 6 and the
    # payments stay right and the disagreement is exactly one transition wide.
    k["d1"][1]["osig_dohod"] = r2(k["d1"][1]["osig_dohod"] - 150.00)
    return 1


def b_sick_days_differ_in_d1(k):
    row = next(r for r in k["d1"] if r["dni_nerabotosposobnost"] > 0)
    row["dni_nerabotosposobnost"] += 2
    return next(i for i, p in enumerate(k["people"]) if p["ident"] == row["ident"])


def b_d6_not_sum_of_d1(k):
    line = next(x for x in k["d6"] if x["vid"] == "ДОО")
    line["dalzhimo"] = r2(line["dalzhimo"] - 45.60)
    k["ledger"] = ledger_from(k["d6"], k["people"])
    k["payments"] = payments_from(k["d6"], k["people"])
    return None


def b_declared_not_paid(k):
    order = next(x for x in k["payments"] if x["vid"] == "НАП ЗО")
    order["suma"] = r2(order["suma"] - 120.00)
    return None


def b_net_not_paid(k):
    order = next(x for x in k["payments"]
                 if x["vid"] == "заплата" and x["poluchatel"] == k["people"][3]["ident"])
    order["suma"] = r2(order["suma"] - 80.00)
    return 3


def b_duplicate_payment(k):
    order = next(x for x in k["payments"]
                 if x["vid"] == "заплата" and x["poluchatel"] == k["people"][4]["ident"])
    k["payments"].append(dict(order))
    return 4


def b_ledger_differs(k):
    """The books carry a liability to НАП that the declaration does not.

    Everything upstream agrees - обр. 1, обр. 6 and the payment file all reconcile - and
    only the ledger disagrees, which is the shape this defect has in a real file: either
    an accrual nobody declared, or a declaration nobody booked.
    """
    line = next(x for x in k["ledger"] if x["smetka"].startswith("461 ДЗПО"))
    line["salddo"] = r2(line["salddo"] + 73.40)
    return None


def b_iban_shared(k):
    # Two people, one account. A9 calls this master data; I10 is where it becomes
    # visible, because it is the usual way a duplicated payment stays invisible.
    k["people"][5]["iban"] = k["people"][4]["iban"]
    k["ledger"] = ledger_from(k["d6"], k["people"])
    k["payments"] = payments_from(k["d6"], k["people"])
    return None            # two people, so the finding is about the file, not a row


def b_midmonth_annex_whole_month(k):
    """The annex applied from the first of the month instead of from its date.

    Everything downstream follows the wrong base - обр. 1, обр. 6 and the payment file
    are all rebuilt from it - so the chain stays consistent with itself and the only
    document that disagrees is the contract. That is the shape this defect has in a
    real file, and the reason A6's "last effective annex" does not catch it.
    """
    p = k["people"][0]
    a = p["annex"]
    p["inputs"]["monthly_salary"] = a["new"]
    p["row"] = M.clean_row(p["inputs"], k["regime"], k["tzpb"], k["policy"], k["norm"])
    k["d1"] = chain(k["people"])
    k["d6"] = d6_from_d1(k["d1"])
    k["ledger"] = ledger_from(k["d6"], k["people"])
    k["payments"] = payments_from(k["d6"], k["people"])
    return 0


BREAKS = {
    "A10_midmonth_annex":          (b_midmonth_annex_whole_month,
                                    "a raise effective mid-month paid for the whole month"),
    "I9_person_missing_in_d1":     (b_person_missing_in_d1,
                                    "someone in the payroll has no row in обр. 1"),
    "I9_extra_person_in_d1":       (b_extra_person_in_d1,
                                    "обр. 1 carries a person the payroll does not"),
    "I9_insurable_differs_in_d1":  (b_insurable_differs_in_d1,
                                    "т. 21 of обр. 1 is not the payroll's insurable income"),
    "I9_sick_days_differ_in_d1":   (b_sick_days_differ_in_d1,
                                    "т. 16.А of обр. 1 is not the payroll's sick days"),
    "I9_d6_not_sum_of_d1":         (b_d6_not_sum_of_d1,
                                    "обр. 6 is not the sum of the обр. 1 rows"),
    "I9_declared_not_paid":        (b_declared_not_paid,
                                    "less was paid to НАП than обр. 6 declares"),
    "I9_net_not_paid":             (b_net_not_paid,
                                    "someone was paid less than their net"),
    "I10_duplicate_payment":       (b_duplicate_payment,
                                    "one net paid twice"),
    "I10_iban_shared":             (b_iban_shared,
                                    "two people, one bank account"),
    "I9_ledger_differs":           (b_ledger_differs,
                                    "the trial balance owes НАП something обр. 6 does not"),
}


# The order breaks are applied in, and the reason there is an order at all: four of
# them rebuild what follows from the link they touch (a filing compiled from wrong data
# is still internally consistent), so they must run before anything that edits обр. 6 or
# the payment file, or they would quietly undo it. Groups are mutually exclusive - two
# breaks from one group would land on the same figure and neither could be attributed.
ORDER = ["A10_midmonth_annex",
         "I9_person_missing_in_d1", "I9_extra_person_in_d1",
         "I9_insurable_differs_in_d1", "I9_sick_days_differ_in_d1",
         "I10_iban_shared", "I9_d6_not_sum_of_d1",
         "I9_declared_not_paid", "I9_net_not_paid", "I10_duplicate_payment",
         "I9_ledger_differs"]

GROUPS = [("A10_midmonth_annex",),
          ("I9_person_missing_in_d1", "I9_extra_person_in_d1"),
          ("I9_insurable_differs_in_d1", "I9_sick_days_differ_in_d1"),
          ("I10_iban_shared",),
          ("I9_d6_not_sum_of_d1",),
          ("I9_declared_not_paid",),
          ("I9_net_not_paid", "I10_duplicate_payment"),
          ("I9_ledger_differs",)]

assert set(ORDER) == set(BREAKS) and len(ORDER) == len(BREAKS)
assert [b for g in GROUPS for b in g] and set(b for g in GROUPS for b in g) == set(BREAKS)


def apply_breaks(k, ids):
    """Break the named links, in ORDER whatever order they were asked for in."""
    unknown = [i for i in ids if i not in BREAKS]
    if unknown:
        raise KeyError(", ".join(unknown))
    k.setdefault("hit", {})
    for ident in ORDER:
        if ident in ids:
            k["hit"][ident] = BREAKS[ident][0](k)
    return k


def breaks_for_seed(seed, groups=4):
    """Which links a seed breaks. One function, because two callers need the same answer.

    `eval_skill.py` builds the set from this and `--covering` scans seeds with it; when
    each had its own copy of the three lines, a change to one would have sent the eval
    to seeds that inject something else entirely and nothing would have said so.
    """
    rnd = random.Random(seed * 7919)
    return [rnd.choice(g) for g in rnd.sample(GROUPS, groups)]


def komplekt(seed=1, year=YEAR, month=MONTH):
    """The clean set: every link follows the one before it."""
    rnd = random.Random(seed)
    norm = M.working_days(year, month)
    regime = M.REGIMES[M.regime_for(year, month)]
    tzpb = 0.5
    policy = dict(bonus_in_base=False, in_kind_in_bases=False,
                  excess_in_insurable=False, excess_in_taxable=False,
                  excess_reading="А")
    people = [_person(rnd, norm, regime, tzpb, policy, i) for i in range(1, 8)]
    _add_annex(people[0], rnd, norm, regime, tzpb, policy, year, month)
    d1 = chain(people)
    d6 = d6_from_d1(d1)
    return dict(people=people, d1=d1, d6=d6, ledger=ledger_from(d6, people),
                payments=payments_from(d6, people),
                year=year, month=month, norm=norm, tzpb=tzpb, regime=regime,
                policy=policy)


def write(k, outdir):
    """Write the five documents. Returns the manifest."""
    os.makedirs(outdir, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{k['month']:02d}-{k['year']}"
    ws["A1"] = 'ВЕДОМОСТ ЗА РАБОТНИ ЗАПЛАТИ — "Измислено" ЕООД'
    ws["A2"] = (f"Месец: {k['month']:02d}.{k['year']} г.  |  Работни дни: {k['norm']}  |  "
                f"Валута: EUR  |  ЕИК: 000000000 (тестов)")
    ws["A3"] = f"Икономическа дейност: тестова  |  ТЗПБ по КИД: {k['tzpb']}%"
    ws["A1"].font = Font(bold=True, size=12)
    for i, column in enumerate(M.COLUMNS, start=1):
        ws.cell(row=HDR, column=i, value=column).font = Font(bold=True, size=8)
    for offset, p in enumerate(k["people"]):
        r = HDR + 1 + offset
        ws.cell(row=r, column=M.COL["№"], value=offset + 1)
        ws.cell(row=r, column=M.COL["Име"], value=p["name"])
        ws.cell(row=r, column=M.COL["Отдел"], value=p["ident"])
        for column in M.COLUMNS:
            if column in ("№", "Име", "Отдел"):
                continue
            v = p["row"].get(column, 0)
            ws.cell(row=r, column=M.COL[column],
                    value=(v if v else (0 if column in M.DAY_COLUMNS else None)))
    total = HDR + 1 + len(k["people"])
    ws.cell(row=total, column=M.COL["Име"], value="ОБЩО").font = Font(bold=True)
    for column in M.SUMMED_COLUMNS:
        ws.cell(row=total, column=M.COL[column],
                value=r2(sum(p["row"].get(column, 0) or 0 for p in k["people"])))
    path = os.path.join(outdir, "vedomost.xlsx")
    wb.save(path)

    _csv(os.path.join(outdir, "dogovori.csv"),
         ["идентификатор", "име", "основна заплата", "клас %", "IBAN",
          "дата на допълнително споразумение", "нова основна заплата"],
         [[p["ident"], p["name"], p["monthly_salary"], p["seniority_pct"], p["iban"],
           (p.get("annex") or {}).get("date", ""),
           (p.get("annex") or {}).get("new", "")] for p in k["people"]])
    _csv(os.path.join(outdir, "deklaracia_1.csv"),
         ["идентификатор", "т.21 осигурителен доход", "т.16.А дни неработоспособност",
          "вноски ДОО", "вноски ЗО", "вноски ДЗПО-УПФ", "ДДФЛ"],
         [[r["ident"], r["osig_dohod"], r["dni_nerabotosposobnost"], r["vnoski_doo"],
           r["vnoski_zo"], r["vnoski_upf"], r["ddfl"]] for r in k["d1"]])
    _csv(os.path.join(outdir, "deklaracia_6.csv"),
         ["вид задължение", "дължима сума"],
         [[x["vid"], x["dalzhimo"]] for x in k["d6"]])
    _csv(os.path.join(outdir, "plateni.csv"),
         ["вид", "получател", "IBAN", "сума"],
         [[x["vid"], x["poluchatel"], x["iban"], x["suma"]] for x in k["payments"]])
    _csv(os.path.join(outdir, "oborotna.csv"),
         ["сметка", "описание", "крайно салдо (кредит)"],
         [[x["smetka"], x["opisanie"], x["salddo"]] for x in k["ledger"]])

    return dict(dir=outdir, sheet=ws.title, year=k["year"], month=k["month"],
                norm=k["norm"], tzpb=k["tzpb"], people=len(k["people"]),
                header_row=HDR)


def _csv(path, header, rows):
    with open(path, "w", encoding="utf8", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)


def build(break_id, outdir, seed=1):
    """The комплект with the named links broken, written to `outdir`.

    `break_id` is one id, a list of them, or None for the clean set.
    """
    ids = [] if break_id is None else ([break_id] if isinstance(break_id, str)
                                       else list(break_id))
    k = apply_breaks(komplekt(seed), ids)
    man = write(k, outdir)
    man["break"] = break_id
    man["breaks"] = ids
    return k, man


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("break_id", nargs="?")
    ap.add_argument("--list", action="store_true")
    ap.add_argument("--clean", action="store_true")
    ap.add_argument("--seed", type=int, default=1)
    ap.add_argument("--out", default="komplekt")
    a = ap.parse_args(argv)
    if a.list or not (a.break_id or a.clean):
        for k, (_, why) in BREAKS.items():
            print(f"{k:30} {why}")
        return 0
    _, man = build(None if a.clean else a.break_id, a.out, a.seed)
    print(f"{man['break'] or 'clean'} -> {os.path.abspath(a.out)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
