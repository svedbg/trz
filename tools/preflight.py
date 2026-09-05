#!/usr/bin/env python3
"""Pre-flight check for a real payroll workbook, run before the audit.

This is not a second auditor. It is a gatekeeper: it answers whether the file can be
audited at all, and names what is missing so the answer is not discovered halfway
through. Half the checks in `references/proverki.md` come back „недостатъчни данни" for
reasons visible before the audit starts and fixable once.

Three rules shape the whole tool:

* **The original is never modified.** The workbook is evidence. It is opened, read and
  closed; the normalised extract goes to a separate file the caller names.
* **Nothing is guessed.** Period boundaries are read from `references/stavki.md`, the
  same file the skill takes its figures from, so this tool cannot drift from it. Values
  that are not in the workbook at all - КИД, квалификационна група, ТЗПБ - are reported
  as required inputs, never inferred from the numbers.
* **No personal data leaves the file.** The report is keyed by sheet, row and column
  header; the extract by sheet and cell reference. The name column is located precisely
  so that it can be left out of both.

**It describes, it does not prescribe.** The report says what is wrong with the shape
and never offers the corrected file, or a patch, or a rewritten header. A suggested fix
is an invitation to edit the evidence before the audit has seen it, and the one thing
worse than an unauditable payroll is a tidied one. Merged cells are reported with
„раздели ги в работно копие, не в оригинала" for the same reason.

The mapping describes columns, and only columns. Internal wage rules, a КТД or the
contracts are documents the audit reads as documents (C6); they are not layout, and
folding them into mapping.yaml would turn a description of a spreadsheet into a second,
unversioned copy of company policy.

The checks emit **signals** - machine-readable ids - and the Bulgarian report is
rendered from them. That split is what lets the shape suite assert a planted defect is
found exactly once and *nothing else* is raised; asserting on rendered prose would pass
on a coincidence of wording.

Code and comments are English, per the repository convention. Two things are Bulgarian
because they are data rather than prose: the column headers, matched by their real text,
and the report, read by Bulgarian payroll staff in the same words as the audit it feeds.

Usage:
    python tools/preflight.py ВЕДОМОСТ.xlsx [--mapping mapping.yaml]
                              [--kid 62] [--group 3] [--tzpb 0.4]
                              [--out report.md] [--extract extract.json]

Exit codes: 0 auditable (warnings allowed), 1 blocked, 2 could not read the file.
"""

import argparse
import datetime as dt
import json
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:                                          # pragma: no cover
    sys.exit("openpyxl is required: pip install -r test/requirements.txt")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STAVKI = os.path.join(ROOT, "skills", "trz-expert", "references", "stavki.md")

# ------------------------------------------------------------------------- signals
# Blocking signals stop the audit; the rest travel into the report's closing section.
# Ids are stable: the shape suite names them, and so will phase 3's consumers.
NO_HEADER = "NO_HEADER"
NO_PERIOD = "NO_PERIOD"
NO_TOTALS = "NO_TOTALS"
NO_FORMULAS = "NO_FORMULAS"
NO_CACHED_VALUES = "NO_CACHED_VALUES"
MERGED_IN_DATA = "MERGED_IN_DATA"
MISSING_REQUIRED = "MISSING_REQUIRED"
UNKNOWN_COLUMNS = "UNKNOWN_COLUMNS"
DUPLICATE_CONCEPT = "DUPLICATE_CONCEPT"
MID_YEAR_BOUNDARY = "MID_YEAR_BOUNDARY"
MAPPING_UNKNOWN_CONCEPT = "MAPPING_UNKNOWN_CONCEPT"
MAPPING_COLUMN_ABSENT = "MAPPING_COLUMN_ABSENT"
NO_KID = "NO_KID"
NO_TZPB = "NO_TZPB"

BLOCKING = {NO_HEADER, NO_PERIOD, MISSING_REQUIRED, DUPLICATE_CONCEPT,
            MAPPING_UNKNOWN_CONCEPT, MAPPING_COLUMN_ABSENT}

# --------------------------------------------------------------- column vocabulary
# concept -> (required, accepted header spellings). The canonical spellings are the ones
# test/trz_model.py generates, which mirror a real accounting-firm layout; the rest are
# the variants real exports use for the same quantity. test/preflight_test.py asserts
# every canonical name is still recognised here, so the two cannot drift apart silently.
CONCEPTS = {
    "име":            (True,  ["име", "имена", "трите имена", "служител", "работник",
                               "лице"]),
    # Neither „раб. дни" nor a bare „отработени": the first reads more naturally as
    # „работни дни" - the month's norm, a different quantity from days actually worked -
    # and it claimed „Извънр. часове (раб. дни)" on a real layout. The second swallows
    # „Отработени часове", which is hours, not days. Days are named in full here.
    "отработени дни": (True,  ["отраб. дни", "отработени дни", "изработени дни",
                               "отработени работни дни"]),
    "основна":        (True,  ["основна за отработеното", "основна заплата",
                               "основно възнаграждение", "основна", "заплата"]),
    "бруто":          (True,  ["бруто", "брутно", "брутно възнаграждение",
                               "общо начислено", "всичко начислено"]),
    "осиг. доход":    (True,  ["осигурителен доход", "осиг. доход", "осиг доход",
                               "доход за осигуряване"]),
    "данъчна основа": (True,  ["данъчна основа", "основа за данък",
                               "облагаема основа"]),
    "данък":          (True,  ["ддфл", "данък", "данък общ доход", "дод"]),
    "лични вноски":   (True,  ["лични вноски общо", "лични осигуровки",
                               "осигуровки лице", "лични вноски"]),
    "нето":           (True,  ["нето за изплащане", "нето преди удръжки", "нето",
                               "сума за получаване", "за получаване"]),
    "вноски раб-л":   (False, ["вноски работодател общо", "вноски работодател",
                               "осигуровки работодател"]),
    # „Клас %" and „Клас сума" are the rate and the amount - different quantities, and
    # collapsing them into one concept made a correct real layout look like it named the
    # same thing twice. Found by running this tool against test/vedomost_05_2026.xlsx.
    "клас %":         (False, ["клас %", "клас процент", "процент клас"]),
    "клас":           (False, ["клас сума", "клас прослужено време",
                               "прослужено време", "стаж"]),
    "дни отпуск":     (False, ["дни платен отпуск", "дни отпуск", "отпуск дни"]),
    "дни болничен":   (False, ["дни болничен", "болнични дни", "дни временна "
                               "неработоспособност"]),
    "болнични":       (False, ["болнични (работодател)", "болнични работодател",
                               "обезщетение чл. 40, ал. 5"]),
    "изплатено":      (False, ["изплатено", "платено", "изплатена сума"]),
}

# Concepts never written into the normalised extract. „име" is the payroll's personal
# data; the extract is keyed by row instead, and the audit asks for a name only if a
# finding actually needs one.
NEVER_EXTRACTED = {"име"}

# Which check groups stop being answerable when a concept is absent. Only groups the
# audit would otherwise run - this is the bridge into the report's closing section, and
# the wording matches „какво не е проверено" in SKILL.md on purpose.
DEPENDS = {
    "осиг. доход":    "B3, B4, F1, F2 - осигурителният доход не може да се сверява",
    "данъчна основа": "F6, I1 - данъчната основа не може да се сверява",
    "данък":          "F6, F7, I1 - данъкът не може да се сверява",
    "лични вноски":   "F2, I1 - разпределението на вноските не може да се провери",
    "нето":           "I1, K7 - вертикалната сверка не може да се затвори",
    "бруто":          "I2, K1 - хоризонталната сверка не може да се затвори",
    "клас":           "C1, C2, C3 - класът не може да се провери",
    "дни болничен":   "F9 - дните за сметка на работодателя не могат да се разделят",
    "изплатено":      "I9 - веригата до изплатеното не може да се затвори",
}

TOTALS_LABEL = re.compile(r"^\s*(общо|всичко|тотал|сума)\b", re.I)
MONTHS = {"януари": 1, "февруари": 2, "март": 3, "април": 4, "май": 5, "юни": 6,
          "юли": 7, "август": 8, "септември": 9, "октомври": 10, "ноември": 11,
          "декември": 12}


def norm(text):
    """Header text reduced to what identifies it: lowercase, no punctuation runs."""
    s = str(text or "").strip().lower().replace("ё", "е")
    s = re.sub(r"[«»\"'`]+", "", s)
    return re.sub(r"\s+", " ", s)


def regime_boundaries(path=STAVKI):
    """Period boundaries as (start, end) dates, read from the МОД table in stavki.md.

    Read rather than hardcoded: the mid-year split is the thing a copied sheet gets
    wrong, and a second copy of those dates here would be one more place to forget when
    the budget moves them. Same reason rates_test.py parses the file instead of trusting
    trz_model.py.
    """
    try:
        with open(path, encoding="utf8") as f:
            text = f.read()
    except OSError:
        return []
    seen, out = set(), []
    for a, b in re.findall(r"\|\s*(\d{2}\.\d{2}\.\d{4})\s*[–-]\s*(\d{2}\.\d{2}\.\d{4})",
                           text):
        if (a, b) in seen:
            continue
        seen.add((a, b))
        try:
            out.append((dt.datetime.strptime(a, "%d.%m.%Y").date(),
                        dt.datetime.strptime(b, "%d.%m.%Y").date()))
        except ValueError:
            continue
    return sorted(out)


# ------------------------------------------------------------------- phase 2: mapping
class Mapping:
    """A company's declared layout, so month two does not re-guess month one.

    Every company's ведомост differs and the checks look columns up by exact Bulgarian
    text. Declaring the layout once turns the unknown-column list from a monthly
    negotiation into a one-off. The file carries no personal data - column headers and
    the company's КИД, nothing per person - so it can live in version control.
    """

    def __init__(self, raw=None, path=None):
        raw = raw or {}
        self.path = path
        self.company = raw.get("company")
        self.kid = raw.get("kid")
        self.group = raw.get("group")
        self.tzpb = raw.get("tzpb")
        self.header_row = raw.get("header_row")
        self.columns = {k: str(v) for k, v in (raw.get("columns") or {}).items()}
        self.ignore = {norm(x) for x in (raw.get("ignore") or [])}
        # A typo in a concept key would silently do nothing, which is the worst
        # outcome for a file whose whole job is to remove ambiguity.
        self.unknown_concepts = sorted(k for k in self.columns if k not in CONCEPTS)
        self._by_header = {norm(v): k for k, v in self.columns.items()
                           if k in CONCEPTS}

    @classmethod
    def load(cls, path):
        import yaml                                            # already a dependency
        with open(path, encoding="utf8") as f:
            return cls(yaml.safe_load(f) or {}, path=path)

    def concept_for(self, header):
        return self._by_header.get(norm(header))

    def ignored(self, header):
        return norm(header) in self.ignore


def classify(header, mapping=None):
    """The concept a header names, or None.

    The declared mapping wins outright - it is the company saying what its own column
    is, against which a guess has no standing. Then three passes, narrowest first:
    exact spelling, substring, and every word of a spelling present in any order. The
    last pass recognises „Болнични от работодател" as „Болнични (работодател)": real
    headers insert a preposition, a bracket or a unit into an otherwise standard name.
    """
    if mapping:
        declared = mapping.concept_for(header)
        if declared:
            return declared
    h = norm(header)
    if not h:
        return None
    for concept, (_, spellings) in CONCEPTS.items():
        if h in spellings:
            return concept
    for concept, (_, spellings) in CONCEPTS.items():
        for s in spellings:
            if len(s) >= 5 and s in h:
                return concept
    words = set(re.findall(r"\w+", h))
    for concept, (_, spellings) in CONCEPTS.items():
        for s in spellings:
            # Tokens of four characters or more, and at least two of them. Three-letter
            # fragments are too generic to carry a match: „раб. дни" claimed
            # „Извънр. часове (раб. дни)" as отработени дни on a real layout, because
            # „раб" and „дни" both happened to appear in it.
            tokens = [t for t in re.findall(r"\w+", s) if len(t) >= 4]
            if len(tokens) >= 2 and set(tokens) <= words:
                return concept
    return None


def sheet_period(sheet_name, ws):
    """(year, month) for a sheet, from its name or the first few cells, else None.

    Real files label the month in the tab, in a title cell, or not at all. Nothing is
    inferred from the numbers: an unlabelled sheet is reported as unlabelled, because
    guessing the period picks the thresholds.
    """
    haystacks = [str(sheet_name)]
    for row in ws.iter_rows(min_row=1, max_row=4, max_col=12, values_only=True):
        haystacks += [str(v) for v in row if isinstance(v, str)]
    for h in haystacks:
        m = re.search(r"(\d{1,2})[-./](\d{4})", h)
        if m and 1 <= int(m.group(1)) <= 12:
            return int(m.group(2)), int(m.group(1))
        m = re.search(r"(\d{4})[-./](\d{1,2})(?!\d)", h)
        if m and 1 <= int(m.group(2)) <= 12:
            return int(m.group(1)), int(m.group(2))
        low = h.lower()
        for name, num in MONTHS.items():
            if name in low:
                y = re.search(r"(20\d{2})", h)
                if y:
                    return int(y.group(1)), num
    return None


def find_header_row(ws, mapping=None, limit=15):
    """The row that names the columns: the one matching most known concepts.

    Scored rather than assumed to be row 1, because real files carry a title, a company
    line and sometimes a blank before the headers start. A declared header_row is taken
    as given - but still scored, so a mapping that has gone stale says so instead of
    quietly reading a blank row.
    """
    if mapping and mapping.header_row:
        r = int(mapping.header_row)
        score = sum(1 for c in ws[r] if classify(c.value, mapping))
        return (r, score) if score >= 3 else (None, score)
    best, best_score = None, 0
    for r in range(1, min(limit, ws.max_row or 1) + 1):
        score = sum(1 for c in ws[r] if classify(c.value, mapping))
        if score > best_score:
            best, best_score = r, score
    return (best, best_score) if best_score >= 3 else (None, best_score)


def data_range(ws, header_row):
    """(first, last, totals): data rows run to the row before the totals row.

    The totals row is excluded because it is not a person, and every per-row check that
    treats it as one produces a finding against nobody.
    """
    first = header_row + 1
    last, totals = ws.max_row, None
    for r in range(first, (ws.max_row or first) + 1):
        for c in range(1, min(4, (ws.max_column or 1) + 1)):
            if TOTALS_LABEL.match(str(ws.cell(r, c).value or "")):
                totals = r
                break
        if totals:
            break
    if totals:
        last = totals - 1
    return first, last, totals


def analyse(path, mapping=None, kid=None, group=None, tzpb=None):
    """Read the workbook and return findings as signals. Never writes."""
    mapping = mapping or Mapping()
    kid = kid or mapping.kid
    group = group or mapping.group
    tzpb = tzpb if tzpb is not None else mapping.tzpb

    formulas = openpyxl.load_workbook(path, data_only=False)
    values = openpyxl.load_workbook(path, data_only=True)
    out = {"file": os.path.basename(path), "sheets": [],
           "boundaries": regime_boundaries(), "kid": kid, "group": group,
           "tzpb": tzpb, "signals": [], "mapping": mapping.path}

    if mapping.unknown_concepts:
        out["signals"].append((MAPPING_UNKNOWN_CONCEPT, mapping.unknown_concepts))
    if not kid or not group:
        out["signals"].append((NO_KID, None))
    if tzpb in (None, ""):
        out["signals"].append((NO_TZPB, None))

    declared_headers = {norm(v) for v in mapping.columns.values()}
    seen_headers = set()

    for name in formulas.sheetnames:
        wf, wv = formulas[name], values[name]
        header_row, score = find_header_row(wv, mapping)
        info = {"name": name, "header_row": header_row, "matched": score,
                "period": sheet_period(name, wv), "rows": 0, "totals_row": None,
                "known": {}, "unknown": [], "formula_cells": 0, "value_cells": 0,
                "merged": [], "name_col": None, "signals": [], "first_row": None,
                "uncached_cells": 0}
        out["sheets"].append(info)

        if header_row is None:
            info["signals"].append((NO_HEADER, score))
            continue

        duplicates = {}
        for c in range(1, (wv.max_column or 1) + 1):
            raw = wv.cell(header_row, c).value
            if raw is None or not str(raw).strip():
                continue
            text = str(raw).strip()
            seen_headers.add(norm(text))
            if mapping.ignored(text):
                continue
            concept = classify(text, mapping)
            if concept:
                if concept in info["known"]:
                    duplicates.setdefault(concept, [info["known"][concept]["header"]])
                    duplicates[concept].append(text)
                else:
                    info["known"][concept] = {"col": c, "header": text}
                if concept == "име":
                    info["name_col"] = c
            else:
                info["unknown"].append(text)

        if duplicates:
            info["signals"].append(
                (DUPLICATE_CONCEPT,
                 sorted(f"{k}: " + ", ".join(v) for k, v in duplicates.items())))

        first, last, totals = data_range(wv, header_row)
        info["first_row"], info["totals_row"] = first, totals
        info["rows"] = max(0, last - first + 1)
        if totals is None:
            info["signals"].append((NO_TOTALS, None))

        if info["period"] is None:
            info["signals"].append((NO_PERIOD, None))
        else:
            y, m = info["period"]
            for a, b in out["boundaries"]:
                if a.year == y and a <= dt.date(y, m, 1) <= b and (a.month, a.day) != (1, 1):
                    info["signals"].append((MID_YEAR_BOUNDARY, f"{a:%d.%m.%Y}"))
                    break

        # Formula coverage over the data block. A values-only export is not a defect in
        # itself, but it removes the evidence the K group works from, and the audit has
        # to say so rather than quietly checking less.
        for r in range(first, last + 1):
            for c in range(1, (wf.max_column or 1) + 1):
                v = wf.cell(r, c).value
                if isinstance(v, str) and v.startswith("="):
                    info["formula_cells"] += 1
                    # Excel stores the last computed result beside each formula. A file
                    # written by a script and never opened in Excel has none, and then
                    # every formula column reads as empty - the audit would see blanks
                    # where the money is and check less without noticing.
                    if wv.cell(r, c).value is None:
                        info["uncached_cells"] += 1
                elif v is not None:
                    info["value_cells"] += 1
        if info["formula_cells"] == 0:
            info["signals"].append((NO_FORMULAS, None))
        elif info["uncached_cells"]:
            info["signals"].append((NO_CACHED_VALUES,
                                    (info["uncached_cells"], info["formula_cells"])))

        info["merged"] = [str(rng) for rng in getattr(wf, "merged_cells", []).ranges
                          if rng.min_row >= header_row] if hasattr(
                              wf, "merged_cells") else []
        if info["merged"]:
            info["signals"].append((MERGED_IN_DATA, info["merged"]))

        missing = [c for c, (req, _) in CONCEPTS.items() if req and c not in info["known"]]
        if missing:
            info["signals"].append((MISSING_REQUIRED, missing))
        if info["unknown"]:
            info["signals"].append((UNKNOWN_COLUMNS, info["unknown"]))

    stale = sorted(declared_headers - seen_headers)
    if stale:
        out["signals"].append((MAPPING_COLUMN_ABSENT, stale))
    return out


def all_signals(data):
    """Every signal id raised, file-level and per sheet."""
    ids = {s for s, _ in data["signals"]}
    for s in data["sheets"]:
        ids |= {sig for sig, _ in s["signals"]}
    return ids


def blocked(data):
    return sorted(all_signals(data) & BLOCKING)


# ------------------------------------------------------------------ phase 3: extract
def extract(data, path):
    """Write the normalised sidecar the audit consumes. The workbook is not touched.

    Every value carries its cell reference, so a finding keeps a pointer back into the
    evidence instead of a number with no provenance. Names are never written: rows are
    identified by sheet and row number, and the audit asks for a name only when a
    finding actually needs one.

    Values come from the cached results Excel stores next to each formula. A workbook
    written by a script and never opened in Excel has formulas but no cached values, and
    those cells are absent from the extract rather than wrong - so the count is carried
    in `uncached_cells` and raised as NO_CACHED_VALUES by analyse(), instead of leaving
    the reader to wonder why a column is thin.
    """
    book = openpyxl.load_workbook(data["_path"], data_only=True)
    doc = {"file": data["file"], "generated": dt.datetime.now().isoformat(timespec="seconds"),
           "kid": data["kid"], "group": data["group"], "tzpb": data["tzpb"],
           "note": "Без имена: редовете се идентифицират по лист и номер на ред.",
           "uncached_cells": sum(s.get("uncached_cells", 0) for s in data["sheets"]),
           "sheets": []}
    for s in data["sheets"]:
        if s["header_row"] is None:
            continue
        ws = book[s["name"]]
        cols = {k: v for k, v in s["known"].items() if k not in NEVER_EXTRACTED}
        sheet = {"name": s["name"], "period": list(s["period"]) if s["period"] else None,
                 "header_row": s["header_row"], "totals_row": s["totals_row"],
                 "columns": {k: {"col": v["col"], "header": v["header"],
                                 "letter": get_column_letter(v["col"])}
                             for k, v in cols.items()},
                 "rows": []}
        last = (s["totals_row"] - 1) if s["totals_row"] else ws.max_row
        for r in range(s["first_row"], last + 1):
            cells = {}
            for concept, meta in cols.items():
                v = ws.cell(r, meta["col"]).value
                if v is None:
                    continue
                cells[concept] = {"ref": f"{get_column_letter(meta['col'])}{r}",
                                  "value": v if isinstance(v, (int, float)) else str(v)}
            if cells:
                sheet["rows"].append({"row": r, "cells": cells})
        doc["sheets"].append(sheet)
    with open(path, "w", encoding="utf8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=1)
    return doc


# ------------------------------------------------------------------------- rendering
def report(data):
    """The pre-flight report, in Bulgarian, rendered from the signals."""
    L = []
    sig = dict(data["signals"])
    L.append(f"# Предварителна проверка — `{data['file']}`\n")
    L.append("Проверката е само за четене: файлът не е променян.\n")
    if data.get("mapping"):
        L.append(f"Приложен опис на колоните: `{os.path.basename(data['mapping'])}`\n")

    if MAPPING_UNKNOWN_CONCEPT in sig:
        L.append(f"\n> **Описът съдържа непознати понятия**: "
                 f"{', '.join(sig[MAPPING_UNKNOWN_CONCEPT])}. Правописна грешка в ключ "
                 f"не прави нищо тихо — затова спира. Допустимите понятия са: "
                 f"{', '.join(CONCEPTS)}.\n")
    if MAPPING_COLUMN_ABSENT in sig:
        L.append(f"\n> **Описът сочи колони, които ги няма във файла**: "
                 f"{', '.join(sig[MAPPING_COLUMN_ABSENT])}. Или ведомостта е сменила "
                 f"формата си, или описът е остарял — и двете се оправят веднъж.\n")

    for s in data["sheets"]:
        ssig = dict(s["signals"])
        L.append(f"\n## Лист „{s['name']}“\n")
        if NO_HEADER in ssig:
            L.append(f"- **Заглавният ред не е намерен** (разпознати {s['matched']} "
                     f"колони). Одитът не може да тръгне по този лист.")
            continue

        L.append(f"- заглавен ред: {s['header_row']}; редове с данни: {s['rows']}"
                 + (f"; ред с общи суми: {s['totals_row']}" if s["totals_row"]
                    else "; **ред с общи суми не е намерен** (K5 няма какво да сверява)"))

        if NO_PERIOD in ssig:
            L.append("- **периодът не е обявен на листа** — не се извежда от числата; "
                     "подай го, иначе всяка проверка срещу праг е недостатъчни данни")
        else:
            y, m = s["period"]
            L.append(f"- период: {m:02d}.{y}")
            if MID_YEAR_BOUNDARY in ssig:
                L.append(f"  - режимът за периода започва на {ssig[MID_YEAR_BOUNDARY]} — "
                         f"праговете се сменят в средата на годината; лист, копиран от "
                         f"предходния месец, носи чужди прагове (K8)")

        total = s["formula_cells"] + s["value_cells"]
        if NO_FORMULAS in ssig:
            L.append("- **няма нито една формула** — файлът е експорт само със "
                     "стойности. Групата K (конструкция на файла) отпада почти изцяло: "
                     "обхват на сумите, твърди стойности, слепи контроли. Одитът остава "
                     "възможен, но го казва изрично.")
        else:
            share = 100.0 * s["formula_cells"] / total if total else 0.0
            L.append(f"- формули: {s['formula_cells']} от {total} клетки "
                     f"({share:.0f}%) — конструкцията може да се провери")

        if NO_CACHED_VALUES in ssig:
            bad, tot = ssig[NO_CACHED_VALUES]
            L.append(f"- **{bad} от {tot} формули са без запазен резултат** — файлът е "
                     f"писан от програма и не е отварян в Excel. Стойността зад формулата "
                     f"я няма: тези клетки се четат като празни, а не като грешни, и "
                     f"одитът ще провери по-малко, без да забележи. Отвори файла в Excel "
                     f"и го запиши, или поискай експорт със запазени стойности.")
        if DUPLICATE_CONCEPT in ssig:
            L.append(f"- **две колони означават едно и също**: "
                     f"{'; '.join(ssig[DUPLICATE_CONCEPT])}. Кое от двете чете одитът не "
                     f"се решава от инструмента — назови ги в описа (K10).")
        if MERGED_IN_DATA in ssig:
            m = ssig[MERGED_IN_DATA]
            L.append(f"- **слети клетки в обхвата на данните**: {len(m)} "
                     f"({', '.join(m[:5])}) — редовете се разместват при четене; "
                     f"раздели ги в работно копие, не в оригинала")
        if MISSING_REQUIRED in ssig:
            L.append(f"- **липсващи задължителни колони**: "
                     f"{', '.join(ssig[MISSING_REQUIRED])}")

        absent = [c for c in DEPENDS if c not in s["known"]]
        if absent:
            L.append("- проверки, които няма да могат да се направят:")
            L += [f"  - няма „{c}“ → {DEPENDS[c]}" for c in absent]
        if UNKNOWN_COLUMNS in ssig:
            u = ssig[UNKNOWN_COLUMNS]
            L.append(f"- неразпознати колони ({len(u)}) — опиши ги в описа на колоните "
                     f"или ги назови при подаването: "
                     + ", ".join(f"„{x}“" for x in u[:12])
                     + (" …" if len(u) > 12 else ""))
        if s["name_col"]:
            L.append(f"- колоната с имена е {s['name_col']} — съдържанието ѝ не се "
                     f"възпроизвежда нито в този доклад, нито в извлека")

    L.append("\n## Данни, които ги няма във файла\n")
    for label, value, why in (
            ("КИД на дружеството", data["kid"],
             "без него B3 (МОД) е недостатъчни данни — не се сравнява с чужд бранш"),
            ("квалификационна група", data["group"],
             "МОД се чете по група; редът от приложението към ЗБДОО се подава заедно с КИД"),
            ("ТЗПБ %", data["tzpb"],
             "без него F5 е недостатъчни данни; изведеният от вноските процент се сверява "
             "с приложението към ЗБДОО")):
        L.append(f"- **{label}**: " + (f"`{value}`" if value not in (None, "")
                                       else f"не е подаден — {why}"))

    L.append("\n## Заключение\n")
    stop = blocked(data)
    if stop:
        L.append("Одитът **не може** да тръгне, докато не се отстрани:")
        L += [f"- `{b}`" for b in stop]
    else:
        L.append("Файлът е годен за одит. Ограниченията по-горе влизат в секцията "
                 "„какво не е проверено“ на отчета.")
    return "\n".join(L) + "\n", stop


def main(argv=None):
    ap = argparse.ArgumentParser(description="Pre-flight check for a payroll workbook.")
    ap.add_argument("workbook")
    ap.add_argument("--mapping", help="mapping.yaml describing this company's layout")
    ap.add_argument("--kid", help="КИД code of the company, e.g. 62")
    ap.add_argument("--group", help="qualification group for the МОД row")
    ap.add_argument("--tzpb", help="accident-insurance percentage for the КИД")
    ap.add_argument("--out", help="write the report here instead of stdout")
    ap.add_argument("--extract", help="write the normalised sidecar here")
    a = ap.parse_args(argv)

    if not os.path.exists(a.workbook):
        print(f"няма такъв файл: {a.workbook}", file=sys.stderr)
        return 2
    mapping = None
    if a.mapping:
        try:
            mapping = Mapping.load(a.mapping)
        except Exception as exc:                              # noqa: BLE001
            print(f"описът не може да бъде прочетен: {exc}", file=sys.stderr)
            return 2
    try:
        data = analyse(a.workbook, mapping, a.kid, a.group, a.tzpb)
        data["_path"] = a.workbook
    except Exception as exc:                                  # noqa: BLE001
        print(f"файлът не може да бъде прочетен: {exc}", file=sys.stderr)
        return 2

    text, stop = report(data)
    if a.out:
        with open(a.out, "w", encoding="utf8") as f:
            f.write(text)
        print(f"докладът е записан в {a.out}")
    else:
        print(text)

    if a.extract:
        if stop:
            print("извлек не се прави, докато одитът е блокиран", file=sys.stderr)
            return 1
        extract(data, a.extract)
        print(f"извлекът е записан в {a.extract}")
    return 1 if stop else 0


if __name__ == "__main__":
    sys.exit(main())
